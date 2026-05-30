# backend/api/v1/endpoints/relatorios.py
"""
Relatorios endpoints mirroring the desktop RelatoriosPage read/report flow.
"""
from __future__ import annotations

import re
import unicodedata
import json
from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import case, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.formatters import normalize_date, safe_float, safe_int
from backend.api.v1.endpoints.programacao import get_programacao_by_codigo, upper_text
from backend.api.v1.endpoints.users import require_admin_user
from backend.config.database import get_db
from backend.models.cadastro import AjudanteDB
from backend.models.despesa import DespesaDB
from backend.models.programacao import ProgramacaoDB, ProgramacaoItemControleDB, ProgramacaoItemDB
from backend.models.recebimento import RecebimentoDB
from backend.models.user import User
from backend.models.venda_importada import VendaImportadaDB
from backend.services.audit import client_ip_from_request, record_audit_log

router = APIRouter()

REPORT_TYPES = [
    "Nota Fiscal / Transbordo",
    "Detalhe Completo da Rota",
    "Programacoes",
    "Prestacao de Contas",
    "Mortalidades",
    "Ocorrencias por Motorista",
    "Rotina Motorista/Ajudantes",
    "KM de Veiculos",
    "Abastecimentos",
    "Banhos",
    "Despesas",
]
CEDULAS = (200, 100, 50, 20, 10, 5, 2)
ACTIVE_STATUSES = {"ATIVA", "EM_ROTA", "CARREGADA", "INICIADA"}


class RelatoriosOptions(BaseModel):
    tipos: list[str]


class RelatorioProgramacaoOption(BaseModel):
    codigo_programacao: str
    motorista: str = ""
    veiculo: str = ""
    nf_numero: str = ""
    data_ref: str = ""
    status: str = ""
    prestacao_status: str = "PENDENTE"
    tipo_estimativa: str = "KG"
    operacao_tipo: str = "VENDA"
    transbordo_modalidade: str = ""
    transbordo_grupo: str = ""


class RelatorioKpi(BaseModel):
    label: str
    value: str


class RelatorioChartItem(BaseModel):
    label: str
    value: float = 0


class RelatorioColumn(BaseModel):
    key: str
    label: str
    kind: str = "text"


class RelatorioSection(BaseModel):
    title: str
    columns: list[RelatorioColumn]
    rows: list[dict[str, Any]]


class RelatorioResumoResponse(BaseModel):
    tipo: str
    programacao: str = ""
    status: str = ""
    kpis: list[RelatorioKpi]
    chart: list[RelatorioChartItem]
    columns: list[RelatorioColumn]
    rows: list[dict[str, Any]]
    text: str
    sections: list[RelatorioSection] = Field(default_factory=list)


def normalize_key(value: Any) -> str:
    text_value = upper_text(value)
    decomposed = unicodedata.normalize("NFKD", text_value)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def money(value: Any) -> float:
    return round(safe_float(value, 0.0), 2)


def number2(value: Any) -> float:
    return round(safe_float(value, 0.0), 2)


def fmt_money(value: Any) -> str:
    return f"R$ {safe_float(value, 0.0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def normalize_unit_price(value: Any) -> float:
    number = safe_float(value, 0.0)
    if abs(number) >= 100 and abs(number - round(number)) < 1e-9:
        return number / 100.0
    return number


def status_ref(programacao: ProgramacaoDB) -> str:
    status_value = upper_text(programacao.status_operacional or programacao.status)
    if not status_value and safe_int(programacao.finalizada_no_app, 0) == 1:
        return "FINALIZADA"
    return status_value or "ATIVA"


def data_ref(programacao: ProgramacaoDB) -> str:
    return str(programacao.data_criacao or programacao.data or programacao.data_saida or "")


def normalize_local_rota_display(value: Any) -> str:
    key = re.sub(r"[^A-Z0-9]", "", normalize_key(value))
    if key.startswith("SERRA"):
        return "SERRA"
    if key.startswith("SERT"):
        return "SERTAO"
    return upper_text(value)


def date_patterns(value: str | None) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    patterns = []
    normalized = normalize_date(raw)
    if normalized:
        patterns.append(normalized)
        patterns.append(f"{normalized[8:10]}/{normalized[5:7]}/{normalized[0:4]}")
    patterns.append(raw)
    return [item for item in dict.fromkeys(patterns) if item]


def matches_date(value: str, patterns: list[str]) -> bool:
    if not patterns:
        return True
    text = upper_text(value)
    return any(upper_text(pattern) in text for pattern in patterns)


def matches_programacao(
    programacao: ProgramacaoDB,
    *,
    tipo: str,
    codigo_like: str,
    motorista_like: str,
    data_like: str,
) -> bool:
    tipo_key = normalize_key(tipo)
    codigo = upper_text(programacao.codigo_programacao)
    motorista = upper_text(programacao.motorista)
    if codigo_like and codigo_like not in codigo:
        return False
    if motorista_like and motorista_like not in motorista:
        return False
    if not matches_date(data_ref(programacao), date_patterns(data_like)):
        return False
    if "PRESTACAO" in tipo_key or "FECHAMENTO" in tipo_key:
        prestacao = upper_text(programacao.prestacao_status or "PENDENTE")
        status_value = status_ref(programacao)
        return prestacao in {"PENDENTE", "FECHADA"} or status_value == "FINALIZADA"
    return True


async def resolve_equipe_nomes(db: AsyncSession, equipe_raw: str | None) -> str:
    raw = str(equipe_raw or "").strip()
    if not raw:
        return ""
    result = await db.execute(select(AjudanteDB))
    nomes = {
        str(item.id): upper_text(f"{item.nome or ''} {item.sobrenome or ''}".strip())
        for item in result.scalars().all()
    }
    out = []
    seen = set()
    for part in re.split(r"[|,;/]+", raw):
        token = part.strip()
        if not token:
            continue
        nome = nomes.get(token) if token.isdigit() else ""
        nome = nome or upper_text(token)
        if nome in seen:
            continue
        seen.add(nome)
        out.append(nome)
    return " / ".join(out)


async def programacao_rows(db: AsyncSession, limit: int = 1000) -> list[ProgramacaoDB]:
    result = await db.execute(select(ProgramacaoDB).order_by(ProgramacaoDB.id.desc()).limit(max(min(limit, 3000), 1)))
    return list(result.scalars().all())


async def itens_for(db: AsyncSession, codigo: str) -> list[ProgramacaoItemDB]:
    result = await db.execute(
        select(ProgramacaoItemDB)
        .where(func.upper(ProgramacaoItemDB.codigo_programacao) == upper_text(codigo))
        .order_by(ProgramacaoItemDB.nome_cliente.asc(), ProgramacaoItemDB.cod_cliente.asc())
    )
    return list(result.scalars().all())


async def recebimentos_for(db: AsyncSession, codigo: str) -> list[RecebimentoDB]:
    result = await db.execute(
        select(RecebimentoDB)
        .where(func.upper(RecebimentoDB.codigo_programacao) == upper_text(codigo))
        .order_by(RecebimentoDB.data_registro.desc(), RecebimentoDB.id.desc())
    )
    return list(result.scalars().all())


async def despesas_for(db: AsyncSession, codigo: str) -> list[DespesaDB]:
    result = await db.execute(
        select(DespesaDB)
        .where(func.upper(DespesaDB.codigo_programacao) == upper_text(codigo))
        .order_by(DespesaDB.data_registro.desc(), DespesaDB.id.desc())
    )
    return list(result.scalars().all())


async def controles_for(db: AsyncSession, codigo: str) -> list[ProgramacaoItemControleDB]:
    result = await db.execute(
        select(ProgramacaoItemControleDB).where(func.upper(ProgramacaoItemControleDB.codigo_programacao) == upper_text(codigo))
    )
    return list(result.scalars().all())


async def vendas_importadas_for(db: AsyncSession, codigo: str) -> list[VendaImportadaDB]:
    result = await db.execute(
        select(VendaImportadaDB)
        .where(func.upper(VendaImportadaDB.codigo_programacao) == upper_text(codigo))
        .order_by(VendaImportadaDB.id.desc())
    )
    return list(result.scalars().all())


def relatorio_status(message: str, count: int = 0) -> str:
    suffix = f" ({count})" if count else ""
    return f"STATUS: {message}{suffix}"


def col(key: str, label: str, kind: str = "text") -> RelatorioColumn:
    return RelatorioColumn(key=key, label=label, kind=kind)


def item_key(codigo: Any, pedido: Any) -> tuple[str, str]:
    return (upper_text(codigo), str(pedido or "").strip().upper())


def item_caixas(item: ProgramacaoItemDB, controle: ProgramacaoItemControleDB | None = None) -> float:
    value = controle.caixas_atual if controle and controle.caixas_atual is not None else item.caixas_atual
    if value is None:
        value = item.qnt_caixas
    return safe_float(value, 0.0)


def item_preco(item: ProgramacaoItemDB, controle: ProgramacaoItemControleDB | None = None) -> float:
    value = controle.preco_atual if controle and controle.preco_atual is not None else item.preco_atual
    if value is None:
        value = item.preco
    return normalize_unit_price(value)


def item_valor_total(item: ProgramacaoItemDB, controle: ProgramacaoItemControleDB | None = None) -> float:
    preco = item_preco(item, controle)
    caixas = item_caixas(item, controle)
    kg = safe_float((controle.peso_previsto if controle and controle.peso_previsto else item.kg), 0.0)
    if caixas > 0 and 0 < preco < 1000:
        return money(caixas * preco)
    if kg > 0 and 0 < preco < 1000:
        return money(kg * preco)
    return money(preco)


def classificar_despesa(categoria: Any, descricao: Any, tipo_despesa: Any = "") -> str:
    tipo = normalize_key(tipo_despesa)
    if tipo in {"VEICULO", "FROTA"}:
        return "VEICULO"
    text = normalize_key(f"{categoria or ''} {descricao or ''}")
    veiculo_tokens = (
        "COMBUST", "DIESEL", "GASOL", "ALCOOL", "ETANOL", "ARLA",
        "OLEO", "PNEU", "BORRACH", "MECAN", "MANUT", "OFICINA",
        "LAVAGEM", "ESTACION", "PEDAGIO", "IPVA", "MULTA",
    )
    return "VEICULO" if any(token in text for token in veiculo_tokens) else "ROTA"


def plain_number(value: Any, suffix: str = "") -> str:
    number = safe_float(value, 0.0)
    if abs(number - round(number)) < 0.000001:
        return f"{int(round(number))}{suffix}"
    return f"{number:.2f}{suffix}"


def is_transbordo_programacao(programacao: ProgramacaoDB) -> bool:
    tipo = normalize_key(getattr(programacao, "operacao_tipo", "")).replace("-", "_").replace(" ", "_")
    return tipo == "TRANSBORDO" or upper_text(getattr(programacao, "tipo_estimativa", "")) == "CX"


def nf_programacao(programacao: ProgramacaoDB) -> str:
    return upper_text(getattr(programacao, "nf_numero", "") or getattr(programacao, "num_nf", ""))


def media_carregada_programacao(programacao: ProgramacaoDB) -> float:
    for field in ("media", "nf_media_carregada", "media_1", "media_2", "media_3"):
        value = safe_float(getattr(programacao, field, 0), 0.0)
        if value > 0:
            return value / 1000.0 if value > 20 else value
    return 0.0


def carga_raiz_from_snapshot(value: Any, fallback: Any = "") -> str:
    try:
        data = json.loads(str(value or "{}"))
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}
    return upper_text(data.get("carga_raiz_programacao") or data.get("carga_origem_programacao") or fallback)


async def transferencias_rows(db: AsyncSession) -> list[dict[str, Any]]:
    try:
        result = await db.execute(
            text(
                """
                SELECT id, codigo_origem, codigo_destino, cod_cliente, pedido, qtd_caixas, qtd_convertida,
                       status, snapshot, obs, motorista_origem, motorista_destino, criado_em, atualizado_em
                  FROM transferencias
                 ORDER BY COALESCE(atualizado_em, criado_em, '') DESC, id DESC
                """
            )
        )
    except Exception:
        return []
    return [dict(row) for row in result.mappings().all()]


def transferencia_qtd_total(row: dict[str, Any]) -> int:
    return max(safe_int(row.get("qtd_caixas"), 0), safe_int(row.get("qtd_convertida"), 0), 0)


def transferencia_qtd_convertida(row: dict[str, Any]) -> int:
    qtd_convertida = max(safe_int(row.get("qtd_convertida"), 0), 0)
    if qtd_convertida > 0:
        return qtd_convertida
    status_value = normalize_key(row.get("status"))
    if status_value == "CONVERTIDA":
        return transferencia_qtd_total(row)
    return 0


def transferencia_kg_estimado(programacoes: dict[str, ProgramacaoDB], row: dict[str, Any], qtd: int | None = None) -> float:
    origem = upper_text(row.get("codigo_origem"))
    raiz = carga_raiz_from_snapshot(row.get("snapshot"), origem)
    origem_prog = programacoes.get(origem) or programacoes.get(raiz)
    if not origem_prog:
        return 0.0
    kg_base = (
        safe_float(getattr(origem_prog, "nf_kg_carregado", 0), 0.0)
        or safe_float(getattr(origem_prog, "kg_carregado", 0), 0.0)
        or safe_float(getattr(origem_prog, "nf_kg", 0), 0.0)
        or safe_float(getattr(origem_prog, "kg_nf", 0), 0.0)
    )
    cx_base = (
        safe_int(getattr(origem_prog, "nf_caixas", 0), 0)
        or safe_int(getattr(origem_prog, "total_caixas", 0), 0)
        or safe_int(getattr(origem_prog, "caixas_carregadas", 0), 0)
        or safe_int(getattr(origem_prog, "qnt_cx_carregada", 0), 0)
    )
    qtd = transferencia_qtd_convertida(row) if qtd is None else max(safe_int(qtd, 0), 0)
    if qtd <= 0 or kg_base <= 0 or cx_base <= 0:
        return 0.0
    return number2(qtd * (kg_base / cx_base))


async def codigos_relacionados_nf(db: AsyncSession, nf_like: str) -> tuple[list[str], list[dict[str, Any]]]:
    nf_norm = upper_text(nf_like)
    if not nf_norm:
        return [], []
    programacoes = await programacao_rows(db, limit=3000)
    by_code = {upper_text(item.codigo_programacao): item for item in programacoes}
    related = {
        code
        for code, prog in by_code.items()
        if nf_norm in nf_programacao(prog)
    }
    transfers = await transferencias_rows(db)
    changed = True
    while changed:
        changed = False
        for row in transfers:
            status_value = normalize_key(row.get("status"))
            if status_value in {"CANCELADA", "CANCELADO", "RECUSADA", "RECUSADO"}:
                continue
            origem = upper_text(row.get("codigo_origem"))
            destino = upper_text(row.get("codigo_destino"))
            raiz = carga_raiz_from_snapshot(row.get("snapshot"), origem)
            if origem in related or destino in related or raiz in related:
                for code in (origem, destino, raiz):
                    if code and code in by_code and code not in related:
                        related.add(code)
                        changed = True
    ordered = [
        upper_text(item.codigo_programacao)
        for item in programacoes
        if upper_text(item.codigo_programacao) in related
    ]
    return ordered, transfers


async def transferencias_programacao(db: AsyncSession, codigo: str) -> tuple[list[dict[str, Any]], float, float]:
    codigo_norm = upper_text(codigo)
    try:
        result = await db.execute(
            text(
                """
                SELECT
                    t.id,
                    t.codigo_origem,
                    t.codigo_destino,
                    t.cod_cliente,
                    t.pedido,
                    t.qtd_caixas,
                    t.qtd_convertida,
                    t.status,
                    t.snapshot,
                    t.obs,
                    t.motorista_origem,
                    t.motorista_destino,
                    t.criado_em,
                    t.atualizado_em,
                    p.nf_kg,
                    p.kg_nf,
                    p.nf_kg_carregado,
                    p.kg_carregado,
                    p.nf_caixas,
                    p.total_caixas,
                    p.caixas_carregadas,
                    p.nf_preco,
                    p.preco_nf
                  FROM transferencias t
                  LEFT JOIN programacoes p
                    ON UPPER(COALESCE(p.codigo_programacao, ''))=UPPER(COALESCE(t.codigo_origem, ''))
                 WHERE UPPER(COALESCE(t.codigo_origem, ''))=:codigo
                    OR UPPER(COALESCE(t.codigo_destino, ''))=:codigo
                 ORDER BY COALESCE(t.atualizado_em, t.criado_em, '') DESC
                """
            ),
            {"codigo": codigo_norm},
        )
    except Exception:
        return [], 0.0, 0.0
    raw_rows = list(result.mappings().all())
    roots = {
        carga_raiz_from_snapshot(row.get("snapshot"), row.get("codigo_origem"))
        for row in raw_rows
        if carga_raiz_from_snapshot(row.get("snapshot"), row.get("codigo_origem"))
    }
    root_map: dict[str, ProgramacaoDB] = {}
    if roots:
        try:
            root_result = await db.execute(select(ProgramacaoDB).where(func.upper(ProgramacaoDB.codigo_programacao).in_(roots)))
            root_map = {upper_text(item.codigo_programacao): item for item in root_result.scalars().all()}
        except Exception:
            root_map = {}
    rows: list[dict[str, Any]] = []
    saida_valor = 0.0
    entrada_valor = 0.0
    for row in raw_rows:
        status_value = normalize_key(row.get("status"))
        if status_value in {"CANCELADA", "CANCELADO", "RECUSADA", "RECUSADO"}:
            continue
        qtd = transferencia_qtd_total(row)
        qtd_convertida = transferencia_qtd_convertida(row)
        qtd_saldo = max(qtd - qtd_convertida, 0)
        raiz = carga_raiz_from_snapshot(row.get("snapshot"), row.get("codigo_origem"))
        root = root_map.get(raiz)
        kg_origem = (
            safe_float(getattr(root, "nf_kg_carregado", 0), 0.0)
            or safe_float(getattr(root, "kg_carregado", 0), 0.0)
            or safe_float(getattr(root, "nf_kg", 0), 0.0)
            or safe_float(getattr(root, "kg_nf", 0), 0.0)
            or safe_float(row.get("nf_kg_carregado"), 0.0)
            or safe_float(row.get("kg_carregado"), 0.0)
            or safe_float(row.get("nf_kg"), 0.0)
            or safe_float(row.get("kg_nf"), 0.0)
        )
        caixas_origem = (
            safe_int(getattr(root, "nf_caixas", 0), 0)
            or safe_int(getattr(root, "total_caixas", 0), 0)
            or safe_int(getattr(root, "caixas_carregadas", 0), 0)
            or safe_int(row.get("nf_caixas"), 0)
            or safe_int(row.get("total_caixas"), 0)
            or safe_int(row.get("caixas_carregadas"), 0)
        )
        preco = (
            safe_float(getattr(root, "nf_preco", 0), 0.0)
            or safe_float(getattr(root, "preco_nf", 0), 0.0)
            or safe_float(row.get("nf_preco"), 0.0)
            or safe_float(row.get("preco_nf"), 0.0)
        )
        kg_por_caixa = (kg_origem / caixas_origem) if kg_origem > 0 and caixas_origem > 0 else 0.0
        kg_total = number2(qtd * kg_por_caixa)
        kg_convertido = number2(qtd_convertida * kg_por_caixa)
        valor_compra = money(qtd_convertida * kg_por_caixa * preco) if qtd_convertida > 0 and kg_por_caixa > 0 and preco > 0 else 0.0
        origem = upper_text(row.get("codigo_origem"))
        destino = upper_text(row.get("codigo_destino"))
        direcao = "SAIDA" if origem == codigo_norm else "ENTRADA"
        if direcao == "SAIDA":
            saida_valor += valor_compra
        else:
            entrada_valor += valor_compra
        rows.append(
            {
                "direcao": direcao,
                "origem": origem,
                "destino": destino,
                "cod_cliente": upper_text(row.get("cod_cliente")),
                "pedido": upper_text(row.get("pedido")) or "-",
                "caixas": qtd,
                "caixas_convertidas": qtd_convertida,
                "caixas_saldo": qtd_saldo,
                "kg_estimado": kg_total,
                "kg_convertido_estimado": kg_convertido,
                "valor_compra": valor_compra,
                "carga_raiz_programacao": raiz,
                "status": upper_text(row.get("status")) or "-",
                "motorista_origem": upper_text(row.get("motorista_origem")) or "-",
                "motorista_destino": upper_text(row.get("motorista_destino")) or "-",
                "obs": str(row.get("obs") or "").strip() or "-",
                "atualizado_em": str(row.get("atualizado_em") or row.get("criado_em") or "")[:19],
            }
        )
    return rows, money(saida_valor), money(entrada_valor)


async def build_detalhe_completo_report(db: AsyncSession, programacao: ProgramacaoDB) -> RelatorioResumoResponse:
    codigo = upper_text(programacao.codigo_programacao)
    itens = await itens_for(db, codigo)
    controles = await controles_for(db, codigo)
    recebimentos = await recebimentos_for(db, codigo)
    despesas = await despesas_for(db, codigo)
    vendas = await vendas_importadas_for(db, codigo)
    equipe = await resolve_equipe_nomes(db, programacao.equipe)
    transferencias, transferencia_saida_compra, transferencia_entrada_compra = await transferencias_programacao(db, codigo)

    controle_map = {item_key(ctrl.cod_cliente, ctrl.pedido): ctrl for ctrl in controles}
    total_receb = money(sum(safe_float(item.valor, 0.0) for item in recebimentos))
    total_desp = money(sum(safe_float(item.valor, 0.0) for item in despesas))
    venda_importada_total = money(sum(safe_float(item.vr_total, 0.0) for item in vendas))

    total_caixas = sum(item_caixas(item, controle_map.get(item_key(item.cod_cliente, item.pedido))) for item in itens)
    total_kg_clientes = sum(
        safe_float((controle_map.get(item_key(item.cod_cliente, item.pedido)).peso_previsto if controle_map.get(item_key(item.cod_cliente, item.pedido)) else item.kg), 0.0)
        for item in itens
    )
    mort_clientes = sum(safe_int(ctrl.mortalidade_aves, 0) for ctrl in controles)
    mort_doa_aves = safe_int(programacao.mortalidade_transbordo_aves, 0)
    mort_doa_kg = safe_float(programacao.mortalidade_transbordo_kg, 0.0)

    despesas_veiculo = 0.0
    despesas_rota = 0.0
    descontos_despesas = 0.0
    despesa_rows = []
    for despesa in despesas:
        classe = classificar_despesa(despesa.categoria, despesa.descricao, despesa.tipo_despesa)
        valor = money(despesa.valor)
        if classe == "VEICULO":
            despesas_veiculo += valor
        else:
            despesas_rota += valor
        descontos_despesas += safe_float(despesa.desconto, 0.0)
        despesa_rows.append(
            {
                "classe": classe,
                "data": str(despesa.data_registro or "")[:19],
                "categoria": upper_text(despesa.categoria) or "OUTROS",
                "descricao": upper_text(despesa.descricao) or "-",
                "valor": valor,
                "desconto": money(despesa.desconto),
                "litros": number2(despesa.litros),
                "obs": despesa.observacao or "-",
            }
        )
    despesas_veiculo = money(despesas_veiculo)
    despesas_rota = money(despesas_rota)

    desconto_itens = 0.0
    valor_programado = 0.0
    venda_base_kg = 0.0
    preco_base_peso = 0.0
    cliente_rows = []
    for item in itens:
        controle = controle_map.get(item_key(item.cod_cliente, item.pedido))
        preco_original = normalize_unit_price(item.preco)
        preco_atual = item_preco(item, controle)
        preco_base_item = preco_original if preco_original > 0 else preco_atual
        caixas = item_caixas(item, controle)
        kg_ref = safe_float((controle.peso_previsto if controle and controle.peso_previsto else item.kg), 0.0)
        base_qtd = kg_ref if kg_ref > 0 and 0 < preco_original < 1000 else 1
        if preco_original > preco_atual:
            desconto_itens += (preco_original - preco_atual) * base_qtd
        valor_item = money(kg_ref * preco_atual) if kg_ref > 0 and 0 < preco_atual < 1000 else money(preco_atual)
        valor_programado += valor_item
        if kg_ref > 0 and 0 < preco_base_item < 1000:
            venda_base_kg += kg_ref
            preco_base_peso += kg_ref * preco_base_item
        cliente_rows.append(
            {
                "cod_cliente": upper_text(item.cod_cliente),
                "nome_cliente": upper_text(item.nome_cliente),
                "pedido": upper_text(item.pedido) or "-",
                "caixas": number2(caixas),
                "kg": number2(kg_ref),
                "preco": money(preco_atual),
                "valor": valor_item,
                "status": upper_text((controle.status_pedido if controle else item.status_pedido)) or "PENDENTE",
                "mortalidade": safe_int((controle.mortalidade_aves if controle else 0), 0),
                "vendedor": upper_text(item.vendedor) or "-",
                "evento": (controle.alterado_em or controle.updated_at) if controle else (item.alterado_em or ""),
            }
        )
    valor_programado = money(valor_programado)

    total_descontos = money(desconto_itens + descontos_despesas)
    peso_nf = safe_float(programacao.nf_kg or programacao.kg_nf, 0.0)
    peso_carregado = safe_float(programacao.nf_kg_carregado or programacao.kg_carregado, 0.0)
    if peso_nf <= 0:
        peso_nf = peso_carregado
    preco_compra = normalize_unit_price(programacao.nf_preco or programacao.preco_nf)
    saldo_nao_carregado_kg = safe_float(programacao.nf_saldo, 0.0)
    if saldo_nao_carregado_kg <= 0 and peso_nf > 0 and peso_carregado > 0:
        saldo_nao_carregado_kg = max(peso_nf - peso_carregado, 0.0)
    valor_compra_bruta = money(peso_nf * preco_compra)
    desconto_saldo_compra = money(saldo_nao_carregado_kg * preco_compra)
    valor_compra_liquida = money(max(valor_compra_bruta - desconto_saldo_compra, 0.0))
    valor_compra_liquida = money(max(valor_compra_liquida - transferencia_saida_compra, 0.0) + transferencia_entrada_compra)
    preco_base_venda = money(preco_base_peso / venda_base_kg) if venda_base_kg > 0 else 0.0
    kg_vendido_operacional = safe_float(programacao.nf_kg_vendido, 0.0)
    if kg_vendido_operacional <= 0:
        kg_vendido_operacional = total_kg_clientes if total_kg_clientes > 0 else peso_carregado
    venda_bruta = money(kg_vendido_operacional * preco_base_venda) if kg_vendido_operacional > 0 and preco_base_venda > 0 else (
        venda_importada_total if venda_importada_total > 0 else valor_programado
    )
    venda_liquida = money(venda_bruta - total_descontos)
    lucro_bruto = money(venda_liquida - valor_compra_liquida)
    despesas_operacionais = money(despesas_rota + despesas_veiculo)
    resultado_liquido = money(lucro_bruto - despesas_operacionais)
    margem = number2((resultado_liquido / venda_liquida * 100.0) if venda_liquida else 0.0)
    impacto_venda_mort = money((valor_programado / total_caixas * mort_clientes) if total_caixas > 0 and valor_programado > 0 else 0.0)
    impacto_compra_mort = money(mort_doa_kg * preco_compra)
    operacao_tipo = "TRANSBORDO" if is_transbordo_programacao(programacao) else "VENDA"
    if operacao_tipo == "TRANSBORDO" and venda_liquida <= 0 and not transferencia_entrada_compra:
        lucro_bruto = 0.0
        resultado_liquido = 0.0
        margem = 0.0

    resultado_rows = [
        {"grupo": "OPERACAO", "indicador": "Tipo", "valor_texto": operacao_tipo, "obs": upper_text(getattr(programacao, "transbordo_modalidade", "") or "") or "-"},
        {"grupo": "COMPRA", "indicador": "Peso da NF", "valor_texto": f"{peso_nf:.2f} KG", "obs": "Peso total da nota fiscal usado como base de compra"},
        {"grupo": "COMPRA", "indicador": "Preco de compra", "valor_texto": fmt_money(preco_compra), "obs": "Valor de compra por KG da NF"},
        {"grupo": "COMPRA", "indicador": "Compra bruta", "valor_texto": fmt_money(valor_compra_bruta), "obs": "Peso da NF x preco de compra"},
        {"grupo": "COMPRA", "indicador": "Saldo nao carregado", "valor_texto": f"{saldo_nao_carregado_kg:.2f} KG / {fmt_money(desconto_saldo_compra)}", "obs": "Abatido pelo preco de compra"},
        {"grupo": "COMPRA", "indicador": "Compra transferida saida", "valor_texto": fmt_money(transferencia_saida_compra), "obs": "Parcela da compra repassada para outros veiculos"},
        {"grupo": "COMPRA", "indicador": "Compra transferida entrada", "valor_texto": fmt_money(transferencia_entrada_compra), "obs": "Parcela de compra recebida de outro veiculo"},
        {"grupo": "COMPRA", "indicador": "Compra liquida", "valor_texto": fmt_money(valor_compra_liquida), "obs": "Compra bruta menos saldo nao carregado"},
        {"grupo": "VENDA", "indicador": "Venda programada", "valor_texto": fmt_money(valor_programado), "obs": "Clientes/itens da programacao"},
        {"grupo": "VENDA", "indicador": "KG vendido/entregue", "valor_texto": f"{kg_vendido_operacional:.2f} KG", "obs": "NF KG vendido; fallback KG dos clientes/carregado"},
        {"grupo": "VENDA", "indicador": "Preco base de venda", "valor_texto": fmt_money(preco_base_venda), "obs": "Preco medio ponderado pelos KG dos pedidos"},
        {"grupo": "VENDA", "indicador": "Venda bruta real", "valor_texto": fmt_money(venda_bruta), "obs": "KG vendido/entregue x preco base de venda"},
        {"grupo": "VENDA", "indicador": "Descontos comerciais", "valor_texto": fmt_money(total_descontos), "obs": "Descontos em itens e descontos lancados nas despesas"},
        {"grupo": "VENDA", "indicador": "Venda liquida", "valor_texto": fmt_money(venda_liquida), "obs": "Venda bruta real menos descontos"},
        {"grupo": "LUCRO", "indicador": "Lucro bruto", "valor_texto": fmt_money(lucro_bruto), "obs": "Venda liquida menos compra liquida"},
        {"grupo": "CUSTOS", "indicador": "Custos do veiculo", "valor_texto": fmt_money(despesas_veiculo), "obs": "Combustivel, manutencao, pedagio e similares"},
        {"grupo": "CUSTOS", "indicador": "Custos da rota", "valor_texto": fmt_money(despesas_rota), "obs": "Diarias e demais gastos operacionais"},
        {"grupo": "LUCRO", "indicador": "Lucro liquido", "valor_texto": fmt_money(resultado_liquido), "obs": f"Lucro bruto menos despesas. Margem: {margem:.2f}%"},
        {"grupo": "OCORRENCIAS", "indicador": "Transbordo/operacao", "valor_texto": f"{mort_doa_aves} unid. / {mort_doa_kg:.2f} KG", "obs": f"Impacto compra: {fmt_money(impacto_compra_mort)}"},
        {"grupo": "OCORRENCIAS", "indicador": "Clientes", "valor_texto": f"{mort_clientes} unid.", "obs": f"Impacto venda: {fmt_money(impacto_venda_mort)}"},
    ]

    local_carreg = upper_text(programacao.local_carregamento or programacao.granja_carregada or programacao.local_carregado or programacao.local_carreg)
    local_rota = normalize_local_rota_display(programacao.local_rota or programacao.tipo_rota)
    lines = [
        f"DETALHE COMPLETO DA ROTA - PROGRAMACAO {codigo}",
        "=" * 100,
        "",
        "[IDENTIFICACAO E OPERACAO]",
        f"Status: {status_ref(programacao)} | Prestacao: {upper_text(programacao.prestacao_status or 'PENDENTE')}",
        f"Operacao: {operacao_tipo} | Modalidade: {upper_text(getattr(programacao, 'transbordo_modalidade', '') or '') or '-'}",
        f"Data: {normalize_date(data_ref(programacao)) or data_ref(programacao) or '-'} | Usuario: {programacao.usuario_criacao or '-'}",
        f"Motorista: {programacao.motorista or '-'} | Veiculo: {programacao.veiculo or '-'} | Equipe: {equipe or '-'}",
        f"Local rota: {local_rota or '-'} | Carregamento: {local_carreg or '-'}",
        f"Saida: {str((programacao.data_saida or '') + ' ' + (programacao.hora_saida or '')).strip() or '-'} | Chegada: {str((programacao.data_chegada or '') + ' ' + (programacao.hora_chegada or '')).strip() or '-'}",
        "",
        "[CARREGAMENTO, CONSUMO E ROTA]",
        f"NF: {upper_text(programacao.nf_numero or programacao.num_nf) or '-'} | NF KG: {safe_float(programacao.nf_kg, 0.0):.2f} | NF caixas: {safe_int(programacao.nf_caixas, 0)}",
        f"KG carregado: {safe_float(programacao.nf_kg_carregado or programacao.kg_carregado, 0.0):.2f} | KG vendido/entregue: {kg_vendido_operacional:.2f} | Saldo KG: {saldo_nao_carregado_kg:.2f}",
        f"Caixas: {plain_number(total_caixas)} | KG clientes: {total_kg_clientes:.2f}",
        f"KM inicial/final: {safe_float(programacao.km_inicial, 0.0):.2f} / {safe_float(programacao.km_final, 0.0):.2f} | KM rodado: {safe_float(programacao.km_rodado, 0.0):.2f} | Litros: {safe_float(programacao.litros, 0.0):.2f} | Media: {safe_float(programacao.media_km_l, 0.0):.2f} km/l",
        "",
        "[RESULTADO FINANCEIRO]",
        f"Compra bruta: {fmt_money(valor_compra_bruta)} = NF {peso_nf:.2f} KG x {fmt_money(preco_compra)}",
        f"Saldo nao carregado: {saldo_nao_carregado_kg:.2f} KG | Abatimento na compra: {fmt_money(desconto_saldo_compra)}",
        f"Transferencias de compra: saida {fmt_money(transferencia_saida_compra)} | entrada {fmt_money(transferencia_entrada_compra)}",
        f"Compra liquida: {fmt_money(valor_compra_liquida)}",
        f"Venda bruta real: {fmt_money(venda_bruta)} = {kg_vendido_operacional:.2f} KG x {fmt_money(preco_base_venda)}",
        f"Descontos comerciais: {fmt_money(total_descontos)}",
        f"Venda liquida: {fmt_money(venda_liquida)}",
        f"Lucro bruto: {fmt_money(lucro_bruto)}",
        f"Custos do veiculo: {fmt_money(despesas_veiculo)}",
        f"Custos de rota: {fmt_money(despesas_rota)}",
        f"Lucro liquido: {fmt_money(resultado_liquido)} | Margem liquida: {margem:.2f}%",
        "",
        "[OCORRENCIAS E IMPACTOS]",
        f"Transbordo/operacao: {mort_doa_aves} unid., {mort_doa_kg:.2f} KG. Impacto compra: {fmt_money(impacto_compra_mort)}.",
        f"Clientes: {mort_clientes} unid. Impacto estimado venda: {fmt_money(impacto_venda_mort)}.",
        "",
        "[TRANSBORDO / TRANSFERENCIAS]",
    ]
    if transferencias:
        lines.append("DIR | ORIGEM -> DESTINO | CX | KG EST. | COMPRA | STATUS | PEDIDO/CLIENTE")
        lines.append("-" * 100)
        for trans in transferencias:
            lines.append(
                f"{trans['direcao']} | {trans['origem']} -> {trans['destino']} | {plain_number(trans['caixas'])} | "
                f"{safe_float(trans['kg_estimado'], 0.0):.2f} | {fmt_money(trans['valor_compra'])} | "
                f"{trans['status']} | raiz {trans.get('carga_raiz_programacao') or '-'} | {trans['pedido']} / {trans['cod_cliente'] or '-'}"
            )
    else:
        lines.append("Sem transferencias vinculadas a esta programacao.")
    lines.extend([
        "",
        "[CLIENTES / ENTREGA]",
        "COD | CLIENTE | CX | KG | VALOR | STATUS | MORT. | PEDIDO | ULTIMO EVENTO",
        "-" * 100,
    ])
    for row in cliente_rows:
        lines.append(
            f"{row['cod_cliente'] or '-'} | {row['nome_cliente'] or '-'} | {plain_number(row['caixas'])} | "
            f"{safe_float(row['kg'], 0.0):.2f} | {fmt_money(row['valor'])} | {row['status']} | "
            f"{safe_int(row['mortalidade'], 0)} | {row['pedido'] or '-'} | {row['evento'] or '-'}"
        )

    columns_main = [
        col("grupo", "GRUPO"),
        col("indicador", "INDICADOR"),
        col("valor_texto", "VALOR"),
        col("obs", "OBSERVACAO"),
    ]
    sections = [
        RelatorioSection(
            title="Transbordo e Transferencias",
            columns=[
                col("direcao", "DIR"), col("origem", "ORIGEM"), col("destino", "DESTINO"),
                col("pedido", "PEDIDO"), col("cod_cliente", "CLIENTE"), col("caixas", "CX", "number"),
                col("kg_estimado", "KG EST.", "number"), col("valor_compra", "COMPRA", "money"),
                col("carga_raiz_programacao", "RAIZ"),
                col("status", "STATUS"), col("motorista_origem", "MOT. ORIGEM"),
                col("motorista_destino", "MOT. DESTINO"), col("obs", "OBS"),
            ],
            rows=transferencias,
        ),
        RelatorioSection(
            title="Clientes e Entregas",
            columns=[
                col("cod_cliente", "COD"), col("nome_cliente", "CLIENTE"), col("pedido", "PEDIDO"),
                col("caixas", "CX", "number"), col("kg", "KG", "number"), col("preco", "PRECO", "money"),
                col("valor", "VALOR", "money"), col("status", "STATUS"), col("mortalidade", "MORT.", "number"),
                col("vendedor", "VENDEDOR"), col("evento", "EVENTO"),
            ],
            rows=cliente_rows,
        ),
        RelatorioSection(
            title="Custos Classificados",
            columns=[
                col("classe", "CLASSE"), col("data", "DATA"), col("categoria", "CATEGORIA"),
                col("descricao", "DESCRICAO"), col("valor", "VALOR", "money"),
                col("desconto", "DESCONTO", "money"), col("litros", "LITROS", "number"), col("obs", "OBS"),
            ],
            rows=despesa_rows,
        ),
        RelatorioSection(title="Resultados", columns=columns_main, rows=resultado_rows),
    ]

    return RelatorioResumoResponse(
        tipo="Detalhe Completo da Rota",
        programacao=codigo,
        status=relatorio_status("Detalhe completo da rota gerado."),
        kpis=[
            RelatorioKpi(label="Venda liquida", value=fmt_money(venda_liquida)),
            RelatorioKpi(label="Compra liquida", value=fmt_money(valor_compra_liquida)),
            RelatorioKpi(label="Lucro bruto", value=fmt_money(lucro_bruto)),
            RelatorioKpi(label="Lucro liquido", value=fmt_money(resultado_liquido)),
        ],
        chart=[
            RelatorioChartItem(label="Venda liquida", value=venda_liquida),
            RelatorioChartItem(label="Compra liquida", value=valor_compra_liquida),
            RelatorioChartItem(label="Custos", value=despesas_operacionais),
            RelatorioChartItem(label="Lucro liquido", value=resultado_liquido),
        ],
        columns=columns_main,
        rows=resultado_rows,
        text="\n".join(lines),
        sections=sections,
    )


async def build_nf_transbordo_report(db: AsyncSession, nf_like: str) -> RelatorioResumoResponse:
    nf_norm = upper_text(nf_like)
    if not nf_norm:
        raise HTTPException(status_code=422, detail="Informe o numero da nota fiscal.")
    codigos, transfers_all = await codigos_relacionados_nf(db, nf_norm)
    if not codigos:
        return RelatorioResumoResponse(
            tipo="Nota Fiscal / Transbordo",
            status=relatorio_status("Nenhuma programacao encontrada para a nota fiscal."),
            kpis=[],
            chart=[],
            columns=[col("mensagem", "RESULTADO")],
            rows=[{"mensagem": f"Nenhuma programacao vinculada a NF {nf_norm}."}],
            text=f"NOTA FISCAL / TRANSBORDO\n\nNenhuma programacao vinculada a NF {nf_norm}.",
        )

    result = await db.execute(select(ProgramacaoDB).where(func.upper(ProgramacaoDB.codigo_programacao).in_(codigos)))
    programacoes = {upper_text(item.codigo_programacao): item for item in result.scalars().all()}
    related_set = set(codigos)
    transfer_entrada_kg_by_codigo: dict[str, float] = {}
    transfer_saida_kg_by_codigo: dict[str, float] = {}
    transfer_rows_base: list[tuple[dict[str, Any], str, str, str, int, int, float, float]] = []
    for transfer in transfers_all:
        status_value = normalize_key(transfer.get("status"))
        if status_value in {"CANCELADA", "CANCELADO", "RECUSADA", "RECUSADO"}:
            continue
        origem = upper_text(transfer.get("codigo_origem"))
        destino = upper_text(transfer.get("codigo_destino"))
        raiz = carga_raiz_from_snapshot(transfer.get("snapshot"), origem)
        if origem not in related_set and destino not in related_set and raiz not in related_set:
            continue
        qtd_convertida = transferencia_qtd_convertida(transfer)
        qtd_total = transferencia_qtd_total(transfer)
        kg_transf = transferencia_kg_estimado(programacoes, transfer, qtd_total)
        kg_convertido = transferencia_kg_estimado(programacoes, transfer, qtd_convertida)
        transfer_rows_base.append((transfer, origem, destino, raiz, qtd_total, qtd_convertida, kg_transf, kg_convertido))
        if destino:
            transfer_entrada_kg_by_codigo[destino] = transfer_entrada_kg_by_codigo.get(destino, 0.0) + kg_convertido
        if origem:
            transfer_saida_kg_by_codigo[origem] = transfer_saida_kg_by_codigo.get(origem, 0.0) + kg_convertido

    rows: list[dict[str, Any]] = []
    transfer_rows: list[dict[str, Any]] = []
    fotos_rows: list[dict[str, Any]] = []
    despesa_rows: list[dict[str, Any]] = []
    total_caixas = 0.0
    total_kg_carregado = 0.0
    total_kg_entregue = 0.0
    total_mort_aves = 0
    total_mort_kg = 0.0
    total_recebido = 0.0
    total_despesas = 0.0

    for codigo in codigos:
        prog = programacoes.get(codigo)
        if not prog:
            continue
        itens = await itens_for(db, codigo)
        controles = await controles_for(db, codigo)
        recebimentos = await recebimentos_for(db, codigo)
        despesas = await despesas_for(db, codigo)
        equipe = await resolve_equipe_nomes(db, prog.equipe)
        controle_map = {item_key(ctrl.cod_cliente, ctrl.pedido): ctrl for ctrl in controles}
        caixas = sum(item_caixas(item, controle_map.get(item_key(item.cod_cliente, item.pedido))) for item in itens)
        kg_itens = sum(
            safe_float((controle_map.get(item_key(item.cod_cliente, item.pedido)).peso_previsto if controle_map.get(item_key(item.cod_cliente, item.pedido)) else item.kg), 0.0)
            for item in itens
        )
        kg_carregado = safe_float(prog.nf_kg_carregado or prog.kg_carregado or prog.nf_kg or prog.kg_nf, 0.0)
        kg_entregue = safe_float(prog.nf_kg_vendido, 0.0) or kg_itens
        kg_transferido_entrada = transfer_entrada_kg_by_codigo.get(codigo, 0.0)
        kg_transferido_saida = transfer_saida_kg_by_codigo.get(codigo, 0.0)
        if kg_entregue <= 0 and kg_transferido_entrada > 0:
            kg_entregue = kg_transferido_entrada
        mort_aves = safe_int(getattr(prog, "mortalidade_transbordo_aves", 0), 0) + sum(safe_int(ctrl.mortalidade_aves, 0) for ctrl in controles)
        mort_kg = safe_float(getattr(prog, "mortalidade_transbordo_kg", 0), 0.0)
        media_ref = media_carregada_programacao(prog)
        if mort_kg <= 0 and mort_aves and media_ref > 0:
            mort_kg = mort_aves * media_ref
        recebido = money(sum(safe_float(item.valor, 0.0) for item in recebimentos))
        despesas_total = money(sum(safe_float(item.valor, 0.0) for item in despesas))
        total_caixas += caixas
        total_kg_carregado += kg_carregado
        total_kg_entregue += kg_entregue
        total_mort_aves += mort_aves
        total_mort_kg += mort_kg
        total_recebido += recebido
        total_despesas += despesas_total
        rows.append(
            {
                "codigo_programacao": codigo,
                "nf_numero": nf_programacao(prog) or "-",
                "operacao": "TRANSBORDO" if is_transbordo_programacao(prog) else "VENDA",
                "modalidade": upper_text(getattr(prog, "transbordo_modalidade", "")) or "-",
                "motorista": upper_text(prog.motorista) or "-",
                "veiculo": upper_text(prog.veiculo) or "-",
                "ajudantes": equipe or "-",
                "status": status_ref(prog),
                "prestacao": upper_text(prog.prestacao_status or "PENDENTE"),
                "caixas": number2(caixas),
                "kg_carregado": number2(kg_carregado),
                "kg_entregue": number2(kg_entregue),
                "kg_transferido_entrada": number2(kg_transferido_entrada),
                "kg_transferido_saida": number2(kg_transferido_saida),
                "divergencia_kg": number2(kg_entregue - kg_carregado),
                "media": number2(media_ref),
                "mortalidade_aves": mort_aves,
                "mortalidade_kg": number2(mort_kg),
                "recebido": recebido,
                "despesas": despesas_total,
                "data_ref": data_ref(prog),
            }
        )
        for despesa in despesas:
            despesa_rows.append(
                {
                    "codigo_programacao": codigo,
                    "motorista": upper_text(prog.motorista) or "-",
                    "veiculo": upper_text(prog.veiculo) or "-",
                    "categoria": upper_text(despesa.categoria) or "OUTROS",
                    "descricao": upper_text(despesa.descricao) or "-",
                    "valor": money(despesa.valor),
                    "data": str(despesa.data_registro or "")[:19],
                    "obs": despesa.observacao or "-",
                }
            )
        try:
            fotos_result = await db.execute(
                text(
                    """
                    SELECT codigo_programacao, categoria, tipo_registro, cod_cliente, cliente_nome,
                           arquivo_nome, storage_path, registrado_em
                      FROM rota_fotos
                     WHERE UPPER(COALESCE(codigo_programacao, ''))=:codigo
                     ORDER BY id DESC
                    """
                ),
                {"codigo": codigo},
            )
            for foto in fotos_result.mappings().all():
                fotos_rows.append(
                    {
                        "codigo_programacao": codigo,
                        "categoria": upper_text(foto.get("categoria") or foto.get("tipo_registro")) or "-",
                        "cliente": upper_text(foto.get("cliente_nome") or foto.get("cod_cliente")) or "-",
                        "arquivo": foto.get("arquivo_nome") or foto.get("storage_path") or "-",
                        "data": str(foto.get("registrado_em") or "")[:19],
                    }
                )
        except Exception:
            pass

    for row, origem, destino, raiz, qtd, qtd_convertida, kg_transf, kg_convertido in transfer_rows_base:
        transfer_rows.append(
            {
                "origem": origem or "-",
                "destino": destino or "-",
                "raiz": raiz or "-",
                "motorista_origem": upper_text(row.get("motorista_origem")) or "-",
                "motorista_destino": upper_text(row.get("motorista_destino")) or "-",
                "pedido": upper_text(row.get("pedido")) or "-",
                "cliente": upper_text(row.get("cod_cliente")) or "-",
                "caixas": qtd,
                "caixas_convertidas": qtd_convertida,
                "caixas_saldo": max(qtd - qtd_convertida, 0),
                "kg_estimado": kg_transf,
                "kg_convertido_estimado": kg_convertido,
                "status": upper_text(row.get("status")) or "-",
                "data": str(row.get("atualizado_em") or row.get("criado_em") or "")[:19],
                "obs": row.get("obs") or "-",
            }
        )

    rows.sort(key=lambda item: (item["operacao"] != "TRANSBORDO", item["codigo_programacao"]))
    lines = [
        f"RELATORIO POR NOTA FISCAL / TRANSBORDO - NF {nf_norm}",
        "=" * 100,
        f"Programacoes vinculadas: {len(rows)}",
        f"Caixas totais programadas/recebidas: {plain_number(total_caixas)}",
        f"KG carregado: {total_kg_carregado:.2f} | KG entregue: {total_kg_entregue:.2f} | Divergencia: {(total_kg_entregue - total_kg_carregado):.2f}",
        f"Mortalidade: {total_mort_aves} aves / {total_mort_kg:.2f} kg",
        f"Recebimentos: {fmt_money(total_recebido)} | Despesas: {fmt_money(total_despesas)}",
        "",
        "[PROGRAMACOES]",
        "CODIGO | OPERACAO | MOTORISTA | VEICULO | CX | KG CARREG. | KG ENTREG. | MORT | STATUS",
        "-" * 100,
    ]
    for row in rows:
        lines.append(
            f"{row['codigo_programacao']} | {row['operacao']} | {row['motorista']} | {row['veiculo']} | "
            f"{plain_number(row['caixas'])} | {safe_float(row['kg_carregado'], 0):.2f} | "
            f"{safe_float(row['kg_entregue'], 0):.2f} | {row['mortalidade_aves']} / {safe_float(row['mortalidade_kg'], 0):.2f}kg | {row['status']}"
        )
    lines.extend(["", "[TRANSBORDO / TRANSFERENCIAS]"])
    if transfer_rows:
        for trans in transfer_rows:
            lines.append(
                f"{trans['origem']} -> {trans['destino']} | raiz {trans['raiz']} | CX {trans['caixas']} | "
                f"CONV {trans['caixas_convertidas']} | SALDO {trans['caixas_saldo']} | "
                f"KG {safe_float(trans['kg_estimado'], 0):.2f} | KG CONV {safe_float(trans['kg_convertido_estimado'], 0):.2f} | "
                f"{trans['motorista_origem']} -> {trans['motorista_destino']} | {trans['status']}"
            )
    else:
        lines.append("Sem transferencias vinculadas.")

    sections = [
        RelatorioSection(
            title="Fluxo de Transbordo e Transferencias",
            columns=[
                col("origem", "ORIGEM"), col("destino", "DESTINO"), col("raiz", "RAIZ"),
                col("motorista_origem", "MOT. ORIGEM"), col("motorista_destino", "MOT. DESTINO"),
                col("caixas", "CX", "number"), col("kg_estimado", "KG EST.", "number"),
                col("caixas_convertidas", "CX CONV.", "number"), col("caixas_saldo", "CX SALDO", "number"),
                col("kg_convertido_estimado", "KG CONV. EST.", "number"),
                col("pedido", "PEDIDO"), col("cliente", "CLIENTE"), col("status", "STATUS"),
                col("data", "DATA"), col("obs", "OBS"),
            ],
            rows=transfer_rows,
        ),
        RelatorioSection(
            title="Despesas Vinculadas",
            columns=[
                col("codigo_programacao", "PROGRAMACAO"), col("motorista", "MOTORISTA"), col("veiculo", "VEICULO"),
                col("categoria", "CATEGORIA"), col("descricao", "DESCRICAO"), col("valor", "VALOR", "money"),
                col("data", "DATA"), col("obs", "OBS"),
            ],
            rows=despesa_rows,
        ),
        RelatorioSection(
            title="Fotos / Evidencias do App",
            columns=[
                col("codigo_programacao", "PROGRAMACAO"), col("categoria", "CATEGORIA"),
                col("cliente", "CLIENTE"), col("arquivo", "ARQUIVO"), col("data", "DATA"),
            ],
            rows=fotos_rows,
        ),
    ]

    return RelatorioResumoResponse(
        tipo="Nota Fiscal / Transbordo",
        programacao="",
        status=relatorio_status("Relatorio por nota fiscal gerado", len(rows)),
        kpis=[
            RelatorioKpi(label="Programacoes", value=str(len(rows))),
            RelatorioKpi(label="KG carregado", value=f"{total_kg_carregado:.2f}"),
            RelatorioKpi(label="KG entregue", value=f"{total_kg_entregue:.2f}"),
            RelatorioKpi(label="Mortalidade", value=f"{total_mort_aves} / {total_mort_kg:.2f} kg"),
            RelatorioKpi(label="Recebimentos", value=fmt_money(total_recebido)),
            RelatorioKpi(label="Despesas", value=fmt_money(total_despesas)),
        ],
        chart=[
            RelatorioChartItem(label="KG carregado", value=total_kg_carregado),
            RelatorioChartItem(label="KG entregue", value=total_kg_entregue),
            RelatorioChartItem(label="Mortalidade KG", value=total_mort_kg),
            RelatorioChartItem(label="Despesas", value=total_despesas),
        ],
        columns=[
            col("codigo_programacao", "PROGRAMACAO"), col("nf_numero", "NF"), col("operacao", "OPERACAO"),
            col("motorista", "MOTORISTA"), col("veiculo", "VEICULO"), col("ajudantes", "AJUDANTES"),
            col("caixas", "CX", "number"), col("kg_carregado", "KG CARREG.", "number"),
            col("kg_entregue", "KG ENTREG.", "number"), col("divergencia_kg", "DIVERG. KG", "number"),
            col("kg_transferido_entrada", "KG TRANSF. ENTR.", "number"),
            col("kg_transferido_saida", "KG TRANSF. SAIDA", "number"),
            col("mortalidade_aves", "MORT AVES", "number"), col("mortalidade_kg", "MORT KG", "number"),
            col("recebido", "RECEBIDO", "money"), col("despesas", "DESPESAS", "money"),
            col("status", "STATUS"), col("prestacao", "PRESTACAO"),
        ],
        rows=rows,
        text="\n".join(lines),
        sections=sections,
    )


async def build_programacao_report(db: AsyncSession, programacao: ProgramacaoDB) -> RelatorioResumoResponse:
    codigo = upper_text(programacao.codigo_programacao)
    itens = await itens_for(db, codigo)
    equipe = await resolve_equipe_nomes(db, programacao.equipe)
    valor_total = sum(normalize_unit_price(item.preco) for item in itens)
    preco_medio = valor_total / max(len(itens), 1)

    lines = [
        f"RELATORIO DE PROGRAMACAO - {codigo}",
        "=" * 90,
        "",
        "[IDENTIFICACAO]",
        f"Codigo: {codigo}",
        f"Status: {status_ref(programacao) or '-'}",
        f"Data: {normalize_date(data_ref(programacao)) or data_ref(programacao) or '-'}",
        f"Usuario criacao: {programacao.usuario_criacao or '-'}",
        f"Motorista: {programacao.motorista or '-'}",
        f"Equipe: {equipe or '-'}",
        f"Veiculo: {programacao.veiculo or '-'}",
        f"Local da rota: {normalize_local_rota_display(programacao.local_rota or programacao.tipo_rota) or '-'}",
        f"Local carregamento: {upper_text(programacao.local_carregamento or programacao.granja_carregada or programacao.local_carregado or programacao.local_carreg) or '-'}",
        f"KG estimado: {safe_float(programacao.kg_estimado, 0.0):.2f}",
        f"Clientes na programacao: {len(itens)}",
        f"Total estimado (clientes): {fmt_money(valor_total)}",
        "",
        "[CLIENTES / PRECO / VENDEDOR]",
    ]
    rows = []
    if not itens:
        lines.append("Sem clientes cadastrados na programacao.")
    else:
        lines.extend(["COD | CLIENTE | PRECO | VENDEDOR", "-" * 90])
        for item in itens:
            preco = normalize_unit_price(item.preco)
            rows.append(
                {
                    "cod_cliente": upper_text(item.cod_cliente),
                    "nome_cliente": upper_text(item.nome_cliente),
                    "preco": money(preco),
                    "vendedor": upper_text(item.vendedor) or "-",
                }
            )
            lines.append(
                f"{upper_text(item.cod_cliente) or '-'} | {upper_text(item.nome_cliente) or '-'} | "
                f"{fmt_money(preco)} | {upper_text(item.vendedor) or '-'}"
            )

    return RelatorioResumoResponse(
        tipo="Planejamentos",
        programacao=codigo,
        status=relatorio_status("Resumo de programacao gerado."),
        kpis=[
            RelatorioKpi(label="Clientes", value=str(len(itens))),
            RelatorioKpi(label="Total estimado", value=fmt_money(valor_total)),
            RelatorioKpi(label="Preco medio", value=fmt_money(preco_medio)),
            RelatorioKpi(label="Motorista", value=upper_text(programacao.motorista) or "-"),
        ],
        chart=[
            RelatorioChartItem(label=upper_text(item.cod_cliente) or "-", value=money(normalize_unit_price(item.preco)))
            for item in itens[:8]
        ],
        columns=[
            col("cod_cliente", "COD"),
            col("nome_cliente", "CLIENTE"),
            col("preco", "PRECO", "money"),
            col("vendedor", "VENDEDOR"),
        ],
        rows=rows,
        text="\n".join(lines),
    )


async def build_prestacao_report(
    db: AsyncSession,
    programacao: ProgramacaoDB,
    *,
    show_recebimentos: bool,
    show_despesas: bool,
) -> RelatorioResumoResponse:
    codigo = upper_text(programacao.codigo_programacao)
    itens = await itens_for(db, codigo)
    recebimentos = await recebimentos_for(db, codigo)
    despesas = await despesas_for(db, codigo)
    equipe = await resolve_equipe_nomes(db, programacao.equipe)
    transferencias, transferencia_saida_compra, transferencia_entrada_compra = await transferencias_programacao(db, codigo)

    total_receb = money(sum(safe_float(item.valor, 0.0) for item in recebimentos))
    total_desp = money(sum(safe_float(item.valor, 0.0) for item in despesas))
    adiantamento = money(safe_float(programacao.adiantamento, 0.0) or safe_float(programacao.adiantamento_rota, 0.0))
    ced_qtd = {ced: safe_int(getattr(programacao, f"ced_{ced}_qtd", 0), 0) for ced in CEDULAS}
    ced_total = money(sum(float(ced) * qtd for ced, qtd in ced_qtd.items()))
    dinheiro_total = money(programacao.valor_dinheiro) if safe_float(programacao.valor_dinheiro, 0.0) > 0 else ced_total
    total_entradas = money(total_receb + adiantamento)
    total_saidas = money(total_desp + ced_total)
    valor_final_caixa = money(total_entradas - total_desp)
    diferenca = money(valor_final_caixa - ced_total)
    resultado = money(total_entradas - total_saidas)
    prestacao = upper_text(programacao.prestacao_status or "PENDENTE")
    operacao_tipo = "TRANSBORDO" if is_transbordo_programacao(programacao) else "VENDA"

    local_rota = normalize_local_rota_display(programacao.local_rota or programacao.tipo_rota)
    local_carreg = upper_text(
        programacao.local_carregamento
        or programacao.granja_carregada
        or programacao.local_carregado
        or programacao.local_carreg
    )
    nf_kg = safe_float(programacao.nf_kg, 0.0)
    nf_caixas = safe_int(programacao.nf_caixas, 0)
    nf_kg_carregado = safe_float(programacao.nf_kg_carregado or programacao.kg_carregado, 0.0)
    nf_kg_vendido = safe_float(programacao.nf_kg_vendido, 0.0)
    nf_saldo = safe_float(programacao.nf_saldo, 0.0)
    km_inicial = safe_float(programacao.km_inicial, 0.0)
    km_final = safe_float(programacao.km_final, 0.0)
    litros = safe_float(programacao.litros, 0.0)
    km_rodado = safe_float(programacao.km_rodado, 0.0)
    media_km_l = safe_float(programacao.media_km_l, 0.0)
    custo_km = safe_float(programacao.custo_km, 0.0)

    lines = [
        f"RELATORIO DE PRESTACAO DE CONTAS / FECHAMENTO OPERACIONAL - PLANEJAMENTO {codigo}",
        "Tipo: Prestacao de Contas",
        "=" * 90,
        "",
        "[IDENTIFICACAO]",
        f"Status: {status_ref(programacao) or '-'}",
        f"Fechamento: {prestacao or '-'}",
        f"Operacao: {operacao_tipo} | Modalidade: {upper_text(getattr(programacao, 'transbordo_modalidade', '') or '') or '-'}",
        f"Data criacao: {normalize_date(data_ref(programacao)) or data_ref(programacao) or '-'}",
        f"Motorista: {programacao.motorista or '-'}",
        f"Veiculo: {programacao.veiculo or '-'}",
        f"Equipe: {equipe or '-'}",
        f"Entregas (itens): {len(itens)}",
        f"KG estimado: {safe_float(programacao.kg_estimado, 0.0):.2f}",
        "",
        "[DADOS DA ROTA]",
        f"NF: {upper_text(programacao.nf_numero or programacao.num_nf) or '-'}",
        f"Local da rota: {local_rota or '-'}",
        f"Local carregamento: {local_carreg or '-'}",
        f"Saida: {str((programacao.data_saida or '') + ' ' + (programacao.hora_saida or '')).strip() or '-'}",
        f"Chegada: {str((programacao.data_chegada or '') + ' ' + (programacao.hora_chegada or '')).strip() or '-'}",
        "",
        "[NOTA FISCAL / CARREGAMENTO]",
        f"NF KG: {nf_kg:.2f}",
        f"NF caixas: {nf_caixas}",
        f"KG carregado: {nf_kg_carregado:.2f}",
        f"KG vendido: {nf_kg_vendido:.2f}",
        f"Saldo (KG): {nf_saldo:.2f}",
        "",
        "[ROTA / KM]",
        f"KM inicial: {km_inicial:.2f}",
        f"KM final: {km_final:.2f}",
        f"Litros: {litros:.2f}",
        f"KM rodado: {km_rodado:.2f}",
        f"Media km/l: {media_km_l:.2f}",
        f"Custo por KM: {custo_km:.2f}",
        "",
        "[CONTAGEM DE CEDULAS]",
    ]
    for ced in CEDULAS:
        qtd = safe_int(ced_qtd.get(ced, 0), 0)
        lines.append(f"R$ {ced:>3},00 -> QTD {qtd:>4} -> TOTAL {fmt_money(qtd * ced)}")
    lines.extend(
        [
            f"Total cedulas: {fmt_money(ced_total)}",
            f"Total dinheiro (campo): {fmt_money(dinheiro_total)}",
            "",
            "[RESUMO FINANCEIRO]",
            f"Recebimentos: {fmt_money(total_receb)}",
            f"Adiantamento: {fmt_money(adiantamento)}",
            f"Custos: {fmt_money(total_desp)}",
            f"Cedulas: {fmt_money(ced_total)}",
            f"Total entradas: {fmt_money(total_entradas)}",
            f"Total saidas: {fmt_money(total_saidas)}",
            f"Valor final caixa: {fmt_money(valor_final_caixa)}",
            f"Diferenca caixa x cedulas: {fmt_money(diferenca)}",
            f"Resultado liquido: {fmt_money(resultado)}",
            "",
            "[TRANSBORDO / TRANSFERENCIAS]",
            f"Compra transferida saida: {fmt_money(transferencia_saida_compra)}",
            f"Compra transferida entrada: {fmt_money(transferencia_entrada_compra)}",
        ]
    )
    if transferencias:
        for trans in transferencias:
            lines.append(
                f"{trans['direcao']} | {trans['origem']} -> {trans['destino']} | "
                f"{plain_number(trans['caixas'])} cx | conv {plain_number(trans.get('caixas_convertidas', 0))} | "
                f"saldo {plain_number(trans.get('caixas_saldo', 0))} | "
                f"KG conv {safe_float(trans.get('kg_convertido_estimado'), 0):.2f} | "
                f"{fmt_money(trans['valor_compra'])} | {trans['status']}"
            )
    else:
        lines.append("Sem transferencias vinculadas.")
    lines.append("")

    detail_rows = []
    for trans in transferencias:
        detail_rows.append(
            {
                "bloco": "TRANSBORDO",
                "data": trans.get("atualizado_em") or "",
                "referencia": f"{trans.get('origem')} -> {trans.get('destino')}",
                "descricao": (
                    f"{trans.get('caixas', 0)} cx / {trans.get('pedido')} "
                    f"(conv {trans.get('caixas_convertidas', 0)} / saldo {trans.get('caixas_saldo', 0)})"
                ),
                "valor": money(trans.get("valor_compra")),
                "categoria": trans.get("direcao") or "TRANSBORDO",
                "observacao": (
                    f"{trans.get('status') or '-'} | "
                    f"KG total {safe_float(trans.get('kg_estimado'), 0):.2f} | "
                    f"KG conv {safe_float(trans.get('kg_convertido_estimado'), 0):.2f}"
                ),
                "caixas": trans.get("caixas", 0),
                "caixas_convertidas": trans.get("caixas_convertidas", 0),
                "caixas_saldo": trans.get("caixas_saldo", 0),
                "kg_estimado": trans.get("kg_estimado", 0),
                "kg_convertido_estimado": trans.get("kg_convertido_estimado", 0),
            }
        )
    if show_recebimentos:
        lines.append("[RECEBIMENTOS DETALHADOS]")
        if not recebimentos:
            lines.extend(["Sem recebimentos registrados.", ""])
        else:
            for item in recebimentos:
                row = {
                    "bloco": "RECEBIMENTO",
                    "data": str(item.data_registro or "")[:19],
                    "referencia": upper_text(item.cod_cliente),
                    "descricao": upper_text(item.nome_cliente),
                    "valor": money(item.valor),
                    "categoria": upper_text(item.forma_pagamento) or "-",
                    "observacao": item.observacao or "-",
                }
                detail_rows.append(row)
                lines.append(
                    f"{row['data']} | {row['referencia']} | {row['descricao']} | "
                    f"{fmt_money(row['valor'])} | {row['categoria']} | {row['observacao']}"
                )
            lines.append("")

    if show_despesas:
        lines.append("[CUSTOS DETALHADOS]")
        if not despesas:
            lines.append("Sem custos registrados.")
        else:
            for item in despesas:
                row = {
                    "bloco": "DESPESA",
                    "data": str(item.data_registro or "")[:19],
                    "referencia": upper_text(item.categoria) or "OUTROS",
                    "descricao": upper_text(item.descricao) or "-",
                    "valor": money(item.valor),
                    "categoria": upper_text(item.categoria) or "OUTROS",
                    "observacao": item.observacao or "-",
                }
                detail_rows.append(row)
                lines.append(
                    f"{row['data']} | {row['categoria']} | {row['descricao']} | "
                    f"{fmt_money(row['valor'])} | {row['observacao']}"
                )

    if prestacao == "FECHADA":
        lines.extend(["", "[ALERTA] Fechamento CONCLUIDO: alteracoes financeiras estao bloqueadas."])

    return RelatorioResumoResponse(
        tipo="Fechamento Operacional",
        programacao=codigo,
        status=relatorio_status("Resumo detalhado gerado."),
        kpis=[
            RelatorioKpi(label="Recebimentos", value=fmt_money(total_receb)),
            RelatorioKpi(label="Adiantamento", value=fmt_money(adiantamento)),
            RelatorioKpi(label="Resultado", value=fmt_money(resultado)),
            RelatorioKpi(label="Custos", value=fmt_money(total_desp)),
            RelatorioKpi(label="Dinheiro contado", value=fmt_money(dinheiro_total)),
            RelatorioKpi(label="Valor final caixa", value=fmt_money(valor_final_caixa)),
            RelatorioKpi(label="Diferenca", value=fmt_money(diferenca)),
            RelatorioKpi(label="Fechamento", value=prestacao or "-"),
            RelatorioKpi(label="KG carregado", value=f"{nf_kg_carregado:.2f}"),
            RelatorioKpi(label="KG vendido", value=f"{nf_kg_vendido:.2f}"),
            RelatorioKpi(label="KM rodado", value=f"{km_rodado:.2f}"),
            RelatorioKpi(label="Media KM/L", value=f"{media_km_l:.2f}"),
        ],
        chart=[
            RelatorioChartItem(label="Entradas", value=total_entradas),
            RelatorioChartItem(label="Saidas", value=total_saidas),
            RelatorioChartItem(label="Resultado", value=abs(resultado)),
        ],
        columns=[
            col("bloco", "BLOCO"),
            col("data", "DATA"),
            col("referencia", "REFERENCIA"),
            col("categoria", "CATEGORIA"),
            col("descricao", "DESCRICAO"),
            col("valor", "VALOR", "money"),
            col("observacao", "OBSERVACAO"),
        ],
        rows=detail_rows,
        text="\n".join(lines),
    )


async def build_rotina_report(db: AsyncSession, motorista_like: str) -> RelatorioResumoResponse:
    rows = await programacao_rows(db)
    programacoes = {upper_text(item.codigo_programacao): item for item in rows}
    transfer_entrada_kg_by_codigo: dict[str, float] = {}
    for transfer in await transferencias_rows(db):
        status_value = normalize_key(transfer.get("status"))
        if status_value in {"CANCELADA", "CANCELADO", "RECUSADA", "RECUSADO"}:
            continue
        destino = upper_text(transfer.get("codigo_destino"))
        if not destino:
            continue
        qtd_convertida = transferencia_qtd_convertida(transfer)
        if qtd_convertida <= 0:
            continue
        kg_convertido = transferencia_kg_estimado(programacoes, transfer, qtd_convertida)
        if kg_convertido > 0:
            transfer_entrada_kg_by_codigo[destino] = transfer_entrada_kg_by_codigo.get(destino, 0.0) + kg_convertido
    mot: dict[str, dict[str, float]] = {}
    aju: dict[str, dict[str, float]] = {}

    for programacao in rows:
        codigo = upper_text(programacao.codigo_programacao)
        motorista = upper_text(programacao.motorista or "SEM MOTORISTA")
        if motorista_like and motorista_like not in motorista:
            continue
        d = mot.setdefault(motorista, {"viagens": 0, "kg": 0.0, "km": 0.0, "em_rota": 0})
        d["viagens"] += 1
        d["kg"] += safe_float(programacao.nf_kg_vendido, 0.0) or transfer_entrada_kg_by_codigo.get(codigo, 0.0)
        d["km"] += safe_float(programacao.km_rodado, 0.0)
        if status_ref(programacao) in ACTIVE_STATUSES:
            d["em_rota"] += 1
        equipe = await resolve_equipe_nomes(db, programacao.equipe)
        for nome in re.split(r"[|,;/]+", equipe or ""):
            ajudante = upper_text(nome)
            if not ajudante or ajudante in {"-", "NAN", "NONE", "SEM EQUIPE"}:
                continue
            a = aju.setdefault(ajudante, {"viagens": 0, "km": 0.0})
            a["viagens"] += 1
            a["km"] += safe_float(programacao.km_rodado, 0.0)

    rank_m = sorted(mot.items(), key=lambda item: (-item[1]["viagens"], item[0]))
    rank_a = sorted(aju.items(), key=lambda item: (-item[1]["viagens"], item[0]))
    lines = [
        "RELATORIO DE ROTINA - MOTORISTAS E AJUDANTES",
        "=" * 100,
        "",
        "[MOTORISTAS]",
        "MOTORISTA | VIAGENS | KG ENTREGUES | KM RODADO | MEDIA KM/VIAGEM | EM ROTA",
        "-" * 100,
    ]
    out_rows = []
    for nome, data in rank_m:
        media_km = data["km"] / max(data["viagens"], 1)
        out_rows.append(
            {
                "tipo": "MOTORISTA",
                "nome": nome,
                "viagens": safe_int(data["viagens"], 0),
                "kg": number2(data["kg"]),
                "km": number2(data["km"]),
                "media_km": number2(media_km),
                "em_rota": safe_int(data["em_rota"], 0),
            }
        )
        lines.append(
            f"{nome} | {safe_int(data['viagens'], 0)} | {data['kg']:.2f} | "
            f"{data['km']:.2f} | {media_km:.2f} | {safe_int(data['em_rota'], 0)}"
        )

    lines.extend(["", "[AJUDANTES]", "AJUDANTE | VIAGENS | KM RODADO | MEDIA KM/VIAGEM", "-" * 100])
    for nome, data in rank_a:
        media_km = data["km"] / max(data["viagens"], 1)
        out_rows.append(
            {
                "tipo": "AJUDANTE",
                "nome": nome,
                "viagens": safe_int(data["viagens"], 0),
                "kg": 0,
                "km": number2(data["km"]),
                "media_km": number2(media_km),
                "em_rota": 0,
            }
        )
        lines.append(f"{nome} | {safe_int(data['viagens'], 0)} | {data['km']:.2f} | {media_km:.2f}")

    top_nome = rank_m[0][0] if rank_m else "-"
    top_viagens = safe_int(rank_m[0][1]["viagens"], 0) if rank_m else 0
    return RelatorioResumoResponse(
        tipo="Rotina Motorista/Ajudantes",
        status=relatorio_status("Relatorio de rotina gerado", len(rank_m)),
        kpis=[
            RelatorioKpi(label="Registros", value=f"{len(rank_m)} motoristas"),
            RelatorioKpi(label="Total viagens", value=str(sum(safe_int(d["viagens"], 0) for _n, d in rank_m))),
            RelatorioKpi(label="Total KG", value=f"{sum(safe_float(d['kg'], 0.0) for _n, d in rank_m):.2f}"),
            RelatorioKpi(label="Destaque", value=f"{top_nome} ({top_viagens} viagens)"),
        ],
        chart=[RelatorioChartItem(label=nome, value=safe_float(data["viagens"], 0.0)) for nome, data in rank_m[:8]],
        columns=[
            col("tipo", "TIPO"),
            col("nome", "NOME"),
            col("viagens", "VIAGENS", "number"),
            col("kg", "KG", "number"),
            col("km", "KM RODADO", "number"),
            col("media_km", "MEDIA KM/VIAGEM", "number"),
            col("em_rota", "EM ROTA", "number"),
        ],
        rows=out_rows,
        text="\n".join(lines),
    )


async def build_km_report(
    db: AsyncSession,
    *,
    codigo_like: str = "",
    motorista_like: str = "",
    data_like: str = "",
) -> RelatorioResumoResponse:
    despesas_result = await db.execute(select(DespesaDB))
    despesas_por_codigo: dict[str, dict[str, float]] = {}
    for despesa in despesas_result.scalars().all():
        codigo = upper_text(despesa.codigo_programacao)
        item = despesas_por_codigo.setdefault(codigo, {"valor": 0.0, "litros": 0.0})
        item["valor"] += safe_float(despesa.valor, 0.0)
        item["litros"] += safe_float(despesa.litros, 0.0)

    grouped: dict[str, dict[str, float]] = {}
    for programacao in await programacao_rows(db):
        codigo = upper_text(programacao.codigo_programacao)
        if codigo_like and codigo_like not in codigo and codigo_like not in nf_programacao(programacao):
            continue
        if not matches_programacao(
            programacao,
            tipo="KM de Veiculos",
            codigo_like="",
            motorista_like=motorista_like,
            data_like=data_like,
        ):
            continue
        veiculo = upper_text(programacao.veiculo) or "-"
        data = grouped.setdefault(
            veiculo,
            {"viagens": 0, "km": 0.0, "litros": 0.0, "custo": 0.0, "media_sum": 0.0, "media_count": 0},
        )
        despesas_prog = despesas_por_codigo.get(codigo, {"valor": 0.0, "litros": 0.0})
        km = safe_float(programacao.km_rodado, 0.0)
        litros = safe_float(programacao.litros, 0.0) or safe_float(despesas_prog["litros"], 0.0)
        custo = safe_float(despesas_prog["valor"], 0.0)
        media = safe_float(programacao.media_km_l, 0.0) or ((km / litros) if litros > 0 else 0.0)
        data["viagens"] += 1
        data["km"] += km
        data["litros"] += litros
        data["custo"] += custo
        if media > 0:
            data["media_sum"] += media
            data["media_count"] += 1
    sorted_rows = sorted(grouped.items(), key=lambda item: (-item[1]["km"], item[0]))
    out_rows = []
    lines = [
        "RELATORIO DE KM POR VEICULO",
        "=" * 90,
        "",
        "VEICULO | VIAGENS | KM RODADO | LITROS | MEDIA KM/L | CUSTO | CUSTO/KM",
        "-" * 90,
    ]
    for veiculo, data in sorted_rows:
        media = (data["km"] / data["litros"]) if safe_float(data["litros"], 0.0) > 0 else (data["media_sum"] / max(data["media_count"], 1))
        custo_km = safe_float(data["custo"], 0.0) / safe_float(data["km"], 1.0) if safe_float(data["km"], 0.0) > 0 else 0.0
        row = {
            "veiculo": veiculo,
            "viagens": safe_int(data["viagens"], 0),
            "km_rodado": number2(data["km"]),
            "litros": number2(data["litros"]),
            "media_km_l": number2(media),
            "custo_total": money(data["custo"]),
            "custo_km": money(custo_km),
        }
        out_rows.append(row)
        lines.append(
            f"{veiculo} | {row['viagens']} | {row['km_rodado']:.2f} | {row['litros']:.2f} | "
            f"{row['media_km_l']:.2f} | {fmt_money(row['custo_total'])} | {fmt_money(row['custo_km'])}"
        )
    total_km = sum(safe_float(data["km"], 0.0) for _veic, data in sorted_rows)
    total_litros = sum(safe_float(data["litros"], 0.0) for _veic, data in sorted_rows)
    total_custo = money(sum(safe_float(data["custo"], 0.0) for _veic, data in sorted_rows))
    top = sorted_rows[0][0] if sorted_rows else "-"
    return RelatorioResumoResponse(
        tipo="KM de Veiculos",
        status=relatorio_status("Relatorio de KM por veiculo gerado", len(sorted_rows)),
        kpis=[
            RelatorioKpi(label="Veiculos", value=str(len(sorted_rows))),
            RelatorioKpi(label="KM total", value=f"{total_km:.2f}"),
            RelatorioKpi(label="Litros", value=f"{total_litros:.2f}"),
            RelatorioKpi(label="Media KM/L", value=f"{(total_km / total_litros if total_litros > 0 else 0.0):.2f}"),
            RelatorioKpi(label="Custo total", value=fmt_money(total_custo)),
            RelatorioKpi(label="Destaque", value=top),
        ],
        chart=[RelatorioChartItem(label=veic, value=number2(data["km"])) for veic, data in sorted_rows[:8]],
        columns=[
            col("veiculo", "VEICULO"),
            col("viagens", "VIAGENS", "number"),
            col("km_rodado", "KM RODADO", "number"),
            col("litros", "LITROS", "number"),
            col("media_km_l", "MEDIA KM/L", "number"),
            col("custo_total", "CUSTO", "money"),
            col("custo_km", "CUSTO/KM", "money"),
        ],
        rows=out_rows,
        text="\n".join(lines),
    )


async def build_despesas_report(db: AsyncSession) -> RelatorioResumoResponse:
    result = await db.execute(select(DespesaDB))
    grouped: dict[str, dict[str, float]] = {}
    for despesa in result.scalars().all():
        categoria = upper_text(despesa.categoria) or "OUTROS"
        item = grouped.setdefault(categoria, {"qtd": 0, "total": 0.0})
        item["qtd"] += 1
        item["total"] += safe_float(despesa.valor, 0.0)
    sorted_rows = sorted(grouped.items(), key=lambda item: (-item[1]["total"], item[0]))
    rows = []
    lines = [
        "RELATORIO GERAL DE CUSTOS",
        "=" * 90,
        "",
        "CATEGORIA | QTD LANCAMENTOS | TOTAL",
        "-" * 90,
    ]
    for categoria, data in sorted_rows:
        row = {"categoria": categoria, "qtd": safe_int(data["qtd"], 0), "total": money(data["total"])}
        rows.append(row)
        lines.append(f"{categoria} | {row['qtd']} | {fmt_money(row['total'])}")
    total_geral = money(sum(safe_float(data["total"], 0.0) for _cat, data in sorted_rows))
    top = sorted_rows[0][0] if sorted_rows else "-"
    return RelatorioResumoResponse(
        tipo="Custos e Despesas",
        status=relatorio_status("Relatorio de custos gerado", len(sorted_rows)),
        kpis=[
            RelatorioKpi(label="Categorias", value=str(len(sorted_rows))),
            RelatorioKpi(label="Total custos", value=fmt_money(total_geral)),
            RelatorioKpi(label="Media/categoria", value=fmt_money(total_geral / max(len(sorted_rows), 1))),
            RelatorioKpi(label="Maior categoria", value=top),
        ],
        chart=[RelatorioChartItem(label=cat, value=money(data["total"])) for cat, data in sorted_rows[:8]],
        columns=[col("categoria", "CATEGORIA"), col("qtd", "QTD", "number"), col("total", "TOTAL", "money")],
        rows=rows,
        text="\n".join(lines),
    )


async def build_despesas_eventos_report(
    db: AsyncSession,
    *,
    tipo: str,
    codigo_like: str = "",
    motorista_like: str = "",
    data_like: str = "",
) -> RelatorioResumoResponse:
    tipo_key = normalize_key(tipo)
    is_abastecimento = "ABASTEC" in tipo_key or "COMBUST" in tipo_key
    tokens = (
        ("ABASTEC", "COMBUST", "DIESEL", "GASOL", "ETANOL", "ALCOOL", "ARLA")
        if is_abastecimento
        else ("BANHO", "LAVAGEM", "HIGIENIZ", "LIMPEZA")
    )
    result = await db.execute(select(DespesaDB).order_by(DespesaDB.data_registro.desc(), DespesaDB.id.desc()))
    programacoes = {upper_text(item.codigo_programacao): item for item in await programacao_rows(db, limit=3000)}
    patterns = date_patterns(data_like)
    rows: list[dict[str, Any]] = []
    grouped: dict[str, dict[str, float]] = {}
    total_valor = 0.0
    total_litros = 0.0
    for despesa in result.scalars().all():
        codigo = upper_text(despesa.codigo_programacao)
        prog = programacoes.get(codigo)
        text_match = normalize_key(f"{despesa.categoria or ''} {despesa.descricao or ''} {despesa.combustivel or ''}")
        if not any(token in text_match for token in tokens):
            continue
        motorista = upper_text(despesa.motorista_nome or despesa.motorista or (prog.motorista if prog else ""))
        veiculo = upper_text(despesa.veiculo or (prog.veiculo if prog else ""))
        data_evento = str(despesa.data_registro or despesa.registrado_em or "")
        nf_ref = nf_programacao(prog) if prog else ""
        if codigo_like and codigo_like not in codigo and codigo_like not in nf_ref:
            continue
        if motorista_like and motorista_like not in motorista:
            continue
        data_match_ref = data_evento or (data_ref(prog) if prog else "")
        if not matches_date(data_match_ref, patterns):
            continue
        valor = money(despesa.valor)
        litros = number2(despesa.litros)
        total_valor += valor
        total_litros += litros
        group_key = veiculo if is_abastecimento else motorista
        grouped_item = grouped.setdefault(group_key or "-", {"qtd": 0, "valor": 0.0, "litros": 0.0})
        grouped_item["qtd"] += 1
        grouped_item["valor"] += valor
        grouped_item["litros"] += litros
        rows.append(
            {
                "codigo_programacao": codigo,
                "nf_numero": nf_programacao(prog) if prog else "-",
                "data": data_evento[:19],
                "motorista": motorista or "-",
                "veiculo": veiculo or "-",
                "categoria": upper_text(despesa.categoria) or "-",
                "descricao": upper_text(despesa.descricao) or "-",
                "combustivel": upper_text(despesa.combustivel) or "-",
                "litros": litros,
                "valor_litro": money(despesa.valor_litro),
                "valor": valor,
                "odometro": number2(despesa.odometro),
                "obs": despesa.observacao or "-",
            }
        )
    ranking = sorted(grouped.items(), key=lambda item: (-safe_float(item[1]["valor"], 0.0), item[0]))
    title = "Abastecimentos" if is_abastecimento else "Banhos"
    lines = [
        f"RELATORIO DE {normalize_key(title)}",
        "=" * 90,
        f"Filtro codigo/NF: {codigo_like or '-'} | Motorista: {motorista_like or '-'} | Data: {data_like or '-'}",
        f"Lancamentos: {len(rows)} | Total: {fmt_money(total_valor)} | Litros: {total_litros:.2f}",
        "",
        "PROGRAMACAO | NF | DATA | MOTORISTA | VEICULO | DESCRICAO | LITROS | VALOR",
        "-" * 90,
    ]
    for row in rows[:250]:
        lines.append(
            f"{row['codigo_programacao']} | {row['nf_numero']} | {row['data']} | {row['motorista']} | "
            f"{row['veiculo']} | {row['descricao']} | {safe_float(row['litros'], 0):.2f} | {fmt_money(row['valor'])}"
        )
    if len(rows) > 250:
        lines.append(f"... e mais {len(rows) - 250} lancamento(s).")

    sections = [
        RelatorioSection(
            title=f"Ranking por {'Veiculo' if is_abastecimento else 'Motorista'}",
            columns=[
                col("nome", "NOME"), col("qtd", "QTD", "number"),
                col("litros", "LITROS", "number"), col("valor", "VALOR", "money"),
            ],
            rows=[
                {"nome": nome, "qtd": safe_int(data["qtd"], 0), "litros": number2(data["litros"]), "valor": money(data["valor"])}
                for nome, data in ranking
            ],
        )
    ]
    return RelatorioResumoResponse(
        tipo=title,
        status=relatorio_status(f"Relatorio de {title.lower()} gerado", len(rows)),
        kpis=[
            RelatorioKpi(label="Lancamentos", value=str(len(rows))),
            RelatorioKpi(label="Total", value=fmt_money(total_valor)),
            RelatorioKpi(label="Litros", value=f"{total_litros:.2f}"),
            RelatorioKpi(label="Maior grupo", value=ranking[0][0] if ranking else "-"),
        ],
        chart=[RelatorioChartItem(label=nome, value=money(data["valor"])) for nome, data in ranking[:8]],
        columns=[
            col("codigo_programacao", "PROGRAMACAO"), col("nf_numero", "NF"), col("data", "DATA"),
            col("motorista", "MOTORISTA"), col("veiculo", "VEICULO"), col("categoria", "CATEGORIA"),
            col("descricao", "DESCRICAO"), col("combustivel", "COMBUSTIVEL"),
            col("litros", "LITROS", "number"), col("valor_litro", "R$/L", "money"),
            col("valor", "VALOR", "money"), col("odometro", "ODOMETRO", "number"), col("obs", "OBS"),
        ],
        rows=rows,
        text="\n".join(lines),
        sections=sections,
    )


async def build_mortalidade_report(
    db: AsyncSession,
    *,
    codigo_like: str,
    motorista_like: str,
    data_like: str,
) -> RelatorioResumoResponse:
    rows: list[dict[str, Any]] = []
    detalhe_rows: list[dict[str, Any]] = []
    foto_rows: list[dict[str, Any]] = []
    for programacao in await programacao_rows(db):
        codigo = upper_text(programacao.codigo_programacao)
        nf_ref = nf_programacao(programacao)
        if codigo_like and codigo_like not in codigo and codigo_like not in nf_ref:
            continue
        if not matches_programacao(
            programacao,
            tipo="Ocorrencias por Motorista",
            codigo_like="",
            motorista_like=motorista_like,
            data_like=data_like,
        ):
            continue
        controles = await controles_for(db, codigo)
        itens = await itens_for(db, codigo)
        item_by_key = {item_key(item.cod_cliente, item.pedido): item for item in itens}
        media_prog = media_carregada_programacao(programacao)
        mort_cliente_aves = 0
        mort_cliente_kg = 0.0
        clientes_mort = 0
        for controle in controles:
            aves = safe_int(controle.mortalidade_aves, 0)
            if aves <= 0:
                continue
            item = item_by_key.get(item_key(controle.cod_cliente, controle.pedido))
            media = safe_float(controle.media_aplicada, 0.0)
            if media > 20:
                media = media / 1000.0
            if media <= 0:
                media = media_prog
            kg = number2(aves * media) if media > 0 else 0.0
            mort_cliente_aves += aves
            mort_cliente_kg += kg
            clientes_mort += 1
            detalhe_rows.append(
                {
                    "codigo_programacao": codigo,
                    "nf_numero": nf_ref or "-",
                    "motorista": upper_text(programacao.motorista) or "-",
                    "veiculo": upper_text(programacao.veiculo) or "-",
                    "cod_cliente": upper_text(controle.cod_cliente) or "-",
                    "cliente": upper_text(item.nome_cliente if item else "") or "-",
                    "pedido": upper_text(controle.pedido) or "-",
                    "mortalidade_aves": aves,
                    "mortalidade_kg": kg,
                    "media": number2(media),
                    "motivo": upper_text(controle.alteracao_detalhe or controle.alteracao_tipo or controle.status_pedido) or "-",
                    "foto": controle.foto_mortalidade_path or controle.mortalidade_foto_path or controle.foto_mortalidade_ref_json or "-",
                    "data": str(controle.updated_at or controle.alterado_em or "")[:19],
                }
            )
        mort_trans_aves = safe_int(programacao.mortalidade_transbordo_aves, 0)
        mort_trans_kg = safe_float(programacao.mortalidade_transbordo_kg, 0.0)
        if mort_trans_kg <= 0 and mort_trans_aves > 0 and media_prog > 0:
            mort_trans_kg = mort_trans_aves * media_prog
        if mort_trans_aves > 0 or mort_trans_kg > 0:
            detalhe_rows.append(
                {
                    "codigo_programacao": codigo,
                    "nf_numero": nf_ref or "-",
                    "motorista": upper_text(programacao.motorista) or "-",
                    "veiculo": upper_text(programacao.veiculo) or "-",
                    "cod_cliente": "TRANSBORDO",
                    "cliente": "TRANSBORDO / OPERACAO",
                    "pedido": "-",
                    "mortalidade_aves": mort_trans_aves,
                    "mortalidade_kg": number2(mort_trans_kg),
                    "media": number2(media_prog),
                    "motivo": upper_text(getattr(programacao, "obs_transbordo", "") or getattr(programacao, "transbordo_observacao", "")) or "MORTALIDADE TRANSBORDO",
                    "foto": getattr(programacao, "mortalidade_transbordo_foto_path", "") or getattr(programacao, "foto_doa_path", "") or getattr(programacao, "doa_foto_path", "") or getattr(programacao, "foto_doa_ref_json", "") or "-",
                    "data": data_ref(programacao),
                }
            )
        try:
            fotos_result = await db.execute(
                text(
                    """
                    SELECT codigo_programacao, categoria, tipo_registro, cod_cliente, cliente_nome,
                           arquivo_nome, storage_path, registrado_em
                      FROM rota_fotos
                     WHERE UPPER(COALESCE(codigo_programacao, ''))=:codigo
                       AND (
                           UPPER(COALESCE(categoria, '')) LIKE '%MORT%'
                        OR UPPER(COALESCE(tipo_registro, '')) LIKE '%MORT%'
                        OR UPPER(COALESCE(categoria, '')) LIKE '%DOA%'
                        OR UPPER(COALESCE(tipo_registro, '')) LIKE '%DOA%'
                       )
                     ORDER BY id DESC
                    """
                ),
                {"codigo": codigo},
            )
            for foto in fotos_result.mappings().all():
                foto_rows.append(
                    {
                        "codigo_programacao": codigo,
                        "nf_numero": nf_ref or "-",
                        "motorista": upper_text(programacao.motorista) or "-",
                        "categoria": upper_text(foto.get("categoria") or foto.get("tipo_registro")) or "-",
                        "cliente": upper_text(foto.get("cliente_nome") or foto.get("cod_cliente")) or "-",
                        "arquivo": foto.get("arquivo_nome") or foto.get("storage_path") or "-",
                        "data": str(foto.get("registrado_em") or "")[:19],
                    }
                )
        except Exception:
            pass
        mort_total_aves = mort_cliente_aves + mort_trans_aves
        mort_total_kg = number2(mort_cliente_kg + mort_trans_kg)
        rows.append(
            {
                "codigo_programacao": codigo,
                "nf_numero": nf_ref or "-",
                "motorista": upper_text(programacao.motorista),
                "veiculo": upper_text(programacao.veiculo),
                "data_ref": data_ref(programacao),
                "status_ref": status_ref(programacao),
                "mortalidade_total": mort_total_aves,
                "mortalidade_kg": mort_total_kg,
                "clientes_com_mortalidade": clientes_mort,
                "mortalidade_transbordo": mort_trans_aves,
                "mortalidade_transbordo_kg": number2(mort_trans_kg),
            }
        )
    rows.sort(key=lambda item: (-safe_int(item["mortalidade_total"], 0), -safe_float(item["mortalidade_kg"], 0.0), item["codigo_programacao"]))
    for pos, row in enumerate(rows, start=1):
        row["pos"] = pos

    lines = [
        "RELATORIO DE MORTALIDADES",
        "Tipo: Mortalidades / Ocorrencias por Motorista",
        "=" * 95,
        "",
    ]
    if not rows:
        lines.append("Nenhum dado encontrado para os filtros informados.")
        return RelatorioResumoResponse(
            tipo="Mortalidades",
            status=relatorio_status("Nenhum dado de ocorrencia encontrado."),
            kpis=[
                RelatorioKpi(label="Rotas analisadas", value="0"),
                RelatorioKpi(label="Motoristas", value="0"),
                RelatorioKpi(label="Mortalidade kg", value="0.00"),
                RelatorioKpi(label="Destaque", value="-"),
            ],
            chart=[],
            columns=[
                col("pos", "POS", "number"),
                col("codigo_programacao", "PROGRAMACAO"),
                col("nf_numero", "NF"),
                col("motorista", "MOTORISTA"),
                col("mortalidade_total", "AVES", "number"),
                col("mortalidade_kg", "KG", "number"),
                col("clientes_com_mortalidade", "CLIENTES C/ OCORR.", "number"),
                col("data_ref", "DATA"),
                col("status_ref", "STATUS"),
            ],
            rows=[],
            text="\n".join(lines),
        )

    lines.extend(
        [
            "[RANKING POR ROTA - MAIOR MORTALIDADE]",
            "POS | PROGRAMACAO | NF | MOTORISTA | AVES | KG | TRANSBORDO | CLIENTES | DATA | STATUS",
            "-" * 95,
        ]
    )
    for row in rows:
        lines.append(
            f"{row['pos']:>3} | {row['codigo_programacao'] or '-'} | {row['nf_numero'] or '-'} | {row['motorista'] or '-'} | "
            f"{safe_int(row['mortalidade_total'], 0):>4} | {safe_float(row['mortalidade_kg'], 0):>8.2f} | "
            f"{safe_int(row['mortalidade_transbordo'], 0):>10} | {safe_int(row['clientes_com_mortalidade'], 0):>8} | "
            f"{str(row['data_ref'] or '-')[:19]} | {row['status_ref'] or '-'}"
        )

    resumo_mot: dict[str, dict[str, float | None]] = {}
    for row in rows:
        motorista = upper_text(row["motorista"] or "SEM MOTORISTA")
        item = resumo_mot.setdefault(motorista, {"rotas": 0, "mort_total": 0, "mort_kg": 0.0, "melhor": None, "pior": 0})
        mort = safe_int(row["mortalidade_total"], 0)
        item["rotas"] = safe_int(item["rotas"], 0) + 1
        item["mort_total"] = safe_int(item["mort_total"], 0) + mort
        item["mort_kg"] = safe_float(item["mort_kg"], 0.0) + safe_float(row["mortalidade_kg"], 0.0)
        item["pior"] = max(safe_int(item["pior"], 0), mort)
        melhor = item["melhor"]
        item["melhor"] = mort if melhor is None or mort < safe_int(melhor, 0) else melhor

    ranking_motorista = sorted(
        resumo_mot.items(),
        key=lambda item: (
            -safe_int(item[1]["mort_total"], 0),
            -safe_float(item[1]["mort_kg"], 0.0),
            item[0],
        ),
    )
    lines.extend(
        [
            "",
            "[CONSOLIDADO POR MOTORISTA]",
            "MOTORISTA | ROTAS | AVES | KG | MEDIA AVES/ROTA | MENOR ROTA | PIOR ROTA",
            "-" * 95,
        ]
    )
    for motorista, data in ranking_motorista:
        media = safe_int(data["mort_total"], 0) / max(safe_int(data["rotas"], 0), 1)
        lines.append(
            f"{motorista} | {safe_int(data['rotas'], 0)} | {safe_int(data['mort_total'], 0)} | {safe_float(data['mort_kg'], 0):.2f} | "
            f"{media:.2f} | {safe_int(data['melhor'], 0)} | {safe_int(data['pior'], 0)}"
        )

    destaque_mot, destaque_data = ranking_motorista[0]
    destaque_media = safe_int(destaque_data["mort_total"], 0) / max(safe_int(destaque_data["rotas"], 0), 1)
    lines.extend(
        [
            "",
            "[DESTAQUE]",
            f"Motorista com maior mortalidade acumulada: {destaque_mot} "
            f"({safe_int(destaque_data['mort_total'], 0)} aves / {safe_float(destaque_data['mort_kg'], 0):.2f} kg em {safe_int(destaque_data['rotas'], 0)} rota(s)).",
        ]
    )

    return RelatorioResumoResponse(
        tipo="Mortalidades",
        status=relatorio_status("Relatorio de mortalidades gerado", len(rows)),
        kpis=[
            RelatorioKpi(label="Rotas analisadas", value=str(len(rows))),
            RelatorioKpi(label="Motoristas", value=str(len(resumo_mot))),
            RelatorioKpi(label="Aves", value=str(sum(safe_int(row["mortalidade_total"], 0) for row in rows))),
            RelatorioKpi(label="KG", value=f"{sum(safe_float(row['mortalidade_kg'], 0.0) for row in rows):.2f}"),
            RelatorioKpi(label="Destaque", value=destaque_mot),
        ],
        chart=[
            RelatorioChartItem(
                label=motorista,
                value=number2(safe_int(data["mort_total"], 0)),
            )
            for motorista, data in ranking_motorista[:8]
        ],
        columns=[
            col("pos", "POS", "number"),
            col("codigo_programacao", "PROGRAMACAO"),
            col("nf_numero", "NF"),
            col("motorista", "MOTORISTA"),
            col("veiculo", "VEICULO"),
            col("mortalidade_total", "AVES", "number"),
            col("mortalidade_kg", "KG", "number"),
            col("mortalidade_transbordo", "TRANSB.", "number"),
            col("mortalidade_transbordo_kg", "KG TRANSB.", "number"),
            col("clientes_com_mortalidade", "CLIENTES C/ OCORR.", "number"),
            col("data_ref", "DATA"),
            col("status_ref", "STATUS"),
        ],
        rows=rows,
        text="\n".join(lines),
        sections=[
            RelatorioSection(
                title="Mortalidades por Cliente / Transbordo",
                columns=[
                    col("codigo_programacao", "PROGRAMACAO"), col("nf_numero", "NF"),
                    col("motorista", "MOTORISTA"), col("veiculo", "VEICULO"),
                    col("cod_cliente", "COD"), col("cliente", "CLIENTE"), col("pedido", "PEDIDO"),
                    col("mortalidade_aves", "AVES", "number"), col("mortalidade_kg", "KG", "number"),
                    col("media", "MEDIA", "number"), col("motivo", "MOTIVO"), col("foto", "FOTO"), col("data", "DATA"),
                ],
                rows=detalhe_rows,
            ),
            RelatorioSection(
                title="Fotos de Mortalidade / DOA",
                columns=[
                    col("codigo_programacao", "PROGRAMACAO"), col("nf_numero", "NF"),
                    col("motorista", "MOTORISTA"), col("categoria", "CATEGORIA"),
                    col("cliente", "CLIENTE"), col("arquivo", "ARQUIVO"), col("data", "DATA"),
                ],
                rows=foto_rows,
            ),
        ],
    )


async def build_report(
    db: AsyncSession,
    *,
    tipo: str,
    programacao: str = "",
    codigo_like: str = "",
    nf: str = "",
    motorista_like: str = "",
    data_like: str = "",
    show_recebimentos: bool = True,
    show_despesas: bool = True,
) -> RelatorioResumoResponse:
    tipo_key = normalize_key(tipo)
    if "NOTA FISCAL" in tipo_key or "TRANSBORDO" in tipo_key:
        return await build_nf_transbordo_report(db, nf or codigo_like)
    if "MORTALIDADE" in tipo_key or "OCORRENCIA" in tipo_key or "OCORRÊNCIA" in tipo_key:
        return await build_mortalidade_report(
            db,
            codigo_like=upper_text(codigo_like),
            motorista_like=upper_text(motorista_like),
            data_like=data_like,
        )
    if "ROTINA" in tipo_key:
        return await build_rotina_report(db, upper_text(motorista_like))
    if "KM DE VEICULOS" in tipo_key:
        return await build_km_report(
            db,
            codigo_like=upper_text(codigo_like),
            motorista_like=upper_text(motorista_like),
            data_like=data_like,
        )
    if "ABASTEC" in tipo_key or "BANHO" in tipo_key:
        return await build_despesas_eventos_report(
            db,
            tipo=tipo,
            codigo_like=upper_text(codigo_like),
            motorista_like=upper_text(motorista_like),
            data_like=data_like,
        )
    if tipo_key in {"DESPESAS", "CUSTOS E DESPESAS"}:
        return await build_despesas_report(db)

    codigo = upper_text(programacao)
    if not codigo:
        raise HTTPException(status_code=422, detail="Selecione um planejamento.")
    prog = await get_programacao_by_codigo(db, codigo)
    if not prog:
        raise HTTPException(status_code=404, detail=f"Planejamento nao encontrado: {codigo}")
    if "DETALHE COMPLETO" in tipo_key:
        return await build_detalhe_completo_report(db, prog)
    if "PRESTACAO" in tipo_key or "FECHAMENTO" in tipo_key:
        return await build_prestacao_report(
            db,
            prog,
            show_recebimentos=show_recebimentos,
            show_despesas=show_despesas,
        )
    return await build_programacao_report(db, prog)


@router.get("/options", response_model=RelatoriosOptions)
async def relatorios_options(current_user: User = Depends(require_admin_user)):
    return RelatoriosOptions(tipos=REPORT_TYPES)


@router.get("/programacoes", response_model=list[RelatorioProgramacaoOption])
async def relatorios_programacoes(
    tipo: str = "Planejamentos",
    codigo: str = "",
    nf: str = "",
    motorista: str = "",
    data: str = "",
    limit: int = 400,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    tipo_key = normalize_key(tipo)
    nf_norm = upper_text(nf)
    nf_relacionados: set[str] = set()
    if nf_norm and ("NOTA FISCAL" in tipo_key or "TRANSBORDO" in tipo_key):
        codigos, _ = await codigos_relacionados_nf(db, nf_norm)
        nf_relacionados = set(codigos)
    rows = []
    for programacao in await programacao_rows(db, limit=max(limit, 1)):
        codigo_prog = upper_text(programacao.codigo_programacao)
        if nf_norm:
            if nf_relacionados:
                if codigo_prog not in nf_relacionados:
                    continue
            elif nf_norm not in nf_programacao(programacao):
                continue
        if not matches_programacao(
            programacao,
            tipo=tipo,
            codigo_like=upper_text(codigo),
            motorista_like=upper_text(motorista),
            data_like=data,
        ):
            continue
        rows.append(
            RelatorioProgramacaoOption(
                codigo_programacao=codigo_prog,
                motorista=upper_text(programacao.motorista),
                veiculo=upper_text(programacao.veiculo),
                nf_numero=nf_programacao(programacao),
                data_ref=data_ref(programacao),
                status=status_ref(programacao),
                prestacao_status=upper_text(programacao.prestacao_status or "PENDENTE"),
                tipo_estimativa=upper_text(programacao.tipo_estimativa or "KG"),
                operacao_tipo="TRANSBORDO" if is_transbordo_programacao(programacao) else "VENDA",
                transbordo_modalidade=upper_text(getattr(programacao, "transbordo_modalidade", "") or ""),
                transbordo_grupo=upper_text(getattr(programacao, "transbordo_grupo", "") or ""),
            )
        )
    return rows[: max(min(limit, 1000), 1)]


@router.get("/resumo", response_model=RelatorioResumoResponse)
async def relatorios_resumo(
    tipo: str = "Planejamentos",
    programacao: str = "",
    codigo: str = "",
    nf: str = "",
    motorista: str = "",
    data: str = "",
    show_recebimentos: bool = True,
    show_despesas: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    return await build_report(
        db,
        tipo=tipo,
        programacao=programacao,
        codigo_like=codigo,
        nf=nf,
        motorista_like=motorista,
        data_like=data,
        show_recebimentos=show_recebimentos,
        show_despesas=show_despesas,
    )


def serialize_model_row(item: Any, fields: list[str]) -> dict[str, Any]:
    return {field: getattr(item, field, None) for field in fields}


def add_sheet(workbook: Any, title: str, rows: list[dict[str, Any]]) -> None:
    safe_title = re.sub(r"[\[\]:*?/\\]", "_", title or "RELATORIO")[:31] or "RELATORIO"
    ws = workbook.create_sheet(title=safe_title)
    if not rows:
        ws.append(["sem_dados"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 46)


def relatorio_response_rows(report: RelatorioResumoResponse) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in report.rows:
        out: dict[str, Any] = {}
        for column in report.columns:
            out[column.label or column.key] = row.get(column.key)
        rows.append(out)
    return rows


@router.get("/exportar-excel")
async def relatorios_exportar_excel_generico(
    tipo: str = "Planejamentos",
    programacao: str = "",
    codigo: str = "",
    nf: str = "",
    motorista: str = "",
    data: str = "",
    show_recebimentos: bool = True,
    show_despesas: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - depends on optional package
        raise HTTPException(status_code=503, detail="Biblioteca openpyxl indisponivel.") from exc

    report = await build_report(
        db,
        tipo=tipo,
        programacao=programacao,
        codigo_like=codigo,
        nf=nf,
        motorista_like=motorista,
        data_like=data,
        show_recebimentos=show_recebimentos,
        show_despesas=show_despesas,
    )

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(
        wb,
        "RESUMO",
        [
            {"indicador": item.label, "valor": item.value}
            for item in report.kpis
        ],
    )
    add_sheet(wb, "DADOS", relatorio_response_rows(report))
    for section in report.sections:
        section_rows = []
        for row in section.rows:
            out: dict[str, Any] = {}
            for column in section.columns:
                out[column.label or column.key] = row.get(column.key)
            section_rows.append(out)
        add_sheet(wb, section.title or "DETALHES", section_rows)
    add_sheet(wb, "MEMORIAL", [{"linha": line} for line in (report.text or "").splitlines()])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    safe_name = upper_text(report.programacao) or normalize_key(report.tipo).replace(" ", "_").replace("/", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="RELATORIO_{safe_name}.xlsx"'},
    )


@router.get("/{codigo_programacao}/exportar-excel")
async def relatorios_exportar_excel(
    codigo_programacao: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - depends on optional package
        raise HTTPException(status_code=503, detail="Biblioteca openpyxl indisponivel.") from exc

    programacao = await get_programacao_by_codigo(db, codigo_programacao)
    if not programacao:
        raise HTTPException(status_code=404, detail="Planejamento nao encontrado")
    codigo = upper_text(programacao.codigo_programacao)

    itens = await itens_for(db, codigo)
    recebimentos = await recebimentos_for(db, codigo)
    despesas = await despesas_for(db, codigo)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(
        wb,
        "PROGRAMACAO",
        [
            serialize_model_row(
                programacao,
                [
                    "codigo_programacao",
                    "data_criacao",
                    "motorista",
                    "veiculo",
                    "equipe",
                    "kg_estimado",
                    "status",
                    "status_operacional",
                    "prestacao_status",
                    "local_rota",
                    "local_carregamento",
                    "nf_numero",
                    "data_saida",
                    "hora_saida",
                    "data_chegada",
                    "hora_chegada",
                    "km_rodado",
                    "custo_km",
                ],
            )
        ],
    )
    add_sheet(
        wb,
        "ITENS",
        [
            serialize_model_row(
                item,
                ["cod_cliente", "nome_cliente", "qnt_caixas", "kg", "preco", "endereco", "vendedor", "pedido", "produto", "observacao"],
            )
            for item in itens
        ],
    )
    add_sheet(
        wb,
        "RECEBIMENTOS",
        [
            serialize_model_row(item, ["cod_cliente", "nome_cliente", "valor", "forma_pagamento", "observacao", "num_nf", "data_registro"])
            for item in recebimentos
        ],
    )
    add_sheet(
        wb,
        "CUSTOS",
        [
            serialize_model_row(item, ["descricao", "valor", "data_registro", "tipo_despesa", "categoria", "motorista", "veiculo", "observacao"])
            for item in despesas
        ],
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"RELATORIO_{codigo}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf")
async def relatorios_pdf(
    tipo: str = "Planejamentos",
    programacao: str = "",
    codigo: str = "",
    nf: str = "",
    motorista: str = "",
    data: str = "",
    show_recebimentos: bool = True,
    show_despesas: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception as exc:  # pragma: no cover - depends on optional package
        raise HTTPException(status_code=503, detail="Biblioteca ReportLab indisponivel.") from exc

    report = await build_report(
        db,
        tipo=tipo,
        programacao=programacao,
        codigo_like=codigo,
        nf=nf,
        motorista_like=motorista,
        data_like=data,
        show_recebimentos=show_recebimentos,
        show_despesas=show_despesas,
    )

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 56
    pdf.setFont("Helvetica-Bold", 13)
    title = f"RELATORIO - {upper_text(report.programacao) or normalize_key(report.tipo)}"
    pdf.drawString(40, y, title[:100])
    y -= 24
    pdf.setFont("Helvetica", 9)
    for line in report.text.splitlines():
        pdf.drawString(40, y, line[:126])
        y -= 13
        if y < 52:
            pdf.showPage()
            y = height - 56
            pdf.setFont("Helvetica", 9)
    pdf.save()
    buffer.seek(0)
    safe_name = upper_text(report.programacao) or normalize_key(report.tipo).replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="RELATORIO_{safe_name}.pdf"'},
    )


@router.post("/{codigo_programacao}/finalizar-rota", response_model=RelatorioProgramacaoOption)
async def finalizar_rota_relatorio(
    codigo_programacao: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    programacao = await get_programacao_by_codigo(db, codigo_programacao)
    if not programacao:
        raise HTTPException(status_code=404, detail="Planejamento nao encontrado")
    if upper_text(programacao.prestacao_status or "PENDENTE") == "FECHADA":
        raise HTTPException(status_code=409, detail="A rota esta com a prestacao FECHADA.")
    programacao.status = "FINALIZADA"
    programacao.status_operacional = "FINALIZADA"
    programacao.finalizada_no_app = 1
    record_audit_log(
        db,
        action="relatorios_rota_finalizada",
        actor_user=current_user,
        entity_type="programacao",
        entity_id=upper_text(programacao.codigo_programacao),
        severity="warning",
        ip_address=client_ip_from_request(request),
        metadata={"status": "FINALIZADA", "finalizada_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    )
    await db.commit()
    await db.refresh(programacao)
    return RelatorioProgramacaoOption(
        codigo_programacao=upper_text(programacao.codigo_programacao),
        motorista=upper_text(programacao.motorista),
        veiculo=upper_text(programacao.veiculo),
        data_ref=data_ref(programacao),
        status=status_ref(programacao),
        prestacao_status=upper_text(programacao.prestacao_status or "PENDENTE"),
    )


@router.post("/{codigo_programacao}/reabrir-rota", response_model=RelatorioProgramacaoOption)
async def reabrir_rota_relatorio(
    codigo_programacao: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    programacao = await get_programacao_by_codigo(db, codigo_programacao)
    if not programacao:
        raise HTTPException(status_code=404, detail="Planejamento nao encontrado")
    if upper_text(programacao.prestacao_status or "PENDENTE") == "FECHADA":
        raise HTTPException(status_code=409, detail="Nao e permitido reabrir rota com prestacao FECHADA.")
    programacao.status = "ATIVA"
    programacao.status_operacional = ""
    programacao.finalizada_no_app = 0
    record_audit_log(
        db,
        action="relatorios_rota_reaberta",
        actor_user=current_user,
        entity_type="programacao",
        entity_id=upper_text(programacao.codigo_programacao),
        ip_address=client_ip_from_request(request),
        metadata={"status": "ATIVA", "reaberta_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    )
    await db.commit()
    await db.refresh(programacao)
    return RelatorioProgramacaoOption(
        codigo_programacao=upper_text(programacao.codigo_programacao),
        motorista=upper_text(programacao.motorista),
        veiculo=upper_text(programacao.veiculo),
        data_ref=data_ref(programacao),
        status=status_ref(programacao),
        prestacao_status=upper_text(programacao.prestacao_status or "PENDENTE"),
    )
