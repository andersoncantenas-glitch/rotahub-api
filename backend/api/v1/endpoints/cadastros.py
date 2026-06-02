# backend/api/v1/endpoints/cadastros.py
"""
Operational cadastro endpoints.

This router intentionally mirrors the desktop CadastroCRUD behavior instead of
acting as a generic table editor. The browser must obey the same required
fields, normalizations, status values, password rules, and delete guards used by
the Tkinter system.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, List
from uuid import uuid4

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, create_model
from sqlalchemy import func, select, text
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.security.passwords import hash_password_pbkdf2
from app.utils.excel_helpers import guess_col
from app.utils.formatters import safe_float, safe_int
from app.utils.validators import (
    is_valid_cpf,
    is_valid_motorista_codigo,
    is_valid_motorista_senha,
    is_valid_phone,
    normalize_cpf,
    normalize_phone,
)
from backend.api.v1.endpoints.users import require_admin_user
from backend.config.database import get_db
from backend.models.cadastro import AjudanteDB, CaixaDB, CaixaMovimentoDB, ClienteDB, FornecedorDB, MotoristaDB, ProdutoDB, VeiculoDB, VendedorDB
from backend.models.user import User, UserDB
from backend.services.audit import client_ip_from_request, record_audit_log
from backend.services.auth import get_password_hash

router = APIRouter()

FORNECEDOR_PERFIS = {
    "FRANGO_VIVO",
    "PRESTADOR_SERVICO",
    "PNEUS",
    "OLEO_LUBRIFICANTES",
    "PECAS",
    "MECANICO",
    "BORRACHEIRO",
    "LAVADOR_CAIXAS",
    "MANUTENCAO",
    "COMBUSTIVEL",
    "SERVICO_SEM_NF",
    "OUTROS",
}

PRODUTO_CATEGORIAS = {"AVES", "INSUMOS", "EMBALAGENS", "SERVICOS", "OUTROS"}
PRODUTO_UNIDADES = {"KG", "CX", "UN", "PC", "LT"}
CAIXA_STATUS = {"EM_ESTOQUE", "VINCULADA", "EM_USO", "QUEBRADA", "BAIXADA"}


class CadastroResource:
    def __init__(
        self,
        *,
        model,
        fields: tuple[str, ...],
        required: tuple[str, ...],
        order_by: str,
        status_values: tuple[str, ...] = (),
        password_field: str | None = None,
    ):
        self.model = model
        self.fields = fields
        self.required = required
        self.order_by = order_by
        self.status_values = status_values
        self.password_field = password_field


RESOURCES: dict[str, CadastroResource] = {
    "motoristas": CadastroResource(
        model=MotoristaDB,
        fields=("nome", "codigo", "senha", "perfil_app", "cpf", "telefone", "status"),
        required=("nome", "telefone"),
        order_by="id",
        status_values=("ATIVO", "INATIVO"),
        password_field="senha",
    ),
    "vendedores": CadastroResource(
        model=VendedorDB,
        fields=("codigo", "nome", "senha", "telefone", "cidade_base", "status"),
        required=("codigo", "nome", "status"),
        order_by="id",
        status_values=("ATIVO", "DESATIVADO"),
        password_field="senha",
    ),
    "usuarios": CadastroResource(
        model=UserDB,
        fields=("nome", "senha", "permissoes", "cpf", "telefone"),
        required=("nome", "permissoes"),
        order_by="id",
        password_field="senha",
    ),
    "veiculos": CadastroResource(
        model=VeiculoDB,
        fields=("placa", "modelo", "capacidade_cx", "status"),
        required=("placa", "modelo", "capacidade_cx"),
        order_by="id",
        status_values=("ATIVO", "DESATIVADO"),
    ),
    "caixas": CadastroResource(
        model=CaixaDB,
        fields=("codigo", "lote", "cor", "veiculo_placa", "status", "data_compra", "observacao"),
        required=("codigo", "lote", "cor", "status"),
        order_by="id",
        status_values=("EM_ESTOQUE", "VINCULADA", "EM_USO", "QUEBRADA", "BAIXADA"),
    ),
    "ajudantes": CadastroResource(
        model=AjudanteDB,
        fields=("nome", "sobrenome", "telefone", "status"),
        required=("nome", "sobrenome", "telefone", "status"),
        order_by="id",
        status_values=("ATIVO", "DESATIVADO"),
    ),
    "clientes": CadastroResource(
        model=ClienteDB,
        fields=("cod_cliente", "nome_cliente", "endereco", "bairro", "cidade", "uf", "telefone", "rota", "vendedor"),
        required=("cod_cliente", "nome_cliente"),
        order_by="id",
    ),
    "fornecedores": CadastroResource(
        model=FornecedorDB,
        fields=(
            "razao_social",
            "nome_fantasia",
            "documento",
            "tipo_pessoa",
            "perfil_fornecedor",
            "telefone",
            "email",
            "cidade",
            "uf",
            "status",
            "observacao",
            "certificado_status",
            "certificado_nome",
            "certificado_instalado_em",
        ),
        required=("razao_social", "documento", "perfil_fornecedor", "status"),
        order_by="id",
        status_values=("ATIVO", "INATIVO"),
    ),
    "produtos": CadastroResource(
        model=ProdutoDB,
        fields=(
            "codigo",
            "nome",
            "descricao",
            "categoria",
            "unidade",
            "unidade_estoque",
            "controla_estoque_fisico",
            "controla_estoque_fiscal",
            "estoque_min_kg",
            "estoque_min_caixas",
            "ncm",
            "cest",
            "cfop_entrada",
            "cfop_saida",
            "ean",
            "custo_padrao",
            "preco_padrao",
            "status",
        ),
        required=("codigo", "nome", "categoria", "unidade", "unidade_estoque", "status"),
        order_by="id",
        status_values=("ATIVO", "INATIVO"),
    ),
}


CadastroPayload = create_model(
    "CadastroPayload",
    __base__=BaseModel,
    nome=(str | None, Field(default=None, max_length=180)),
    codigo=(str | None, Field(default=None, max_length=80)),
    senha=(str | None, Field(default=None, max_length=128)),
    perfil_app=(str | None, Field(default=None, max_length=40)),
    cpf=(str | None, Field(default=None, max_length=20)),
    telefone=(str | None, Field(default=None, max_length=40)),
    status=(str | None, Field(default=None, max_length=40)),
    cidade_base=(str | None, Field(default=None, max_length=120)),
    placa=(str | None, Field(default=None, max_length=20)),
    modelo=(str | None, Field(default=None, max_length=120)),
    capacidade_cx=(int | None, Field(default=None, ge=0)),
    lote=(str | None, Field(default=None, max_length=80)),
    cor=(str | None, Field(default=None, max_length=60)),
    veiculo_placa=(str | None, Field(default=None, max_length=20)),
    data_compra=(str | None, Field(default=None, max_length=30)),
    sobrenome=(str | None, Field(default=None, max_length=120)),
    permissoes=(str | None, Field(default=None, max_length=40)),
    cod_cliente=(str | None, Field(default=None, max_length=80)),
    nome_cliente=(str | None, Field(default=None, max_length=180)),
    endereco=(str | None, Field(default=None, max_length=220)),
    bairro=(str | None, Field(default=None, max_length=120)),
    cidade=(str | None, Field(default=None, max_length=120)),
    uf=(str | None, Field(default=None, max_length=2)),
    rota=(str | None, Field(default=None, max_length=120)),
    vendedor=(str | None, Field(default=None, max_length=160)),
    razao_social=(str | None, Field(default=None, max_length=180)),
    nome_fantasia=(str | None, Field(default=None, max_length=180)),
    documento=(str | None, Field(default=None, max_length=20)),
    tipo_pessoa=(str | None, Field(default=None, max_length=10)),
    perfil_fornecedor=(str | None, Field(default=None, max_length=40)),
    email=(str | None, Field(default=None, max_length=180)),
    observacao=(str | None, Field(default=None, max_length=300)),
    certificado_status=(str | None, Field(default=None, max_length=40)),
    certificado_nome=(str | None, Field(default=None, max_length=180)),
    certificado_instalado_em=(str | None, Field(default=None, max_length=30)),
    descricao=(str | None, Field(default=None, max_length=300)),
    categoria=(str | None, Field(default=None, max_length=40)),
    unidade=(str | None, Field(default=None, max_length=12)),
    unidade_estoque=(str | None, Field(default=None, max_length=12)),
    controla_estoque_fisico=(int | None, Field(default=None, ge=0, le=1)),
    controla_estoque_fiscal=(int | None, Field(default=None, ge=0, le=1)),
    estoque_min_kg=(float | None, Field(default=None, ge=0)),
    estoque_min_caixas=(int | None, Field(default=None, ge=0)),
    ncm=(str | None, Field(default=None, max_length=20)),
    cest=(str | None, Field(default=None, max_length=20)),
    cfop_entrada=(str | None, Field(default=None, max_length=10)),
    cfop_saida=(str | None, Field(default=None, max_length=10)),
    ean=(str | None, Field(default=None, max_length=30)),
    custo_padrao=(float | None, Field(default=None, ge=0)),
    preco_padrao=(float | None, Field(default=None, ge=0)),
)


class CadastroPasswordPayload(BaseModel):
    nova_senha: str = Field(min_length=1, max_length=128)


class FornecedorPerfilPayload(BaseModel):
    codigo: str = Field(min_length=2, max_length=40)
    nome: str = Field(min_length=2, max_length=120)
    categoria: str | None = Field(default="OUTROS", max_length=40)
    status: str | None = Field(default="ATIVO", max_length=20)


class ClienteImportRow(BaseModel):
    id: int | None = None
    cod_cliente: str | None = None
    nome_cliente: str | None = None
    endereco: str | None = None
    telefone: str | None = None
    vendedor: str | None = None


class ClientesImportPayload(BaseModel):
    rows: list[ClienteImportRow] = Field(default_factory=list)


class CaixasBulkPayload(BaseModel):
    prefixo: str = Field(default="CX", min_length=1, max_length=40)
    lote: str = Field(min_length=1, max_length=80)
    cor: str = Field(min_length=1, max_length=60)
    quantidade: int = Field(ge=1, le=5000)
    numero_inicial: int = Field(default=1, ge=1)
    digitos: int = Field(default=4, ge=1, le=8)
    veiculo_placa: str | None = Field(default=None, max_length=20)
    status: str | None = Field(default="EM_ESTOQUE", max_length=40)
    data_compra: str | None = Field(default=None, max_length=30)
    observacao: str | None = Field(default=None, max_length=300)


class CaixasBulkResult(BaseModel):
    criadas: int = 0
    primeiro_codigo: str = ""
    ultimo_codigo: str = ""


class CaixasMovimentarPayload(BaseModel):
    quantidade: int = Field(ge=1, le=5000)
    lote: str | None = Field(default=None, max_length=80)
    cor: str | None = Field(default=None, max_length=60)
    veiculo_origem: str | None = Field(default=None, max_length=20)
    status_origem: str | None = Field(default="TODOS", max_length=40)
    veiculo_destino: str | None = Field(default=None, max_length=20)
    status_destino: str = Field(default="VINCULADA", max_length=40)
    observacao: str | None = Field(default=None, max_length=300)


class CaixasMovimentarResult(BaseModel):
    movimentadas: int = 0
    codigos: list[str] = Field(default_factory=list)


class ClientesImportResult(BaseModel):
    total: int = 0
    inseridos: int = 0
    atualizados: int = 0
    ignorados: int = 0


class ClienteDashboardResponse(BaseModel):
    total_clientes: int = 0
    clientes_com_historico: int = 0
    amostras_localizacao: int = 0
    clientes_com_localizacao: int = 0


class ClienteLookupResponse(BaseModel):
    cod_cliente: str
    nome_cliente: str


class ClienteHistoricoResponse(BaseModel):
    resumo: dict[str, Any]
    rows: list[dict[str, Any]]


class ClienteLocalizacoesResponse(BaseModel):
    resumo: dict[str, Any]
    rows: list[dict[str, Any]]


class CadastroItemResponse(BaseModel):
    id: int
    data: dict[str, Any]


def get_resource(resource: str) -> CadastroResource:
    config = RESOURCES.get(resource)
    if config is None:
        raise HTTPException(status_code=404, detail="Cadastro nao encontrado")
    return config


def upper_text(value: Any) -> str:
    return str(value or "").strip().upper()


async def ensure_fornecedor_perfis(db: AsyncSession) -> None:
    await db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS fornecedor_perfis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL UNIQUE,
                nome TEXT NOT NULL,
                categoria TEXT DEFAULT 'OUTROS',
                status TEXT DEFAULT 'ATIVO',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    )
    for codigo, nome, categoria in (
        ("FRANGO_VIVO", "Frango vivo", "FRANGO"),
        ("PRESTADOR_SERVICO", "Prestador de servico", "SERVICO"),
        ("PECAS", "Pecas", "MANUTENCAO"),
        ("MECANICO", "Mecanico", "MANUTENCAO"),
        ("BORRACHEIRO", "Borracheiro", "MANUTENCAO"),
        ("LAVADOR_CAIXAS", "Lavador de caixas", "OPERACIONAL"),
        ("PNEUS", "Pneus", "MANUTENCAO"),
        ("OLEO_LUBRIFICANTES", "Oleo e lubrificantes", "MANUTENCAO"),
        ("COMBUSTIVEL", "Combustivel", "OPERACIONAL"),
        ("SERVICO_SEM_NF", "Servico sem NF", "SERVICO"),
        ("OUTROS", "Outros", "OUTROS"),
    ):
        await db.execute(
            text(
                """
                INSERT OR IGNORE INTO fornecedor_perfis (codigo, nome, categoria, status)
                VALUES (:codigo, :nome, :categoria, 'ATIVO')
                """
            ),
            {"codigo": codigo, "nome": nome, "categoria": categoria},
        )


async def fornecedor_perfis_ativos(db: AsyncSession) -> set[str]:
    await ensure_fornecedor_perfis(db)
    result = await db.execute(
        text(
            """
            SELECT codigo
              FROM fornecedor_perfis
             WHERE UPPER(TRIM(COALESCE(status, 'ATIVO')))='ATIVO'
            """
        )
    )
    return {upper_text(row[0]) for row in result.all()} | FORNECEDOR_PERFIS


async def ensure_caixas_movimentos(db: AsyncSession) -> None:
    await db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS caixas_movimentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                caixa_id INTEGER,
                codigo TEXT,
                movimento TEXT,
                veiculo_origem TEXT,
                veiculo_destino TEXT,
                status_origem TEXT,
                status_destino TEXT,
                observacao TEXT,
                criado_em TEXT,
                company_id INTEGER
            )
            """
        )
    )
    for column, definition in {
        "caixa_id": "INTEGER",
        "codigo": "TEXT",
        "movimento": "TEXT",
        "veiculo_origem": "TEXT",
        "veiculo_destino": "TEXT",
        "status_origem": "TEXT",
        "status_destino": "TEXT",
        "observacao": "TEXT",
        "criado_em": "TEXT",
        "company_id": "INTEGER",
    }.items():
        if column not in await table_columns(db, "caixas_movimentos"):
            await db.execute(text(f"ALTER TABLE caixas_movimentos ADD COLUMN {column} {definition}"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_caixas_movimentos_caixa ON caixas_movimentos(caixa_id, criado_em)"))
    await db.execute(text("CREATE INDEX IF NOT EXISTS idx_caixas_movimentos_codigo ON caixas_movimentos(codigo, criado_em)"))


async def plan_limit(db: AsyncSession, company_id: int, field: str) -> int | None:
    if field not in {"vehicle_limit", "user_limit"}:
        return None
    result = await db.execute(
        text(
            f"""
            SELECT p.{field}
              FROM subscriptions s
              JOIN plans p ON p.id=s.plan_id
             WHERE s.company_id=:company_id
               AND s.status IN ('active', 'trialing', 'past_due')
             ORDER BY s.id DESC
             LIMIT 1
            """
        ),
        {"company_id": int(company_id or 0)},
    )
    value = result.scalar_one_or_none()
    return None if value is None else safe_int(value, 0)


async def enforce_create_limit(resource: str, db: AsyncSession, current_user: User) -> None:
    company_id = int(current_user.company_id or 1)
    if resource == "veiculos":
        limit = await plan_limit(db, company_id, "vehicle_limit")
        if limit is None:
            return
        count = safe_int(
            (
                await db.execute(
                    text("SELECT COUNT(*) FROM veiculos WHERE company_id=:company_id"),
                    {"company_id": company_id},
                )
            ).scalar(),
            0,
        )
        if count >= int(limit):
            raise HTTPException(
                status_code=403,
                detail=f"Limite de {int(limit)} veiculos atingido no plano atual. Faça upgrade para cadastrar mais veículos.",
            )
    elif resource == "usuarios":
        limit = await plan_limit(db, company_id, "user_limit")
        if limit is None:
            return
        count = safe_int(
            (
                await db.execute(
                    text("SELECT COUNT(*) FROM usuarios WHERE company_id=:company_id AND COALESCE(is_active, 1)=1"),
                    {"company_id": company_id},
                )
            ).scalar(),
            0,
        )
        if count >= int(limit):
            raise HTTPException(
                status_code=403,
                detail=f"Limite de {int(limit)} usuarios atingido no plano atual. Faça upgrade para cadastrar mais usuários.",
            )


async def record_caixa_movimento(
    db: AsyncSession,
    *,
    caixa: CaixaDB,
    movimento: str,
    veiculo_origem: str | None = None,
    status_origem: str | None = None,
    observacao: str | None = None,
) -> None:
    await ensure_caixas_movimentos(db)
    db.add(
        CaixaMovimentoDB(
            caixa_id=caixa.id,
            codigo=upper_text(caixa.codigo),
            movimento=upper_text(movimento),
            veiculo_origem=upper_text(veiculo_origem),
            veiculo_destino=upper_text(caixa.veiculo_placa),
            status_origem=upper_text(status_origem),
            status_destino=upper_text(caixa.status),
            observacao=upper_text(observacao or caixa.observacao)[:300] or None,
            criado_em=datetime.now().isoformat(timespec="seconds"),
        )
    )


def clean_base_payload(payload: BaseModel, config: CadastroResource, *, partial: bool) -> dict[str, Any]:
    raw = payload.model_dump(exclude_unset=partial)
    data = {field: raw.get(field) for field in config.fields if field in raw}
    for field, value in list(data.items()):
        if isinstance(value, str):
            value = value.strip()
            data[field] = value or None

    missing = [field for field in config.required if not data.get(field)] if not partial else []
    if missing:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Preencha o campo: {missing[0].upper()}",
        )
    return data


def serialize_item(item: Any, config: CadastroResource) -> CadastroItemResponse:
    data = {field: getattr(item, field, None) for field in config.fields}
    if config.password_field and data.get(config.password_field):
        data[config.password_field] = "******"
    return CadastroItemResponse(id=item.id, data=data)


def current_company_id(current_user: User) -> int:
    try:
        return int(getattr(current_user, "company_id", None) or 1)
    except Exception:
        return 1


def assign_company_id(item: Any, current_user: User) -> None:
    if hasattr(item, "company_id") and not getattr(item, "company_id", None):
        setattr(item, "company_id", current_company_id(current_user))


def cadastro_integrity_message(resource: str, exc: Exception) -> str:
    detail = str(getattr(exc, "orig", None) or exc)
    lower = detail.lower()
    label = resource.rstrip("s")
    if "unique" in lower or "constraint" in lower:
        if "codigo" in lower:
            return f"Ja existe {label} com este codigo."
        if "placa" in lower:
            return "Ja existe veiculo com esta placa."
        if "cpf" in lower:
            return f"Ja existe {label} com este CPF."
        if "telefone" in lower:
            return f"Ja existe {label} com este telefone."
        return f"Cadastro duplicado em {resource}. Atualize a lista e tente novamente."
    if "no such column" in lower or "has no column" in lower or "no such table" in lower:
        return "Banco do servidor precisa atualizar a estrutura de cadastros. Recarregue o sistema e tente novamente."
    return "Nao foi possivel salvar o cadastro. Verifique os dados e tente novamente."


async def get_item_or_404(db: AsyncSession, config: CadastroResource, item_id: int):
    item = await db.get(config.model, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Cadastro nao encontrado")
    return item


async def table_columns(db: AsyncSession, table_name: str) -> set[str]:
    try:
        result = await db.execute(text(f"PRAGMA table_info({table_name})"))
        return {str(row[1]).lower() for row in result.fetchall()}
    except Exception:
        return set()


async def count_sql(db: AsyncSession, sql: str, params: dict[str, Any]) -> int:
    try:
        result = await db.execute(text(sql), params)
        return int((result.scalar() or 0))
    except Exception:
        return 0


def excel_cell_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text_value = str(value).strip()
    return "" if text_value.upper() in {"NAN", "NAT", "NONE", "NULL", "<NA>"} else text_value


def normalize_cliente_import_row(row: ClienteImportRow, *, require_required: bool = True) -> dict[str, Any] | None:
    data = {
        "cod_cliente": upper_text(excel_cell_text(row.cod_cliente)),
        "nome_cliente": upper_text(excel_cell_text(row.nome_cliente)),
        "endereco": upper_text(excel_cell_text(row.endereco)),
        "telefone": upper_text(excel_cell_text(row.telefone)),
        "vendedor": upper_text(excel_cell_text(row.vendedor)),
    }
    data["nome"] = data["nome_cliente"]
    if not any(data.values()):
        return None
    if require_required and (not data["cod_cliente"] or not data["nome_cliente"]):
        raise HTTPException(
            status_code=422,
            detail="Todas as linhas precisam ter pelo menos COD CLIENTE e NOME CLIENTE.",
        )
    if not require_required and not data["cod_cliente"]:
        return None
    return data


def cliente_import_response(item: ClienteDB) -> dict[str, Any]:
    return {
        "id": item.id,
        "cod_cliente": item.cod_cliente or "",
        "nome_cliente": item.nome_cliente or "",
        "endereco": item.endereco or "",
        "telefone": item.telefone or "",
        "vendedor": item.vendedor or "",
    }


def clientes_rows_from_dataframe(df: pd.DataFrame) -> list[ClienteImportRow]:
    columns = list(df.columns)
    col_cod = guess_col(columns, ["cod cliente", "codigo cliente", "cliente codigo", "codigo", "cod", "cód", "cliente"])
    col_nome = guess_col(
        columns,
        ["nome cliente", "nome completo", "razao social", "razao", "fantasia", "nome fantasia", "nome", "cliente"],
    )
    col_end = guess_col(columns, ["endereco", "endereço", "rua", "logradouro"])
    col_tel = guess_col(columns, ["telefone", "fone", "celular", "contato", "tel"])
    col_vendedor = guess_col(columns, ["vendedor", "vend", "representante", "rca"])

    if col_cod and col_nome == col_cod:
        col_nome = next((column for column in columns if column != col_cod), col_nome)

    if not col_cod or not col_nome:
        cols = columns
        if len(cols) >= 2:
            col_cod = col_cod or cols[0]
            col_nome = col_nome or cols[1]
        else:
            raise HTTPException(
                status_code=422,
                detail="Nao identifiquei as colunas de codigo e nome do cliente no Excel.",
            )

    rows: list[ClienteImportRow] = []
    for _, source in df.iterrows():
        rows.append(
            ClienteImportRow(
                cod_cliente=excel_cell_text(source.get(col_cod, "")),
                nome_cliente=excel_cell_text(source.get(col_nome, "")),
                endereco=excel_cell_text(source.get(col_end, "")) if col_end else "",
                telefone=excel_cell_text(source.get(col_tel, "")) if col_tel else "",
                vendedor=excel_cell_text(source.get(col_vendedor, "")) if col_vendedor else "",
            )
        )
    return rows


async def bulk_upsert_clientes(
    db: AsyncSession,
    rows: list[ClienteImportRow],
    *,
    merge_duplicate_rows: bool = False,
) -> ClientesImportResult:
    normalized_by_code: dict[str, dict[str, Any]] = {}
    ordered_codes: list[str] = []
    ignored = 0
    for row in rows:
        data = normalize_cliente_import_row(row, require_required=not merge_duplicate_rows)
        if data is None:
            ignored += 1
            continue
        key = data["cod_cliente"]
        if key in normalized_by_code:
            if not merge_duplicate_rows:
                raise HTTPException(
                    status_code=422,
                    detail=f"COD CLIENTE duplicado na tabela: {key}",
                )
            current = normalized_by_code[key]
            for field, value in data.items():
                if field == "cod_cliente":
                    continue
                if value:
                    current[field] = value
            ignored += 1
            continue
        normalized_by_code[key] = data
        ordered_codes.append(key)

    result = ClientesImportResult(ignorados=ignored)
    if not ordered_codes:
        return result

    seen = set(ordered_codes)
    existing_result = await db.execute(
        select(ClienteDB).where(func.upper(func.coalesce(ClienteDB.cod_cliente, "")).in_(seen))
    )
    existing_by_code = {upper_text(item.cod_cliente): item for item in existing_result.scalars().all()}

    for key in ordered_codes:
        data = normalized_by_code[key]
        item = existing_by_code.get(data["cod_cliente"])
        if item is None:
            if not data["nome_cliente"]:
                result.ignorados += 1
                continue
            item = ClienteDB(**data)
            db.add(item)
            result.inseridos += 1
            continue
        changed = False
        for field, value in data.items():
            if merge_duplicate_rows and field != "cod_cliente" and not value:
                continue
            if getattr(item, field, None) != value:
                setattr(item, field, value)
                changed = True
        result.atualizados += 1 if changed else 0
        result.ignorados += 0 if changed else 1

    result.total = result.inseridos + result.atualizados
    return result


async def exists_ci(
    db: AsyncSession,
    model,
    field: str,
    value: str,
    *,
    exclude_id: int | None = None,
) -> bool:
    value_norm = upper_text(value)
    if not value_norm:
        return False
    column = getattr(model, field)
    stmt = select(model.id).where(func.upper(func.coalesce(column, "")) == value_norm)
    if exclude_id:
        stmt = stmt.where(model.id != exclude_id)
    result = await db.execute(stmt.limit(1))
    return result.scalar_one_or_none() is not None


async def exists_exact(
    db: AsyncSession,
    model,
    field: str,
    value: str,
    *,
    exclude_id: int | None = None,
) -> bool:
    if not value:
        return False
    column = getattr(model, field)
    stmt = select(model.id).where(column == value)
    if exclude_id:
        stmt = stmt.where(model.id != exclude_id)
    result = await db.execute(stmt.limit(1))
    return result.scalar_one_or_none() is not None


async def next_motorista_codigo(db: AsyncSession) -> str:
    result = await db.execute(select(MotoristaDB.codigo).where(MotoristaDB.codigo.is_not(None)))
    max_seq = 0
    for codigo in result.scalars().all():
        text_codigo = upper_text(codigo)
        if not text_codigo.startswith("MOT-"):
            continue
        try:
            max_seq = max(max_seq, int(text_codigo.split("-", 1)[1]))
        except Exception:
            continue
    return f"MOT-{max(1, max_seq + 1):02d}"


@router.get("/motoristas/proximo-codigo")
async def get_next_motorista_codigo(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    return {"codigo": await next_motorista_codigo(db)}


def normalize_common_fields(resource: str, data: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(data)
    for key, value in list(normalized.items()):
        if key == "senha" or value is None:
            continue
        if isinstance(value, str):
            normalized[key] = upper_text(value)

    if resource in {"motoristas", "vendedores", "ajudantes"} and not normalized.get("status"):
        normalized["status"] = "ATIVO"
    if resource == "produtos":
        normalized.setdefault("categoria", "AVES")
        normalized.setdefault("unidade", "KG")
        normalized.setdefault("unidade_estoque", "KG")
        normalized.setdefault("controla_estoque_fisico", 1)
        normalized.setdefault("controla_estoque_fiscal", 1)
        normalized.setdefault("status", "ATIVO")
    if resource == "motoristas" and not normalized.get("perfil_app"):
        normalized["perfil_app"] = "MOTORISTA"
    if resource == "usuarios" and not normalized.get("permissoes"):
        normalized["permissoes"] = "OPERADOR"
    return normalized


async def validate_and_prepare(
    *,
    resource: str,
    config: CadastroResource,
    db: AsyncSession,
    data: dict[str, Any],
    partial: bool,
    item: Any | None = None,
) -> dict[str, Any]:
    item_id = int(getattr(item, "id", 0) or 0) if item else None
    data = normalize_common_fields(resource, data)

    if resource == "motoristas":
        if not partial:
            data["codigo"] = await next_motorista_codigo(db)
        elif "codigo" in data:
            data.pop("codigo", None)

        nome = upper_text(data.get("nome") if "nome" in data else getattr(item, "nome", ""))
        if ("nome" in data or not partial) and len(nome) < 3:
            raise HTTPException(status_code=422, detail="NOME deve ter pelo menos 3 caracteres.")
        if "nome" in data:
            data["nome"] = nome

        codigo = upper_text(data.get("codigo") if "codigo" in data else getattr(item, "codigo", ""))
        if ("codigo" in data or not partial) and not is_valid_motorista_codigo(codigo):
            raise HTTPException(
                status_code=422,
                detail="CODIGO invalido. Use letras/numeros/._- e 3 a 24 caracteres.",
            )
        if "codigo" in data:
            if await exists_ci(db, MotoristaDB, "codigo", codigo, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe motorista com este codigo: {codigo}")
            data["codigo"] = codigo

        if "cpf" in data:
            cpf = normalize_cpf(data.get("cpf"))
            if cpf and not is_valid_cpf(cpf):
                raise HTTPException(status_code=422, detail="CPF invalido.")
            if cpf and await exists_exact(db, MotoristaDB, "cpf", cpf, exclude_id=item_id):
                raise HTTPException(status_code=409, detail="Ja existe motorista com este CPF.")
            data["cpf"] = cpf or None

        telefone_source = data.get("telefone") if "telefone" in data else getattr(item, "telefone", "")
        telefone = normalize_phone(telefone_source)
        if ("telefone" in data or not partial) and not is_valid_phone(telefone):
            raise HTTPException(status_code=422, detail="TELEFONE invalido. Informe DDD+numero.")
        if "telefone" in data:
            data["telefone"] = telefone

        if "status" in data:
            status_value = upper_text(data.get("status") or "ATIVO")
            if status_value not in {"ATIVO", "INATIVO"}:
                raise HTTPException(status_code=422, detail="Status invalido. Use ATIVO ou INATIVO.")
            data["status"] = status_value

        if "perfil_app" in data:
            perfil = upper_text(data.get("perfil_app") or "MOTORISTA")
            if perfil not in {"MOTORISTA", "ADMIN"}:
                perfil = "MOTORISTA"
            data["perfil_app"] = perfil

        if "senha" in data:
            senha = str(data.get("senha") or "").strip()
            if senha == "******":
                data.pop("senha", None)
            elif senha:
                if not is_valid_motorista_senha(senha):
                    raise HTTPException(status_code=422, detail="SENHA invalida. Use 4 a 24 caracteres.")
                data["senha"] = hash_password_pbkdf2(senha)
            elif not partial:
                raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")
            else:
                data.pop("senha", None)
        elif not partial:
            raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")

    elif resource == "vendedores":
        codigo = upper_text(data.get("codigo") if "codigo" in data else getattr(item, "codigo", ""))
        nome = upper_text(data.get("nome") if "nome" in data else getattr(item, "nome", ""))
        if ("codigo" in data or not partial) and not is_valid_motorista_codigo(codigo):
            raise HTTPException(
                status_code=422,
                detail="CODIGO invalido. Use letras/numeros/._- e 3 a 24 caracteres.",
            )
        if ("nome" in data or not partial) and len(nome) < 3:
            raise HTTPException(status_code=422, detail="NOME deve ter pelo menos 3 caracteres.")
        if "codigo" in data:
            if await exists_ci(db, VendedorDB, "codigo", codigo, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe vendedor com este codigo: {codigo}")
            data["codigo"] = codigo
        if "nome" in data:
            data["nome"] = nome
        if "telefone" in data:
            telefone = normalize_phone(data.get("telefone"))
            if telefone and not is_valid_phone(telefone):
                raise HTTPException(status_code=422, detail="TELEFONE invalido. Informe DDD+numero ou deixe vazio.")
            data["telefone"] = telefone or None
        if "status" in data:
            status_value = upper_text(data.get("status") or "ATIVO")
            if status_value not in {"ATIVO", "DESATIVADO"}:
                raise HTTPException(status_code=422, detail="STATUS invalido para vendedor.")
            data["status"] = status_value
        if "senha" in data:
            senha = str(data.get("senha") or "").strip()
            if senha == "******":
                data.pop("senha", None)
            elif senha:
                if not is_valid_motorista_senha(senha):
                    raise HTTPException(status_code=422, detail="SENHA invalida. Use 4 a 24 caracteres.")
                data["senha"] = hash_password_pbkdf2(senha)
            elif not partial:
                raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")
            else:
                data.pop("senha", None)
        elif not partial:
            raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")

    elif resource == "usuarios":
        nome = upper_text(data.get("nome") if "nome" in data else getattr(item, "nome", ""))
        if ("nome" in data or not partial) and not nome:
            raise HTTPException(status_code=422, detail="Preencha o campo: NOME")
        if "nome" in data:
            if await exists_ci(db, UserDB, "nome", nome, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe usuario com este nome: {nome}")
            if await exists_ci(db, UserDB, "username", nome, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe usuario com este login: {nome}")
            data["nome"] = nome
            data["username"] = nome
        if "permissoes" in data:
            data["permissoes"] = upper_text(data.get("permissoes") or "OPERADOR")
        if "senha" in data:
            senha = str(data.get("senha") or "").strip()
            if senha == "******":
                data.pop("senha", None)
            elif senha:
                if len(senha) < 6:
                    raise HTTPException(status_code=422, detail="SENHA deve ter pelo menos 6 caracteres.")
                data["senha"] = get_password_hash(senha)
            elif not partial:
                raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")
            else:
                data.pop("senha", None)
        elif not partial:
            raise HTTPException(status_code=422, detail="Preencha o campo: SENHA")

    elif resource == "ajudantes":
        nome = upper_text(data.get("nome") if "nome" in data else getattr(item, "nome", ""))
        sobrenome = upper_text(data.get("sobrenome") if "sobrenome" in data else getattr(item, "sobrenome", ""))
        if ("nome" in data or not partial) and len(nome) < 2:
            raise HTTPException(status_code=422, detail="NOME deve ter pelo menos 2 caracteres.")
        if ("sobrenome" in data or not partial) and len(sobrenome) < 2:
            raise HTTPException(status_code=422, detail="SOBRENOME deve ter pelo menos 2 caracteres.")
        if ("nome" in data or "sobrenome" in data or not partial) and nome and sobrenome:
            stmt = select(AjudanteDB.id).where(
                func.upper(func.coalesce(AjudanteDB.nome, "")) == nome,
                func.upper(func.coalesce(AjudanteDB.sobrenome, "")) == sobrenome,
            )
            if item_id:
                stmt = stmt.where(AjudanteDB.id != item_id)
            result = await db.execute(stmt.limit(1))
            if result.scalar_one_or_none() is not None:
                raise HTTPException(status_code=409, detail="Ja existe ajudante com este nome/sobrenome.")
        if "nome" in data:
            data["nome"] = nome
        if "sobrenome" in data:
            data["sobrenome"] = sobrenome
        if "telefone" in data or not partial:
            telefone = normalize_phone(data.get("telefone") if "telefone" in data else getattr(item, "telefone", ""))
            if not is_valid_phone(telefone):
                raise HTTPException(status_code=422, detail="TELEFONE invalido. Informe DDD+numero.")
            if "telefone" in data:
                data["telefone"] = telefone
        if "status" in data:
            status_value = upper_text(data.get("status") or "ATIVO")
            if status_value not in {"ATIVO", "DESATIVADO"}:
                raise HTTPException(status_code=422, detail="STATUS invalido para ajudante.")
            data["status"] = status_value

    elif resource == "veiculos":
        placa = upper_text(data.get("placa") if "placa" in data else getattr(item, "placa", ""))
        if ("placa" in data or not partial) and not placa:
            raise HTTPException(status_code=422, detail="Preencha o campo: PLACA")
        if "placa" in data:
            if await exists_ci(db, VeiculoDB, "placa", placa, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe veiculo com esta placa: {placa}")
            data["placa"] = placa
        if "modelo" in data:
            data["modelo"] = upper_text(data.get("modelo"))
        if "capacidade_cx" in data or not partial:
            cap = safe_int(data.get("capacidade_cx") if "capacidade_cx" in data else getattr(item, "capacidade_cx", 0), -1)
            if cap < 0:
                raise HTTPException(status_code=422, detail="CAPACIDADE (CX) deve ser inteiro >= 0.")
            if "capacidade_cx" in data:
                data["capacidade_cx"] = cap

    elif resource == "caixas":
        codigo = upper_text(data.get("codigo") if "codigo" in data else getattr(item, "codigo", ""))
        if ("codigo" in data or not partial) and len(codigo) < 2:
            raise HTTPException(status_code=422, detail="CODIGO da caixa deve ter pelo menos 2 caracteres.")
        if "codigo" in data:
            codigo = "".join(ch if ch.isalnum() else "-" for ch in codigo).strip("-")
            if await exists_ci(db, CaixaDB, "codigo", codigo, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe caixa com este codigo: {codigo}")
            data["codigo"] = codigo
        if "lote" in data or not partial:
            lote = upper_text(data.get("lote") if "lote" in data else getattr(item, "lote", ""))
            if not lote:
                raise HTTPException(status_code=422, detail="Preencha o campo: LOTE")
            if "lote" in data:
                data["lote"] = lote
        if "cor" in data or not partial:
            cor = upper_text(data.get("cor") if "cor" in data else getattr(item, "cor", ""))
            if not cor:
                raise HTTPException(status_code=422, detail="Preencha o campo: COR")
            if "cor" in data:
                data["cor"] = cor
        if "veiculo_placa" in data:
            placa = upper_text(data.get("veiculo_placa"))
            if placa:
                result = await db.execute(
                    select(VeiculoDB.id).where(func.upper(func.coalesce(VeiculoDB.placa, "")) == placa).limit(1)
                )
                if result.scalar_one_or_none() is None:
                    raise HTTPException(status_code=422, detail=f"Veiculo nao encontrado para vinculo: {placa}")
            data["veiculo_placa"] = placa or None
        if "status" in data:
            status_value = upper_text(data.get("status") or "EM_ESTOQUE")
            if status_value not in CAIXA_STATUS:
                raise HTTPException(status_code=422, detail="STATUS invalido para caixa.")
            data["status"] = status_value
            if status_value in {"EM_ESTOQUE", "BAIXADA"} and not data.get("veiculo_placa"):
                data["veiculo_placa"] = None
        if "data_compra" in data:
            data["data_compra"] = str(data.get("data_compra") or "").strip()[:30] or None
        if "observacao" in data:
            data["observacao"] = upper_text(data.get("observacao"))[:300] or None

    elif resource == "clientes":
        if "cod_cliente" in data:
            cod = upper_text(data.get("cod_cliente"))
            if await exists_ci(db, ClienteDB, "cod_cliente", cod, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe cliente com este codigo: {cod}")
            data["cod_cliente"] = cod
        if "nome_cliente" in data:
            nome = upper_text(data.get("nome_cliente"))
            if not nome:
                raise HTTPException(status_code=422, detail="Preencha o campo: NOME_CLIENTE")
            data["nome_cliente"] = nome

    elif resource == "fornecedores":
        if "razao_social" in data:
            razao = upper_text(data.get("razao_social"))
            if not razao:
                raise HTTPException(status_code=422, detail="Preencha o campo: RAZAO SOCIAL")
            data["razao_social"] = razao
        if "nome_fantasia" in data:
            data["nome_fantasia"] = upper_text(data.get("nome_fantasia"))
        if "documento" in data or not partial:
            documento = "".join(ch for ch in str(data.get("documento") if "documento" in data else getattr(item, "documento", "")) if ch.isdigit())
            if len(documento) not in {11, 14}:
                raise HTTPException(status_code=422, detail="CPF/CNPJ do fornecedor invalido.")
            if await exists_exact(db, FornecedorDB, "documento", documento, exclude_id=item_id):
                raise HTTPException(status_code=409, detail="Ja existe fornecedor com este CPF/CNPJ.")
            if "documento" in data:
                data["documento"] = documento
                data["tipo_pessoa"] = "CPF" if len(documento) == 11 else "CNPJ"
        if "perfil_fornecedor" in data or not partial:
            perfil = upper_text(data.get("perfil_fornecedor") if "perfil_fornecedor" in data else getattr(item, "perfil_fornecedor", "OUTROS"))
            perfis_validos = await fornecedor_perfis_ativos(db)
            if perfil not in perfis_validos:
                raise HTTPException(status_code=422, detail="Perfil de fornecedor invalido.")
            if "perfil_fornecedor" in data:
                data["perfil_fornecedor"] = perfil
        if "status" in data:
            status_value = upper_text(data.get("status") or "ATIVO")
            if status_value not in {"ATIVO", "INATIVO"}:
                raise HTTPException(status_code=422, detail="STATUS invalido para fornecedor.")
            data["status"] = status_value
        if "uf" in data:
            data["uf"] = upper_text(data.get("uf"))[:2] or None
        if "cidade" in data:
            data["cidade"] = upper_text(data.get("cidade"))
        if "telefone" in data:
            data["telefone"] = normalize_phone(data.get("telefone")) or None
        if "email" in data:
            data["email"] = str(data.get("email") or "").strip().lower() or None
        for readonly in ("certificado_status", "certificado_nome", "certificado_instalado_em"):
            data.pop(readonly, None)

    elif resource == "produtos":
        if "codigo" in data or not partial:
            codigo = upper_text(data.get("codigo") if "codigo" in data else getattr(item, "codigo", ""))
            if len(codigo) < 2:
                raise HTTPException(status_code=422, detail="CODIGO do produto deve ter pelo menos 2 caracteres.")
            if await exists_ci(db, ProdutoDB, "codigo", codigo, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe produto com este codigo: {codigo}")
            if "codigo" in data:
                data["codigo"] = codigo
        if "nome" in data or not partial:
            nome = upper_text(data.get("nome") if "nome" in data else getattr(item, "nome", ""))
            if len(nome) < 2:
                raise HTTPException(status_code=422, detail="NOME do produto deve ter pelo menos 2 caracteres.")
            if await exists_ci(db, ProdutoDB, "nome", nome, exclude_id=item_id):
                raise HTTPException(status_code=409, detail=f"Ja existe produto com este nome: {nome}")
            if "nome" in data:
                data["nome"] = nome
        if "categoria" in data or not partial:
            categoria = upper_text(data.get("categoria") if "categoria" in data else getattr(item, "categoria", "AVES"))
            if categoria not in PRODUTO_CATEGORIAS:
                raise HTTPException(status_code=422, detail="Categoria de produto invalida.")
            if "categoria" in data:
                data["categoria"] = categoria
        for unidade_field in ("unidade", "unidade_estoque"):
            if unidade_field in data or not partial:
                unidade = upper_text(data.get(unidade_field) if unidade_field in data else getattr(item, unidade_field, "KG"))
                if unidade not in PRODUTO_UNIDADES:
                    raise HTTPException(status_code=422, detail=f"{unidade_field.upper()} invalida.")
                if unidade_field in data:
                    data[unidade_field] = unidade
        if "status" in data:
            status_value = upper_text(data.get("status") or "ATIVO")
            if status_value not in {"ATIVO", "INATIVO"}:
                raise HTTPException(status_code=422, detail="STATUS invalido para produto.")
            data["status"] = status_value
        for flag in ("controla_estoque_fisico", "controla_estoque_fiscal"):
            if flag in data:
                data[flag] = 1 if int(data.get(flag) or 0) else 0
        for numeric in ("estoque_min_kg", "custo_padrao", "preco_padrao"):
            if numeric in data:
                data[numeric] = max(float(data.get(numeric) or 0), 0)
        if "estoque_min_caixas" in data:
            data["estoque_min_caixas"] = max(safe_int(data.get("estoque_min_caixas"), 0), 0)
        for code_field in ("ncm", "cest", "cfop_entrada", "cfop_saida", "ean"):
            if code_field in data:
                data[code_field] = "".join(ch for ch in str(data.get(code_field) or "") if ch.isdigit()) or None

    return data


async def delete_block_reason(db: AsyncSession, resource: str, item: Any) -> str:
    if resource == "motoristas":
        codigo = upper_text(getattr(item, "codigo", ""))
        nome = upper_text(getattr(item, "nome", ""))
        cols = await table_columns(db, "programacoes")
        conds: list[str] = []
        params: dict[str, Any] = {}
        if codigo:
            params["codigo"] = codigo
            if "motorista_codigo" in cols:
                conds.append("UPPER(COALESCE(motorista_codigo,''))=:codigo")
            if "codigo_motorista" in cols:
                conds.append("UPPER(COALESCE(codigo_motorista,''))=:codigo")
            if "motorista" in cols:
                conds.append("UPPER(COALESCE(motorista,''))=:codigo")
        if nome and "motorista" in cols:
            params["nome"] = nome
            conds.append("UPPER(COALESCE(motorista,''))=:nome")
        if conds and await count_sql(db, f"SELECT COUNT(*) FROM programacoes WHERE {' OR '.join(conds)}", params):
            return "Motorista vinculado a programacao/rota. Use status INATIVO."

    if resource == "vendedores":
        codigo = upper_text(getattr(item, "codigo", ""))
        nome = upper_text(getattr(item, "nome", ""))
        if codigo and "vendedor" in await table_columns(db, "clientes"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM clientes WHERE UPPER(COALESCE(vendedor,''))=:codigo",
                {"codigo": codigo},
            ):
                return "Vendedor vinculado ao cadastro de clientes. Use status DESATIVADO."
        if nome and "vendedor" in await table_columns(db, "clientes"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM clientes WHERE UPPER(COALESCE(vendedor,''))=:nome",
                {"nome": nome},
            ):
                return "Vendedor vinculado ao cadastro de clientes. Use status DESATIVADO."
        if codigo and "usuario_criacao" in await table_columns(db, "programacoes"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM programacoes WHERE UPPER(COALESCE(usuario_criacao,''))=:codigo",
                {"codigo": codigo},
            ):
                return "Vendedor vinculado a programacao/rota. Use status DESATIVADO."

    if resource == "veiculos":
        placa = upper_text(getattr(item, "placa", ""))
        if placa and "veiculo" in await table_columns(db, "programacoes"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM programacoes WHERE UPPER(COALESCE(veiculo,''))=:placa",
                {"placa": placa},
            ):
                return "Veiculo vinculado a programacao/rota."

    if resource == "clientes":
        cod = upper_text(getattr(item, "cod_cliente", ""))
        if cod and "cod_cliente" in await table_columns(db, "programacao_itens"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM programacao_itens WHERE UPPER(COALESCE(cod_cliente,''))=:cod",
                {"cod": cod},
            ):
                return "Cliente vinculado a programacao."
        if cod and "cod_cliente" in await table_columns(db, "recebimentos"):
            if await count_sql(
                db,
                "SELECT COUNT(*) FROM recebimentos WHERE UPPER(COALESCE(cod_cliente,''))=:cod",
                {"cod": cod},
            ):
                return "Cliente vinculado a recebimentos."

    if resource == "ajudantes":
        cols = await table_columns(db, "equipes")
        if not cols:
            return ""
        nome = upper_text(getattr(item, "nome", ""))
        sobrenome = upper_text(getattr(item, "sobrenome", ""))
        alvo = f"{nome} {sobrenome}".strip()
        params = {"id": str(item.id), "nome": nome, "alvo": alvo}
        conds = []
        for column in ("ajudante1", "ajudante2", "ajudante_1", "ajudante_2"):
            if column in cols:
                conds.append(f"UPPER(COALESCE({column},''))=UPPER(:id)")
                if nome:
                    conds.append(f"UPPER(COALESCE({column},''))=:nome")
                if alvo:
                    conds.append(f"UPPER(COALESCE({column},''))=:alvo")
        if conds and await count_sql(db, f"SELECT COUNT(*) FROM equipes WHERE {' OR '.join(conds)}", params):
            return "Ajudante vinculado a equipe."

    if resource == "produtos":
        produto_id = int(getattr(item, "id", 0) or 0)
        nome = upper_text(getattr(item, "nome", ""))
        linked_checks = [
            ("programacao_itens", "produto_id", "Produto vinculado a itens de programacao."),
            ("vendas_importadas", "produto_id", "Produto vinculado a vendas importadas."),
            ("compras_nfe", "produto_id", "Produto vinculado a compras/NF-e."),
            ("compras_nfe_itens", "produto_id", "Produto vinculado a itens de compras/NF-e."),
            ("estoque_movimentos", "produto_id", "Produto vinculado ao estoque."),
        ]
        for table_name, column, message in linked_checks:
            if column in await table_columns(db, table_name):
                if await count_sql(db, f"SELECT COUNT(*) FROM {table_name} WHERE COALESCE({column}, 0)=:produto_id", {"produto_id": produto_id}):
                    return message
        for table_name, message in [
            ("programacao_itens", "Produto vinculado a itens de programacao."),
            ("vendas_importadas", "Produto vinculado a vendas importadas."),
            ("compras_nfe", "Produto vinculado a compras/NF-e."),
            ("compras_nfe_itens", "Produto vinculado a itens de compras/NF-e."),
            ("estoque_movimentos", "Produto vinculado ao estoque."),
        ]:
            if nome and "produto" in await table_columns(db, table_name):
                if await count_sql(db, f"SELECT COUNT(*) FROM {table_name} WHERE UPPER(COALESCE(produto,''))=:nome", {"nome": nome}):
                    return message

    return ""


def safe_filename_part(value: Any, fallback: str = "arquivo") -> str:
    text_value = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in str(value or "").strip())
    return text_value.strip("_")[:80] or fallback


@router.post("/fornecedores/{item_id}/certificado", response_model=CadastroItemResponse)
async def instalar_certificado_fornecedor(
    item_id: int,
    request: Request,
    certificado: UploadFile = File(...),
    apelido: str = Form(default=""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    fornecedor = await get_item_or_404(db, RESOURCES["fornecedores"], item_id)
    filename = str(certificado.filename or "").strip()
    if not filename.lower().endswith((".pfx", ".p12")):
        raise HTTPException(status_code=422, detail="Envie um certificado A1 no formato .pfx ou .p12.")
    contents = await certificado.read()
    if not contents:
        raise HTTPException(status_code=422, detail="Arquivo de certificado vazio.")
    if len(contents) > 8 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="Certificado maior que 8 MB.")

    root = Path(".rotahub_runtime") / "certificados" / "fornecedores"
    root.mkdir(parents=True, exist_ok=True)
    ext = ".p12" if filename.lower().endswith(".p12") else ".pfx"
    stored_name = f"{safe_filename_part(fornecedor.documento, 'fornecedor')}_{uuid4().hex[:12]}{ext}"
    target = root / stored_name
    target.write_bytes(contents)

    fornecedor.certificado_nome = upper_text(apelido) or filename
    fornecedor.certificado_path = str(target)
    fornecedor.certificado_status = "INSTALADO"
    fornecedor.certificado_instalado_em = datetime.now().isoformat(timespec="seconds")
    record_audit_log(
        db,
        action="fornecedor_certificado_instalado",
        actor_user=current_user,
        entity_type="fornecedores",
        entity_id=fornecedor.id,
        severity="warning",
        ip_address=client_ip_from_request(request),
        metadata={
            "fornecedor_id": fornecedor.id,
            "documento": fornecedor.documento,
            "arquivo_original": filename,
            "senha_armazenada": False,
        },
    )
    await db.commit()
    await db.refresh(fornecedor)
    return serialize_item(fornecedor, RESOURCES["fornecedores"])


@router.get("/{resource}", response_model=List[CadastroItemResponse])
async def list_cadastro_items(
    resource: str,
    skip: int = 0,
    limit: int = 200,
    status_filter: str = "TODOS",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    config = get_resource(resource)
    order_column = getattr(config.model, config.order_by)
    stmt = select(config.model)
    if resource in {"ajudantes", "fornecedores", "produtos", "caixas"} and upper_text(status_filter) in {"ATIVO", "DESATIVADO", "INATIVO", *CAIXA_STATUS}:
        status_column = getattr(config.model, "status")
        stmt = stmt.where(func.upper(func.coalesce(status_column, "ATIVO")) == upper_text(status_filter))
    stmt = stmt.order_by(order_column.desc()).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return [serialize_item(item, config) for item in result.scalars().all()]


@router.get("/meta/fornecedor-perfis")
async def list_fornecedor_perfis(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    await ensure_fornecedor_perfis(db)
    result = await db.execute(
        text(
            """
            SELECT id, codigo, nome, categoria, status
              FROM fornecedor_perfis
             ORDER BY CASE WHEN codigo='FRANGO_VIVO' THEN 0 WHEN codigo='OUTROS' THEN 2 ELSE 1 END, nome
            """
        )
    )
    return [
        {
            "id": int(row[0] or 0),
            "codigo": upper_text(row[1]),
            "nome": str(row[2] or "").strip(),
            "categoria": upper_text(row[3] or "OUTROS"),
            "status": upper_text(row[4] or "ATIVO"),
        }
        for row in result.all()
    ]


@router.post("/meta/fornecedor-perfis", status_code=status.HTTP_201_CREATED)
async def create_fornecedor_perfil(
    payload: FornecedorPerfilPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    await ensure_fornecedor_perfis(db)
    codigo = upper_text(payload.codigo)
    codigo = "".join(ch if ch.isalnum() else "_" for ch in codigo).strip("_")
    nome = str(payload.nome or "").strip()
    categoria = upper_text(payload.categoria or "OUTROS")
    status_value = upper_text(payload.status or "ATIVO")
    if len(codigo) < 2:
        raise HTTPException(status_code=422, detail="Codigo do perfil invalido.")
    if status_value not in {"ATIVO", "INATIVO"}:
        raise HTTPException(status_code=422, detail="STATUS invalido para perfil.")
    try:
        await db.execute(
            text(
                """
                INSERT INTO fornecedor_perfis (codigo, nome, categoria, status)
                VALUES (:codigo, :nome, :categoria, :status)
                """
            ),
            {"codigo": codigo, "nome": nome, "categoria": categoria, "status": status_value},
        )
        record_audit_log(
            db,
            action="fornecedor_perfil_criado",
            actor_user=current_user,
            entity_type="fornecedor_perfis",
            ip_address=client_ip_from_request(request),
            metadata={"codigo": codigo, "status": status_value},
        )
        await db.commit()
    except SQLAlchemyError as exc:
        await db.rollback()
        raise HTTPException(status_code=409, detail=f"Nao foi possivel criar o perfil: {exc}") from exc
    return {"codigo": codigo, "nome": nome, "categoria": categoria, "status": status_value}


@router.get("/clientes/importacao/rows")
async def list_clientes_importacao_rows(
    limit: int = 5000,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    result = await db.execute(
        select(ClienteDB)
        .order_by(func.upper(func.coalesce(ClienteDB.nome_cliente, "")), func.upper(func.coalesce(ClienteDB.cod_cliente, "")))
        .limit(max(1, min(limit, 10000)))
    )
    return [cliente_import_response(item) for item in result.scalars().all()]


@router.get("/clientes/dashboard", response_model=ClienteDashboardResponse)
async def clientes_dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    total_clientes = safe_int((await db.execute(text("SELECT COUNT(*) FROM clientes"))).scalar(), 0)
    clientes_com_historico = safe_int(
        (
            await db.execute(
                text(
                    """
                    SELECT COUNT(DISTINCT NULLIF(TRIM(COALESCE(cod_cliente,'')), ''))
                    FROM programacao_itens
                    WHERE NULLIF(TRIM(COALESCE(cod_cliente,'')), '') IS NOT NULL
                    """
                )
            )
        ).scalar(),
        0,
    )
    amostras_localizacao = safe_int((await db.execute(text("SELECT COUNT(*) FROM cliente_localizacao_amostras"))).scalar(), 0)
    clientes_com_localizacao = safe_int(
        (
            await db.execute(
                text(
                    """
                    SELECT COUNT(DISTINCT NULLIF(TRIM(COALESCE(cod_cliente,'')), ''))
                    FROM cliente_localizacao_amostras
                    WHERE (latitude IS NOT NULL OR longitude IS NOT NULL)
                      AND NULLIF(TRIM(COALESCE(cod_cliente,'')), '') IS NOT NULL
                    """
                )
            )
        ).scalar(),
        0,
    )
    return ClienteDashboardResponse(
        total_clientes=total_clientes,
        clientes_com_historico=clientes_com_historico,
        amostras_localizacao=amostras_localizacao,
        clientes_com_localizacao=clientes_com_localizacao,
    )


@router.get("/clientes/lookup", response_model=list[ClienteLookupResponse])
async def clientes_lookup(
    limit: int = 5000,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    result = await db.execute(
        text(
            """
            SELECT cod_cliente, COALESCE(NULLIF(nome_cliente,''), nome, '') AS nome_cliente
            FROM clientes
            WHERE NULLIF(TRIM(COALESCE(cod_cliente,'')), '') IS NOT NULL
            ORDER BY nome_cliente ASC
            LIMIT :limit
            """
        ),
        {"limit": max(1, min(int(limit or 5000), 10000))},
    )
    return [
        ClienteLookupResponse(cod_cliente=str(row["cod_cliente"] or ""), nome_cliente=str(row["nome_cliente"] or ""))
        for row in result.mappings().all()
    ]


@router.get("/clientes/{cod_cliente}/historico", response_model=ClienteHistoricoResponse)
async def cliente_historico(
    cod_cliente: str,
    limit: int = 300,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    cod = upper_text(cod_cliente)
    result = await db.execute(
        text(
            """
            SELECT
                COALESCE(NULLIF(p.data_criacao,''), NULLIF(p.data,''), '') AS data_ref,
                pi.codigo_programacao,
                pi.pedido,
                pi.nome_cliente,
                pi.vendedor,
                COALESCE(NULLIF(TRIM(pc.status_pedido), ''), NULLIF(TRIM(pi.status_pedido), ''), 'PENDENTE') AS status_pedido,
                COALESCE(pi.qnt_caixas, 0) AS caixas_programadas,
                COALESCE(pc.caixas_atual, pi.caixas_atual, pi.qnt_caixas, 0) AS caixas_atuais,
                COALESCE(pi.kg, 0) AS kg_programado,
                COALESCE(pc.peso_previsto, pi.kg, 0) AS kg_atual,
                COALESCE(pc.mortalidade_aves, 0) AS mortalidade_aves,
                COALESCE(pc.valor_recebido, 0) AS valor_recebido,
                COALESCE(pc.lat_entrega, pc.lat_evento) AS latitude,
                COALESCE(pc.lon_entrega, pc.lon_evento) AS longitude,
                COALESCE(pc.alteracao_tipo, pc.alteracao_detalhe, pi.alteracao_tipo, pi.alteracao_detalhe, '') AS alteracao,
                COALESCE(pc.alterado_em, pc.updated_at, '') AS alterado_em,
                COALESCE(p.motorista, p.motorista_codigo, p.codigo_motorista, '') AS motorista
            FROM programacao_itens pi
            LEFT JOIN programacao_itens_controle pc
              ON UPPER(TRIM(COALESCE(pc.codigo_programacao,''))) = UPPER(TRIM(COALESCE(pi.codigo_programacao,'')))
             AND UPPER(TRIM(COALESCE(pc.cod_cliente,''))) = UPPER(TRIM(COALESCE(pi.cod_cliente,'')))
             AND UPPER(TRIM(COALESCE(pc.pedido,''))) = UPPER(TRIM(COALESCE(pi.pedido,'')))
            LEFT JOIN programacoes p
              ON UPPER(TRIM(COALESCE(p.codigo_programacao, p.codigo, ''))) = UPPER(TRIM(COALESCE(pi.codigo_programacao,'')))
            WHERE UPPER(TRIM(COALESCE(pi.cod_cliente,''))) = :cod
            ORDER BY COALESCE(NULLIF(data_ref,''), pi.codigo_programacao) DESC, pi.id DESC
            LIMIT :limit
            """
        ),
        {"cod": cod, "limit": max(1, min(int(limit or 300), 1000))},
    )
    rows = []
    for row in result.mappings().all():
        status_value = upper_text(row["status_pedido"] or "PENDENTE")
        kg_prog = safe_float(row["kg_programado"], 0.0)
        kg_atual = safe_float(row["kg_atual"], kg_prog)
        entregue = status_value in {"ENTREGUE", "FINALIZADO", "FINALIZADA", "CONCLUIDO", "CONCLUÍDO"}
        cancelado = status_value in {"CANCELADO", "CANCELADA"}
        alterado = bool(str(row["alteracao"] or "").strip()) or safe_int(row["caixas_atuais"], 0) != safe_int(row["caixas_programadas"], 0)
        rows.append(
            {
                "data_ref": str(row["data_ref"] or ""),
                "codigo_programacao": str(row["codigo_programacao"] or ""),
                "pedido": str(row["pedido"] or ""),
                "nome_cliente": str(row["nome_cliente"] or ""),
                "vendedor": str(row["vendedor"] or ""),
                "status_pedido": status_value,
                "caixas_programadas": safe_int(row["caixas_programadas"], 0),
                "caixas_atuais": safe_int(row["caixas_atuais"], 0),
                "kg_recebido": kg_atual if entregue and not cancelado else 0.0,
                "kg_descontado": kg_prog if cancelado else max(kg_prog - kg_atual, 0.0),
                "mortalidade_aves": safe_int(row["mortalidade_aves"], 0),
                "valor_recebido": safe_float(row["valor_recebido"], 0.0),
                "latitude": row["latitude"],
                "longitude": row["longitude"],
                "alterado": alterado,
                "alterado_em": str(row["alterado_em"] or ""),
                "motorista": str(row["motorista"] or ""),
            }
        )
    resumo = {
        "total_programacoes": len(rows),
        "entregues": sum(1 for row in rows if row["status_pedido"] in {"ENTREGUE", "FINALIZADO", "FINALIZADA", "CONCLUIDO", "CONCLUÍDO"}),
        "canceladas": sum(1 for row in rows if row["status_pedido"] in {"CANCELADO", "CANCELADA"}),
        "alteradas": sum(1 for row in rows if row["alterado"]),
        "mortalidade_aves": sum(safe_int(row["mortalidade_aves"], 0) for row in rows),
        "kg_recebidos": sum(safe_float(row["kg_recebido"], 0.0) for row in rows),
        "kg_descontados": sum(safe_float(row["kg_descontado"], 0.0) for row in rows),
    }
    return ClienteHistoricoResponse(resumo=resumo, rows=rows)


@router.get("/clientes/{cod_cliente}/localizacoes", response_model=ClienteLocalizacoesResponse)
async def cliente_localizacoes(
    cod_cliente: str,
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    result = await db.execute(
        text(
            """
            SELECT codigo_programacao, pedido, latitude, longitude, endereco, cidade, bairro,
                   status_pedido, motorista_codigo, motorista_nome, origem, registrado_em
            FROM cliente_localizacao_amostras
            WHERE UPPER(TRIM(COALESCE(cod_cliente,''))) = :cod
            ORDER BY registrado_em DESC, id DESC
            LIMIT :limit
            """
        ),
        {"cod": upper_text(cod_cliente), "limit": max(1, min(int(limit or 200), 1000))},
    )
    rows = [dict(row) for row in result.mappings().all()]
    resumo = {
        "amostras": len(rows),
        "ultima": rows[0] if rows else {},
        "com_coordenada": sum(1 for row in rows if row.get("latitude") is not None or row.get("longitude") is not None),
    }
    return ClienteLocalizacoesResponse(resumo=resumo, rows=rows)


@router.get("/clientes/importacao/modelo")
async def download_clientes_importacao_modelo(
    current_user: User = Depends(require_admin_user),
):
    del current_user
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Biblioteca openpyxl indisponivel.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CLIENTES"
    headers = ["COD CLIENTE", "NOME CLIENTE", "ENDERECO", "TELEFONE", "VENDEDOR"]
    sheet.append(headers)
    sheet.append(["123", "CLIENTE MODELO", "RUA EXEMPLO, 100", "(88) 99999-9999", "VENDEDOR MODELO"])
    sheet.append(["124", "CLIENTE COM MAIS DE UM TELEFONE", "AV EXEMPLO, 200", "(88) 98888-8888 / (88) 97777-7777", ""])
    sheet.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    widths = {
        "A": 18,
        "B": 34,
        "C": 40,
        "D": 34,
        "E": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="MODELO_IMPORTACAO_CLIENTES.xlsx"'},
    )


@router.post("/clientes/importacao/bulk-upsert", response_model=ClientesImportResult)
async def bulk_upsert_clientes_importacao(
    payload: ClientesImportPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    try:
        result = await bulk_upsert_clientes(db, payload.rows)
        record_audit_log(
            db,
            action="clientes_importacao_salva",
            actor_user=current_user,
            entity_type="clientes",
            ip_address=client_ip_from_request(request),
            metadata=result.model_dump(),
        )
        await db.commit()
        return result
    except SQLAlchemyError as exc:
        await db.rollback()
        raise HTTPException(status_code=422, detail=f"Nao foi possivel salvar os clientes no banco: {exc}") from exc


@router.post("/clientes/importacao/upload", response_model=ClientesImportResult, status_code=status.HTTP_201_CREATED)
async def upload_clientes_importacao(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    filename = str(file.filename or "").lower()
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=422, detail="Arquivo vazio.")
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(contents))
        else:
            engine = "xlrd" if filename.endswith(".xls") and not filename.endswith(".xlsx") else "openpyxl"
            df = pd.read_excel(BytesIO(contents), engine=engine)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Nao foi possivel ler o arquivo: {exc}") from exc

    try:
        rows = clientes_rows_from_dataframe(df)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Nao foi possivel interpretar as colunas do arquivo: {exc}") from exc

    try:
        result = await bulk_upsert_clientes(db, rows, merge_duplicate_rows=True)
        record_audit_log(
            db,
            action="clientes_importacao_excel",
            actor_user=current_user,
            entity_type="clientes",
            ip_address=client_ip_from_request(request),
            metadata={**result.model_dump(), "arquivo": file.filename},
        )
        await db.commit()
        return result
    except SQLAlchemyError as exc:
        await db.rollback()
        raise HTTPException(status_code=422, detail=f"Nao foi possivel salvar os clientes no banco: {exc}") from exc


@router.post("/caixas/bulk", response_model=CaixasBulkResult, status_code=status.HTTP_201_CREATED)
async def create_caixas_bulk(
    payload: CaixasBulkPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    prefixo = upper_text(payload.prefixo)
    prefixo = "".join(ch if ch.isalnum() else "-" for ch in prefixo).strip("-") or "CX"
    lote = upper_text(payload.lote)
    cor = upper_text(payload.cor)
    placa = upper_text(payload.veiculo_placa)
    status_value = upper_text(payload.status or "EM_ESTOQUE")
    if status_value not in CAIXA_STATUS:
        raise HTTPException(status_code=422, detail="STATUS invalido para caixa.")
    if placa:
        result = await db.execute(select(VeiculoDB.id).where(func.upper(func.coalesce(VeiculoDB.placa, "")) == placa).limit(1))
        if result.scalar_one_or_none() is None:
            raise HTTPException(status_code=422, detail=f"Veiculo nao encontrado para vinculo: {placa}")

    codigos = [
        f"{prefixo}-{str(payload.numero_inicial + index).zfill(payload.digitos)}"
        for index in range(payload.quantidade)
    ]
    result = await db.execute(select(CaixaDB.codigo).where(func.upper(func.coalesce(CaixaDB.codigo, "")).in_(codigos)))
    existentes = [upper_text(row[0]) for row in result.all()]
    if existentes:
        raise HTTPException(status_code=409, detail=f"Ja existem caixas nesta numeracao: {', '.join(existentes[:5])}")

    criadas: list[CaixaDB] = []
    for codigo in codigos:
        caixa = CaixaDB(
            codigo=codigo,
            lote=lote,
            cor=cor,
            veiculo_placa=placa or None,
            status=status_value,
            data_compra=str(payload.data_compra or "").strip()[:30] or None,
            observacao=upper_text(payload.observacao)[:300] or None,
        )
        db.add(caixa)
        criadas.append(caixa)
    await db.flush()
    for caixa in criadas:
        await record_caixa_movimento(db, caixa=caixa, movimento="CADASTRO_LOTE", observacao=payload.observacao)
    record_audit_log(
        db,
        action="caixas_lote_criado",
        actor_user=current_user,
        entity_type="caixas",
        ip_address=client_ip_from_request(request),
        metadata={"quantidade": len(criadas), "lote": lote, "prefixo": prefixo, "primeiro": codigos[0], "ultimo": codigos[-1]},
    )
    await db.commit()
    return CaixasBulkResult(criadas=len(criadas), primeiro_codigo=codigos[0], ultimo_codigo=codigos[-1])


@router.post("/caixas/movimentar", response_model=CaixasMovimentarResult)
async def movimentar_caixas(
    payload: CaixasMovimentarPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    lote = upper_text(payload.lote)
    cor = upper_text(payload.cor)
    veiculo_origem = upper_text(payload.veiculo_origem)
    veiculo_destino = upper_text(payload.veiculo_destino)
    status_origem = upper_text(payload.status_origem or "TODOS")
    status_destino = upper_text(payload.status_destino or "VINCULADA")
    observacao = upper_text(payload.observacao)[:300] or None

    if status_origem != "TODOS" and status_origem not in CAIXA_STATUS:
        raise HTTPException(status_code=422, detail="STATUS origem invalido para caixa.")
    if status_destino not in CAIXA_STATUS:
        raise HTTPException(status_code=422, detail="STATUS destino invalido para caixa.")
    if status_destino in {"VINCULADA", "EM_USO"} and not veiculo_destino:
        raise HTTPException(status_code=422, detail="Informe o veiculo destino para vincular/em uso.")
    if veiculo_destino:
        result = await db.execute(
            select(VeiculoDB.id).where(func.upper(func.coalesce(VeiculoDB.placa, "")) == veiculo_destino).limit(1)
        )
        if result.scalar_one_or_none() is None:
            raise HTTPException(status_code=422, detail=f"Veiculo destino nao encontrado: {veiculo_destino}")
    if veiculo_origem:
        result = await db.execute(
            select(VeiculoDB.id).where(func.upper(func.coalesce(VeiculoDB.placa, "")) == veiculo_origem).limit(1)
        )
        if result.scalar_one_or_none() is None:
            raise HTTPException(status_code=422, detail=f"Veiculo origem nao encontrado: {veiculo_origem}")

    stmt = select(CaixaDB).order_by(CaixaDB.id.asc()).limit(payload.quantidade)
    if lote:
        stmt = stmt.where(func.upper(func.coalesce(CaixaDB.lote, "")) == lote)
    if cor:
        stmt = stmt.where(func.upper(func.coalesce(CaixaDB.cor, "")) == cor)
    if veiculo_origem:
        stmt = stmt.where(func.upper(func.coalesce(CaixaDB.veiculo_placa, "")) == veiculo_origem)
    if status_origem != "TODOS":
        stmt = stmt.where(func.upper(func.coalesce(CaixaDB.status, "EM_ESTOQUE")) == status_origem)

    result = await db.execute(stmt)
    caixas = list(result.scalars().all())
    if not caixas:
        raise HTTPException(status_code=404, detail="Nenhuma caixa encontrada para os filtros informados.")
    if len(caixas) < payload.quantidade:
        raise HTTPException(status_code=409, detail=f"Filtros encontraram apenas {len(caixas)} caixa(s).")

    codigos: list[str] = []
    for caixa in caixas:
        origem_placa = caixa.veiculo_placa
        origem_status = caixa.status
        if status_destino in {"EM_ESTOQUE", "BAIXADA"} and not veiculo_destino:
            caixa.veiculo_placa = None
        elif veiculo_destino:
            caixa.veiculo_placa = veiculo_destino
        caixa.status = status_destino
        if observacao:
            caixa.observacao = observacao
        await record_caixa_movimento(
            db,
            caixa=caixa,
            movimento="MOVIMENTACAO_LOTE",
            veiculo_origem=origem_placa,
            status_origem=origem_status,
            observacao=observacao,
        )
        codigos.append(upper_text(caixa.codigo))

    record_audit_log(
        db,
        action="caixas_movimentacao_lote",
        actor_user=current_user,
        entity_type="caixas",
        ip_address=client_ip_from_request(request),
        metadata={
            "quantidade": len(codigos),
            "lote": lote,
            "cor": cor,
            "veiculo_origem": veiculo_origem,
            "veiculo_destino": veiculo_destino,
            "status_origem": status_origem,
            "status_destino": status_destino,
        },
    )
    await db.commit()
    return CaixasMovimentarResult(movimentadas=len(codigos), codigos=codigos)


@router.get("/caixas/{item_id}/historico")
async def caixa_historico(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    del current_user
    await ensure_caixas_movimentos(db)
    caixa = await db.get(CaixaDB, item_id)
    if caixa is None:
        raise HTTPException(status_code=404, detail="Caixa nao encontrada")
    result = await db.execute(
        text(
            """
            SELECT id, movimento, veiculo_origem, veiculo_destino, status_origem, status_destino, observacao, criado_em
              FROM caixas_movimentos
             WHERE caixa_id=:caixa_id
             ORDER BY COALESCE(criado_em, '') DESC, id DESC
            """
        ),
        {"caixa_id": item_id},
    )
    return {
        "caixa": serialize_item(caixa, RESOURCES["caixas"]).model_dump(),
        "rows": [dict(row) for row in result.mappings().all()],
    }


@router.post("/{resource}", response_model=CadastroItemResponse, status_code=status.HTTP_201_CREATED)
async def create_cadastro_item(
    resource: str,
    payload: CadastroPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    config = get_resource(resource)
    await enforce_create_limit(resource, db, current_user)
    data = clean_base_payload(payload, config, partial=False)
    data = await validate_and_prepare(resource=resource, config=config, db=db, data=data, partial=False)
    item = config.model(**data)
    assign_company_id(item, current_user)
    db.add(item)
    try:
        await db.flush()
        if resource == "caixas":
            await record_caixa_movimento(db, caixa=item, movimento="CADASTRO", observacao=data.get("observacao"))
        record_audit_log(
            db,
            action=f"{resource}_criado",
            actor_user=current_user,
            entity_type=resource,
            entity_id=item.id,
            ip_address=client_ip_from_request(request),
            metadata={"fields": sorted(field for field in data.keys() if field != "senha")},
        )
        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(status_code=409, detail=cadastro_integrity_message(resource, exc)) from exc
    except SQLAlchemyError as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=cadastro_integrity_message(resource, exc)) from exc
    await db.refresh(item)
    return serialize_item(item, config)


@router.patch("/{resource}/{item_id}", response_model=CadastroItemResponse)
async def update_cadastro_item(
    resource: str,
    item_id: int,
    payload: CadastroPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    config = get_resource(resource)
    item = await get_item_or_404(db, config, item_id)
    caixa_veiculo_origem = getattr(item, "veiculo_placa", None) if resource == "caixas" else None
    caixa_status_origem = getattr(item, "status", None) if resource == "caixas" else None
    data = clean_base_payload(payload, config, partial=True)
    data = await validate_and_prepare(resource=resource, config=config, db=db, data=data, partial=True, item=item)
    changed_fields: list[str] = []
    for field, value in data.items():
        if getattr(item, field, None) != value:
            setattr(item, field, value)
            changed_fields.append(field)

    if changed_fields:
        if resource == "caixas" and any(field in changed_fields for field in ("veiculo_placa", "status", "observacao")):
            await record_caixa_movimento(
                db,
                caixa=item,
                movimento="MOVIMENTACAO",
                veiculo_origem=caixa_veiculo_origem,
                status_origem=caixa_status_origem,
                observacao=data.get("observacao"),
            )
        record_audit_log(
            db,
            action=f"{resource}_alterado",
            actor_user=current_user,
            entity_type=resource,
            entity_id=item.id,
            ip_address=client_ip_from_request(request),
            metadata={"changed_fields": sorted(field for field in changed_fields if field != "senha")},
        )
    await db.commit()
    await db.refresh(item)
    return serialize_item(item, config)


@router.post("/{resource}/{item_id}/senha", response_model=CadastroItemResponse)
async def change_cadastro_password(
    resource: str,
    item_id: int,
    payload: CadastroPasswordPayload,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    if resource not in {"usuarios", "motoristas", "vendedores"}:
        raise HTTPException(status_code=404, detail="Este cadastro nao possui senha")
    config = get_resource(resource)
    item = await get_item_or_404(db, config, item_id)
    senha = str(payload.nova_senha or "").strip()
    if resource == "usuarios":
        if len(senha) < 6:
            raise HTTPException(status_code=422, detail="SENHA deve ter pelo menos 6 caracteres.")
        item.senha = get_password_hash(senha)
    else:
        if not is_valid_motorista_senha(senha):
            raise HTTPException(status_code=422, detail="SENHA invalida. Use 4 a 24 caracteres.")
        item.senha = hash_password_pbkdf2(senha)

    record_audit_log(
        db,
        action=f"{resource}_senha_alterada",
        actor_user=current_user,
        entity_type=resource,
        entity_id=item.id,
        severity="warning",
        ip_address=client_ip_from_request(request),
        metadata={"password_changed": True},
    )
    await db.commit()
    await db.refresh(item)
    return serialize_item(item, config)


@router.delete("/{resource}/{item_id}", response_model=CadastroItemResponse)
async def delete_cadastro_item(
    resource: str,
    item_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_user),
):
    config = get_resource(resource)
    item = await get_item_or_404(db, config, item_id)
    block_reason = await delete_block_reason(db, resource, item)
    if block_reason:
        raise HTTPException(status_code=409, detail=block_reason)

    response = serialize_item(item, config)
    await db.delete(item)
    record_audit_log(
        db,
        action=f"{resource}_excluido",
        actor_user=current_user,
        entity_type=resource,
        entity_id=item_id,
        severity="warning",
        ip_address=client_ip_from_request(request),
        metadata={"deleted": response.data},
    )
    await db.commit()
    return response
