from __future__ import annotations

import asyncio
import os
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from sqlalchemy import text


fd, DB_PATH = tempfile.mkstemp(suffix=".db")
os.close(fd)

os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{Path(DB_PATH).as_posix()}"
os.environ["DEBUG"] = "false"
os.environ["JWT_SECRET_KEY"] = "test-secret"
os.environ["ALLOWED_HOSTS"] = "localhost,127.0.0.1,testserver"
os.environ["RATE_LIMIT_REQUESTS"] = "10000"

from fastapi.testclient import TestClient  # noqa: E402

from backend.config.database import async_session, create_tables, engine  # noqa: E402
from backend.main import app  # noqa: E402
from backend.models.despesa import DespesaDB  # noqa: E402
from backend.models.programacao import ProgramacaoDB, ProgramacaoItemControleDB, ProgramacaoItemDB  # noqa: E402
from backend.models.recebimento import RecebimentoDB  # noqa: E402
from backend.models.system import SistemaLogDB  # noqa: E402
from backend.models.user import UserDB  # noqa: E402
from backend.models.venda_importada import VendaImportadaDB  # noqa: E402
from backend.api.v1.endpoints.centro_custos import transferencias_compra_por_programacao  # noqa: E402
from backend.api.v1.endpoints.despesas import transferencias_operacionais_pdf  # noqa: E402
from backend.services.auth import get_password_hash  # noqa: E402


class BackendAuthFlowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        asyncio.run(create_tables())
        asyncio.run(cls._seed_admin())

    @classmethod
    def tearDownClass(cls):
        asyncio.run(engine.dispose())
        try:
            os.unlink(DB_PATH)
        except OSError:
            pass

    @classmethod
    async def _seed_admin(cls):
        async with async_session() as session:
            session.add(
                UserDB(
                    username="admin",
                    nome="ADMIN",
                    senha=get_password_hash("Admin@123456"),
                    permissoes="ADMIN",
                )
            )
            await session.commit()

    def _auth_headers(self, client: TestClient, username: str = "admin", password: str = "Admin@123456"):
        login = client.post(
            "/api/v1/auth/login",
            data={"username": username, "password": password},
        )
        self.assertEqual(login.status_code, 200)
        token = login.json()["access_token"]
        return {"Authorization": f"Bearer {token}"}

    def test_login_token_and_protected_user_endpoints(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            me = client.get("/api/v1/users/me", headers=headers)
            self.assertEqual(me.status_code, 200)
            self.assertEqual(me.json()["username"], "admin")

            users = client.get("/api/v1/users/", headers=headers)
            self.assertEqual(users.status_code, 200)
            self.assertEqual(users.json()[0]["username"], "admin")

            user = client.get("/api/v1/users/1", headers=headers)
            self.assertEqual(user.status_code, 200)
            self.assertEqual(user.json()["permissoes"], "ADMIN")

            refreshed = client.post("/api/v1/auth/refresh", headers=headers)
            self.assertEqual(refreshed.status_code, 200)
            refreshed_body = refreshed.json()
            self.assertEqual(refreshed_body["token_type"], "bearer")
            self.assertIn("access_token", refreshed_body)
            self.assertGreater(refreshed_body["expires_in"], 0)

            refreshed_headers = {"Authorization": f"Bearer {refreshed_body['access_token']}"}
            me_refreshed = client.get("/api/v1/users/me", headers=refreshed_headers)
            self.assertEqual(me_refreshed.status_code, 200)
            self.assertEqual(me_refreshed.json()["username"], "admin")

            logged_out = client.post("/api/v1/auth/logout", headers=refreshed_headers)
            self.assertEqual(logged_out.status_code, 200)
            self.assertEqual(logged_out.json()["username"], "admin")

    def test_protected_endpoint_rejects_missing_token(self):
        with TestClient(app, base_url="http://testserver") as client:
            response = client.get("/api/v1/users/me")
            self.assertEqual(response.status_code, 401)

    def test_web_shell_is_served(self):
        with TestClient(app, base_url="http://testserver") as client:
            index = client.get("/")
            self.assertEqual(index.status_code, 200)
            self.assertIn("RotaHub", index.text)

            app_index = client.get("/app/index.html")
            self.assertEqual(app_index.status_code, 200)
            self.assertRegex(app_index.text, r'href="styles\.css(?:\?[^"]*)?"')

            script = client.get("/app/app.js")
            self.assertEqual(script.status_code, 200)
            self.assertIn("rotahub_access_token", script.text)

    def test_admin_can_create_update_and_login_new_user(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            created = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "operador",
                    "password": "Operador@123",
                    "nome": "OPERADOR UM",
                    "permissoes": "OPERADOR",
                    "telefone": "88999999999",
                },
            )
            self.assertEqual(created.status_code, 201)
            created_body = created.json()
            self.assertEqual(created_body["username"], "operador")
            self.assertTrue(created_body["is_active"])
            self.assertNotIn("senha", created_body)

            duplicate = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "operador",
                    "password": "OutraSenha@123",
                    "nome": "OUTRO",
                },
            )
            self.assertEqual(duplicate.status_code, 409)

            operador_headers = self._auth_headers(client, "operador", "Operador@123")
            forbidden = client.get("/api/v1/users/", headers=operador_headers)
            self.assertEqual(forbidden.status_code, 403)

            updated = client.patch(
                f"/api/v1/users/{created_body['id']}",
                headers=headers,
                json={
                    "nome": "OPERADOR ATUALIZADO",
                    "password": "NovaSenha@123",
                },
            )
            self.assertEqual(updated.status_code, 200)
            self.assertEqual(updated.json()["nome"], "OPERADOR ATUALIZADO")

            old_login = client.post(
                "/api/v1/auth/login",
                data={"username": "operador", "password": "Operador@123"},
            )
            self.assertEqual(old_login.status_code, 401)
            self._auth_headers(client, "operador", "NovaSenha@123")

    def test_admin_can_deactivate_and_reactivate_user(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            created = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "temporario",
                    "password": "Temporario@123",
                    "nome": "TEMPORARIO",
                    "permissoes": "OPERADOR",
                },
            )
            self.assertEqual(created.status_code, 201)
            user_id = created.json()["id"]

            deactivated = client.delete(f"/api/v1/users/{user_id}", headers=headers)
            self.assertEqual(deactivated.status_code, 200)
            self.assertFalse(deactivated.json()["is_active"])

            inactive_login = client.post(
                "/api/v1/auth/login",
                data={"username": "temporario", "password": "Temporario@123"},
            )
            self.assertEqual(inactive_login.status_code, 401)

            hidden = client.get("/api/v1/users/", headers=headers)
            self.assertEqual(hidden.status_code, 200)
            self.assertNotIn("temporario", [user["username"] for user in hidden.json()])

            visible = client.get("/api/v1/users/?include_inactive=true", headers=headers)
            self.assertEqual(visible.status_code, 200)
            self.assertIn("temporario", [user["username"] for user in visible.json()])

            reactivated = client.patch(
                f"/api/v1/users/{user_id}",
                headers=headers,
                json={"is_active": True},
            )
            self.assertEqual(reactivated.status_code, 200)
            self.assertTrue(reactivated.json()["is_active"])
            self._auth_headers(client, "temporario", "Temporario@123")

            self_deactivate = client.delete("/api/v1/users/1", headers=headers)
            self.assertEqual(self_deactivate.status_code, 400)

    def test_user_management_actions_are_audited(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            created = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "auditado",
                    "password": "Auditado@123",
                    "nome": "AUDITADO",
                    "permissoes": "OPERADOR",
                },
            )
            self.assertEqual(created.status_code, 201)
            user_id = created.json()["id"]

            updated = client.patch(
                f"/api/v1/users/{user_id}",
                headers=headers,
                json={"nome": "AUDITADO ALTERADO"},
            )
            self.assertEqual(updated.status_code, 200)

            deactivated = client.delete(f"/api/v1/users/{user_id}", headers=headers)
            self.assertEqual(deactivated.status_code, 200)

            reactivated = client.patch(
                f"/api/v1/users/{user_id}",
                headers=headers,
                json={"is_active": True},
            )
            self.assertEqual(reactivated.status_code, 200)

            logs = client.get(
                f"/api/v1/audit-logs/?entity_type=user&entity_id={user_id}",
                headers=headers,
            )
            self.assertEqual(logs.status_code, 200)
            actions = [log["action"] for log in logs.json()]
            self.assertIn("usuario_criado", actions)
            self.assertIn("usuario_alterado", actions)
            self.assertIn("usuario_desativado", actions)
            self.assertIn("usuario_reativado", actions)

            update_log = next(log for log in logs.json() if log["action"] == "usuario_alterado")
            self.assertEqual(update_log["user_id"], 1)
            self.assertEqual(update_log["metadata"]["target_username"], "auditado")
            self.assertIn("nome", update_log["metadata"]["changed_fields"])

            operador_headers = self._auth_headers(client, "auditado", "Auditado@123")
            forbidden = client.get("/api/v1/audit-logs/", headers=operador_headers)
            self.assertEqual(forbidden.status_code, 403)

    def test_admin_can_manage_fine_permissions(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            overview = client.get("/api/v1/permissoes/overview", headers=headers)
            self.assertEqual(overview.status_code, 200)
            overview_body = overview.json()
            self.assertIn("sistema", overview_body["modulos"])
            self.assertTrue(any(item["nome"] == "fazer_backup" for item in overview_body["permissoes"]))
            admin_row = next(user for user in overview_body["usuarios"] if user["username"] == "admin")
            self.assertGreaterEqual(admin_row["granted_count"], len(overview_body["permissoes"]))

            created = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "permissao_user",
                    "password": "Permissao@123",
                    "nome": "USUARIO PERMISSAO",
                    "permissoes": "VISUALIZADOR",
                },
            )
            self.assertEqual(created.status_code, 201)
            user_id = created.json()["id"]

            user_permissions = client.get(f"/api/v1/permissoes/usuarios/{user_id}", headers=headers)
            self.assertEqual(user_permissions.status_code, 200)
            permission_names = {item["nome"] for item in user_permissions.json()}
            self.assertIn("visualizar_programacoes", permission_names)
            self.assertIn("gerar_relatorios", permission_names)

            sistema_permissions = client.get("/api/v1/permissoes/disponiveis?modulo=sistema", headers=headers)
            self.assertEqual(sistema_permissions.status_code, 200)
            backup_permission = next(item for item in sistema_permissions.json() if item["nome"] == "fazer_backup")

            granted = client.post(
                f"/api/v1/permissoes/usuarios/{user_id}/conceder",
                headers=headers,
                json={"permissao_id": backup_permission["id"]},
            )
            self.assertEqual(granted.status_code, 200)
            self.assertFalse(granted.json()["ja_existia"])

            duplicate = client.post(
                f"/api/v1/permissoes/usuarios/{user_id}/conceder",
                headers=headers,
                json={"permissao_id": backup_permission["id"]},
            )
            self.assertEqual(duplicate.status_code, 200)
            self.assertTrue(duplicate.json()["ja_existia"])

            after_grant = client.get(f"/api/v1/permissoes/usuarios/{user_id}", headers=headers)
            self.assertEqual(after_grant.status_code, 200)
            self.assertIn("fazer_backup", {item["nome"] for item in after_grant.json()})

            revoked = client.delete(
                f"/api/v1/permissoes/usuarios/{user_id}/{backup_permission['id']}",
                headers=headers,
            )
            self.assertEqual(revoked.status_code, 200)

            after_revoke = client.get(f"/api/v1/permissoes/usuarios/{user_id}", headers=headers)
            self.assertEqual(after_revoke.status_code, 200)
            self.assertNotIn("fazer_backup", {item["nome"] for item in after_revoke.json()})

            profile = client.post(
                f"/api/v1/permissoes/usuarios/{user_id}/perfil",
                headers=headers,
                json={"perfil": "GERENTE"},
            )
            self.assertEqual(profile.status_code, 200)
            self.assertEqual(profile.json()["perfil"], "GERENTE")
            self.assertGreater(profile.json()["permissoes_atribuidas"], len(after_revoke.json()))

            updated_user = client.get(f"/api/v1/users/{user_id}", headers=headers)
            self.assertEqual(updated_user.status_code, 200)
            self.assertEqual(updated_user.json()["permissoes"], "GERENTE")

            module_users = client.get("/api/v1/permissoes/modulos/cadastros/usuarios", headers=headers)
            self.assertEqual(module_users.status_code, 200)
            self.assertIn(user_id, [item["id"] for item in module_users.json()])

            audit = client.get("/api/v1/audit-logs/?action=permissao_concedida", headers=headers)
            self.assertEqual(audit.status_code, 200)
            self.assertTrue(any(log["metadata"].get("permission_id") == backup_permission["id"] for log in audit.json()))

            gerente_headers = self._auth_headers(client, "permissao_user", "Permissao@123")
            forbidden = client.get("/api/v1/permissoes/overview", headers=gerente_headers)
            self.assertEqual(forbidden.status_code, 403)

    def test_admin_can_use_saas_admin_web(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            dashboard = client.get("/api/v1/saas-admin/dashboard", headers=headers)
            self.assertEqual(dashboard.status_code, 200)
            dashboard_body = dashboard.json()
            company_id = dashboard_body["company"]["id"]
            self.assertEqual(dashboard_body["subscription"]["plan_code"], "starter")
            self.assertGreaterEqual(len(dashboard_body["plans"]), 4)
            self.assertTrue(any(plan["code"] == "professional" for plan in dashboard_body["plans"]))
            self.assertIn("companies", dashboard_body)

            features = client.get(f"/api/v1/saas-admin/companies/{company_id}/features", headers=headers)
            self.assertEqual(features.status_code, 200)
            self.assertEqual(features.json()["plan_code"], "starter")
            self.assertIn("advanced_reports", features.json()["features"])

            suspended = client.put(
                f"/api/v1/saas-admin/companies/{company_id}/status",
                headers=headers,
                json={"status": "suspended", "reason": "teste web"},
            )
            self.assertEqual(suspended.status_code, 200)
            self.assertEqual(suspended.json()["company"]["status"], "suspended")

            created_payment = client.post(
                "/api/v1/saas-admin/payments",
                headers=headers,
                json={"company_id": company_id, "amount": 25.5, "due_date": "2026-05-20", "notes": "TESTE WEB"},
            )
            self.assertEqual(created_payment.status_code, 200)
            payment = created_payment.json()["payment"]
            self.assertEqual(payment["status"], "pending")
            self.assertEqual(payment["company_id"], company_id)

            registered = client.post(
                f"/api/v1/saas-admin/payments/{payment['id']}/registrar-pagamento",
                headers=headers,
                json={"method": "manual", "reference": "REC-WEB"},
            )
            self.assertEqual(registered.status_code, 200)
            self.assertEqual(registered.json()["payment"]["status"], "paid")

            company = client.get(f"/api/v1/saas-admin/companies/{company_id}", headers=headers)
            self.assertEqual(company.status_code, 200)
            self.assertEqual(company.json()["status"], "active")

            changed_plan = client.put(
                f"/api/v1/saas-admin/companies/{company_id}/plan",
                headers=headers,
                json={"plan_code": "professional", "reason": "upgrade teste"},
            )
            self.assertEqual(changed_plan.status_code, 200)
            self.assertEqual(changed_plan.json()["subscription"]["plan_code"], "professional")

            payments = client.get(f"/api/v1/saas-admin/payments?company_id={company_id}", headers=headers)
            self.assertEqual(payments.status_code, 200)
            self.assertIn(payment["id"], [item["id"] for item in payments.json()])

            audit = client.get(f"/api/v1/saas-admin/audit-logs?company_id={company_id}", headers=headers)
            self.assertEqual(audit.status_code, 200)
            self.assertIn("pagamento_registrado", [item["action"] for item in audit.json()])

            overdue = client.post(
                "/api/v1/saas-admin/billing/run-overdue-check",
                headers=headers,
                json={"grace_days": 0},
            )
            self.assertEqual(overdue.status_code, 200)
            self.assertIn("suspended", overdue.json()["summary"])

            created_user = client.post(
                "/api/v1/users/",
                headers=headers,
                json={
                    "username": "saas_viewer",
                    "password": "SaasViewer@123",
                    "nome": "SAAS VIEWER",
                    "permissoes": "OPERADOR",
                },
            )
            self.assertEqual(created_user.status_code, 201)
            viewer_headers = self._auth_headers(client, "saas_viewer", "SaasViewer@123")
            forbidden = client.get("/api/v1/saas-admin/dashboard", headers=viewer_headers)
            self.assertEqual(forbidden.status_code, 403)

    def test_admin_can_manage_operational_cadastros(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            created = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA TESTE",
                    "codigo": "MT999",
                    "senha": "1234",
                    "telefone": "88999999999",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(created.status_code, 201)
            item_id = created.json()["id"]
            self.assertEqual(created.json()["data"]["nome"], "MOTORISTA TESTE")
            self.assertEqual(created.json()["data"]["senha"], "******")

            listed = client.get("/api/v1/cadastros/motoristas", headers=headers)
            self.assertEqual(listed.status_code, 200)
            self.assertIn(item_id, [item["id"] for item in listed.json()])

            updated = client.patch(
                f"/api/v1/cadastros/motoristas/{item_id}",
                headers=headers,
                json={"telefone": "88777777777"},
            )
            self.assertEqual(updated.status_code, 200)
            self.assertEqual(updated.json()["data"]["telefone"], "88777777777")

            deleted = client.delete(f"/api/v1/cadastros/motoristas/{item_id}", headers=headers)
            self.assertEqual(deleted.status_code, 200)
            self.assertEqual(deleted.json()["id"], item_id)

            audit = client.get(
                f"/api/v1/audit-logs/?entity_type=motoristas&entity_id={item_id}",
                headers=headers,
            )
            self.assertEqual(audit.status_code, 200)
            self.assertIn("motoristas_criado", [log["action"] for log in audit.json()])

    def test_admin_can_use_clientes_importacao_web(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            saved = client.post(
                "/api/v1/cadastros/clientes/importacao/bulk-upsert",
                headers=headers,
                json={
                    "rows": [
                        {
                            "cod_cliente": "cli-imp-991",
                            "nome_cliente": "cliente importacao web",
                            "endereco": "rua importacao",
                            "telefone": "88999999111",
                            "vendedor": "vend web",
                        },
                        {
                            "cod_cliente": "CLI-IMP-992",
                            "nome_cliente": "CLIENTE IMPORTACAO WEB DOIS",
                            "endereco": "",
                            "telefone": "",
                            "vendedor": "VEND WEB",
                        },
                    ]
                },
            )
            self.assertEqual(saved.status_code, 200)
            self.assertEqual(saved.json()["inseridos"], 2)

            duplicate = client.post(
                "/api/v1/cadastros/clientes/importacao/bulk-upsert",
                headers=headers,
                json={
                    "rows": [
                        {"cod_cliente": "CLI-DUP", "nome_cliente": "CLIENTE DUP A"},
                        {"cod_cliente": "cli-dup", "nome_cliente": "CLIENTE DUP B"},
                    ]
                },
            )
            self.assertEqual(duplicate.status_code, 422)

            updated = client.post(
                "/api/v1/cadastros/clientes/importacao/bulk-upsert",
                headers=headers,
                json={
                    "rows": [
                        {
                            "cod_cliente": "CLI-IMP-991",
                            "nome_cliente": "CLIENTE IMPORTACAO WEB ALTERADO",
                            "endereco": "RUA ALTERADA",
                            "telefone": "88777777111",
                            "vendedor": "VEND ALTERADO",
                        },
                        {"cod_cliente": "", "nome_cliente": "", "endereco": "", "telefone": "", "vendedor": ""},
                    ]
                },
            )
            self.assertEqual(updated.status_code, 200)
            self.assertEqual(updated.json()["atualizados"], 1)

            rows = client.get("/api/v1/cadastros/clientes/importacao/rows", headers=headers)
            self.assertEqual(rows.status_code, 200)
            row = next(item for item in rows.json() if item["cod_cliente"] == "CLI-IMP-991")
            self.assertEqual(row["nome_cliente"], "CLIENTE IMPORTACAO WEB ALTERADO")
            self.assertEqual(row["telefone"], "88777777111")

            from openpyxl import Workbook, load_workbook

            template = client.get("/api/v1/cadastros/clientes/importacao/modelo", headers=headers)
            self.assertEqual(template.status_code, 200)
            self.assertIn("spreadsheetml", template.headers["content-type"])
            template_book = load_workbook(BytesIO(template.content))
            template_sheet = template_book.active
            self.assertEqual(
                [template_sheet.cell(row=1, column=idx).value for idx in range(1, 6)],
                ["COD CLIENTE", "NOME CLIENTE", "ENDERECO", "TELEFONE", "VENDEDOR"],
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["COD CLIENTE", "NOME CLIENTE", "ENDERECO", "TELEFONE", "VENDEDOR"])
            sheet.append(["CLI-IMP-993", "CLIENTE EXCEL WEB", "RUA EXCEL", "88666666111", "VEND EXCEL"])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            uploaded = client.post(
                "/api/v1/cadastros/clientes/importacao/upload",
                headers=headers,
                files={
                    "file": (
                        "clientes.xlsx",
                        buffer.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            self.assertEqual(uploaded.status_code, 201)
            self.assertEqual(uploaded.json()["inseridos"], 1)

            duplicate_workbook = Workbook()
            duplicate_sheet = duplicate_workbook.active
            duplicate_sheet.append(["COD CLIENTE", "NOME CLIENTE", "ENDERECO", "TELEFONE", "VENDEDOR"])
            duplicate_sheet.append(["CLI-IMP-994", "CLIENTE DUP EXCEL A", "RUA DUP A", "", "VEND A"])
            duplicate_sheet.append(["cli-imp-994", "CLIENTE DUP EXCEL B", "", "88999999444", "VEND B"])
            duplicate_sheet.append(["CLI-IMP-995", "", "", "88999999555", ""])
            duplicate_sheet.append(["cli-imp-995", "CLIENTE PARCIAL EXCEL", "RUA PARCIAL", "", "VEND PARCIAL"])
            duplicate_buffer = BytesIO()
            duplicate_workbook.save(duplicate_buffer)
            duplicate_buffer.seek(0)

            uploaded_duplicate = client.post(
                "/api/v1/cadastros/clientes/importacao/upload",
                headers=headers,
                files={
                    "file": (
                        "clientes-duplicados.xlsx",
                        duplicate_buffer.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            self.assertEqual(uploaded_duplicate.status_code, 201)
            self.assertEqual(uploaded_duplicate.json()["inseridos"], 2)
            self.assertEqual(uploaded_duplicate.json()["ignorados"], 2)

            after_upload = client.get("/api/v1/cadastros/clientes/importacao/rows", headers=headers)
            self.assertEqual(after_upload.status_code, 200)
            self.assertIn("CLI-IMP-993", [item["cod_cliente"] for item in after_upload.json()])
            duplicate_row = next(item for item in after_upload.json() if item["cod_cliente"] == "CLI-IMP-994")
            self.assertEqual(duplicate_row["nome_cliente"], "CLIENTE DUP EXCEL B")
            self.assertEqual(duplicate_row["endereco"], "RUA DUP A")
            self.assertEqual(duplicate_row["telefone"], "88999999444")
            partial_row = next(item for item in after_upload.json() if item["cod_cliente"] == "CLI-IMP-995")
            self.assertEqual(partial_row["nome_cliente"], "CLIENTE PARCIAL EXCEL")
            self.assertEqual(partial_row["endereco"], "RUA PARCIAL")
            self.assertEqual(partial_row["telefone"], "88999999555")

    def test_admin_can_create_update_and_delete_programacao(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA PROGRAMACAO",
                    "codigo": "MP901",
                    "senha": "1234",
                    "telefone": "88999999001",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            motorista_codigo = motorista.json()["data"]["codigo"]
            self.assertTrue(motorista_codigo.startswith("MOT-"))
            self.assertNotEqual(motorista_codigo, "MP901")

            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT901", "modelo": "TRUCK", "capacidade_cx": 20},
            )
            self.assertEqual(veiculo.status_code, 201)

            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "UM", "telefone": "88999999002", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "DOIS", "telefone": "88999999003", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            options = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertTrue(options.json()["proximo_codigo"].startswith("PG"))

            rankings = client.get("/api/v1/programacao/rankings?periodo=30", headers=headers)
            self.assertEqual(rankings.status_code, 200)
            rankings_body = rankings.json()
            self.assertEqual(rankings_body["periodo_dias"], 30)
            self.assertIn(motorista_codigo, [item["codigo"] for item in rankings_body["motoristas"]])
            self.assertGreaterEqual(len(rankings_body["ajudantes"]), 2)
            self.assertIn("Top motoristas", rankings_body["resumo_motoristas"])

            imported = client.post(
                "/api/v1/importar-vendas/importar",
                headers=headers,
                json={
                    "rows": [
                        {
                            "pedido": "PED-01",
                            "cliente": "C-PROG-01",
                            "nome_cliente": "CLIENTE PROGRAMACAO",
                            "vendedor": "VENDEDOR A",
                            "produto": "FRANGO",
                            "vr_total": 100,
                            "qnt": 10,
                            "cidade": "RUA TESTE",
                            "observacao": "5 cx 100 kg OBS TESTE",
                        }
                    ]
                },
            )
            self.assertEqual(imported.status_code, 201)
            marked = client.post("/api/v1/importar-vendas/marcar-todas?selected=1", headers=headers)
            self.assertEqual(marked.status_code, 200)
            selected_vendas = client.get("/api/v1/programacao/vendas-selecionadas", headers=headers)
            self.assertEqual(selected_vendas.status_code, 200)
            selected_body = selected_vendas.json()
            self.assertEqual(len(selected_body["itens"]), 1)
            self.assertEqual(selected_body["itens"][0]["qnt_caixas"], 5)
            self.assertEqual(selected_body["itens"][0]["kg"], 100)

            selected_item = selected_body["itens"][0]
            sugestao_sem_veiculo = client.post(
                "/api/v1/programacao/sugestao",
                headers=headers,
                json={"local_rota": "SERRA", "itens": [selected_item]},
            )
            self.assertEqual(sugestao_sem_veiculo.status_code, 422)
            self.assertIn("veiculo", sugestao_sem_veiculo.json()["detail"].lower())

            sugestao = client.post(
                "/api/v1/programacao/sugestao",
                headers=headers,
                json={"veiculo": "PGT901", "local_rota": "SERRA", "itens": [selected_item]},
            )
            self.assertEqual(sugestao.status_code, 200)
            self.assertEqual(sugestao.json()["veiculo"], "PGT901")
            self.assertEqual(sugestao.json()["capacidade_cx"], 20)

            created = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA PROGRAMACAO",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "PGT901",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERTÃO",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 1000,
                    "local_carregamento": "GRANJA A",
                    "adiantamento": 50,
                    "adiantamento_origem": "CAIXA",
                    "itens": [
                        {
                            "cod_cliente": selected_item["cod_cliente"],
                            "nome_cliente": selected_item["nome_cliente"],
                            "produto": selected_item["produto"],
                            "endereco": selected_item["endereco"],
                            "qnt_caixas": selected_item["qnt_caixas"],
                            "kg": selected_item["kg"],
                            "preco": selected_item["preco"],
                            "vendedor": selected_item["vendedor"],
                            "pedido": selected_item["pedido"],
                            "obs": selected_item["obs"],
                        }
                    ],
                    "venda_ids": selected_body["ids"],
                },
            )
            self.assertEqual(created.status_code, 201)
            body = created.json()
            codigo = body["codigo_programacao"]
            self.assertTrue(codigo.startswith("PG"))
            self.assertEqual(body["motorista"], "MOTORISTA PROGRAMACAO")
            self.assertEqual(body["motorista_codigo"], motorista_codigo)
            self.assertEqual(body["local_rota"], "SERTAO")
            self.assertEqual(body["total_caixas"], 5)
            self.assertEqual(body["quilos"], 100)
            self.assertEqual(len(body["itens"]), 1)

            vendas_livres = client.get("/api/v1/importar-vendas/", headers=headers)
            self.assertEqual(vendas_livres.status_code, 200)
            self.assertNotIn("PED-01", [item["pedido"] for item in vendas_livres.json()])

            loaded = client.get(f"/api/v1/programacao/{codigo}", headers=headers)
            self.assertEqual(loaded.status_code, 200)
            self.assertEqual(loaded.json()["itens"][0]["cod_cliente"], "C-PROG-01")

            programacao_pdf = client.get(f"/api/v1/programacao/{codigo}/pdf", headers=headers)
            self.assertEqual(programacao_pdf.status_code, 200)
            self.assertEqual(programacao_pdf.headers["content-type"], "application/pdf")
            self.assertIn("PROGRAMACAO_", programacao_pdf.headers["content-disposition"])
            self.assertTrue(programacao_pdf.content.startswith(b"%PDF"))
            self.assertIn(b"PROGRAMACAO", programacao_pdf.content)
            self.assertIn(b"C-PROG-01", programacao_pdf.content)

            recibo = client.get(f"/api/v1/programacao/{codigo}/recibo-adiantamento-pdf", headers=headers)
            self.assertEqual(recibo.status_code, 200)
            self.assertEqual(recibo.headers["content-type"], "application/pdf")
            self.assertIn("RECIBO_ADIANTAMENTO_", recibo.headers["content-disposition"])
            self.assertTrue(recibo.content.startswith(b"%PDF"))
            self.assertIn(b"RECIBO DE ADIANTAMENTO", recibo.content)

            romaneios = client.get(f"/api/v1/programacao/{codigo}/romaneios-pdf", headers=headers)
            self.assertEqual(romaneios.status_code, 200)
            self.assertEqual(romaneios.headers["content-type"], "application/pdf")
            self.assertIn("ROMANEIOS_", romaneios.headers["content-disposition"])
            self.assertTrue(romaneios.content.startswith(b"%PDF"))
            self.assertIn(b"ROMANEIO DE ENTREGA", romaneios.content)

            updated = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "codigo_programacao": codigo,
                    "motorista": "MOTORISTA PROGRAMACAO",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "PGT901",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "CX",
                    "caixas_estimado": 6,
                    "local_carregamento": "GRANJA B",
                    "itens": [
                        {
                            "cod_cliente": "C-PROG-01",
                            "nome_cliente": "CLIENTE PROGRAMACAO",
                            "qnt_caixas": 6,
                            "kg": 0,
                            "preco": 12,
                        }
                    ],
                },
            )
            self.assertEqual(updated.status_code, 201)
            self.assertEqual(updated.json()["tipo_estimativa"], "CX")
            self.assertEqual(updated.json()["caixas_estimado"], 6)
            self.assertEqual(updated.json()["total_caixas"], 6)
            self.assertEqual(updated.json()["local_carregamento"], "GRANJA B")

            deleted = client.delete(f"/api/v1/programacao/{codigo}?devolver_vendas=true", headers=headers)
            self.assertEqual(deleted.status_code, 200)
            self.assertEqual(deleted.json()["codigo_programacao"], codigo)

            missing = client.get(f"/api/v1/programacao/{codigo}", headers=headers)
            self.assertEqual(missing.status_code, 404)

    def test_programacao_options_only_show_active_available_resources(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            busy_motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={"nome": "MOTORISTA OCUPADO", "codigo": "MO901", "senha": "1234", "telefone": "88999999101", "status": "ATIVO"},
            )
            self.assertEqual(busy_motorista.status_code, 201)
            busy_motorista_codigo = busy_motorista.json()["data"]["codigo"]
            free_motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={"nome": "MOTORISTA LIVRE", "codigo": "ML901", "senha": "1234", "telefone": "88999999102", "status": "ATIVO"},
            )
            self.assertEqual(free_motorista.status_code, 201)
            free_motorista_codigo = free_motorista.json()["data"]["codigo"]
            inactive_motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={"nome": "MOTORISTA INATIVO", "codigo": "MI901", "senha": "1234", "telefone": "88999999103", "status": "INATIVO"},
            )
            self.assertEqual(inactive_motorista.status_code, 201)
            inactive_motorista_codigo = inactive_motorista.json()["data"]["codigo"]

            busy_veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "BUS901", "modelo": "TRUCK", "capacidade_cx": 20, "status": "ATIVO"},
            )
            self.assertEqual(busy_veiculo.status_code, 201)
            free_veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "FRE901", "modelo": "TRUCK", "capacidade_cx": 20, "status": "ATIVO"},
            )
            self.assertEqual(free_veiculo.status_code, 201)
            inactive_veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "INA901", "modelo": "TRUCK", "capacidade_cx": 20, "status": "DESATIVADO"},
            )
            self.assertEqual(inactive_veiculo.status_code, 201)

            busy_ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "OCUPADO UM", "telefone": "88999999104", "status": "ATIVO"},
            )
            self.assertEqual(busy_ajudante_1.status_code, 201)
            busy_ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "OCUPADO DOIS", "telefone": "88999999105", "status": "ATIVO"},
            )
            self.assertEqual(busy_ajudante_2.status_code, 201)
            free_ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "LIVRE UM", "telefone": "88999999106", "status": "ATIVO"},
            )
            self.assertEqual(free_ajudante_1.status_code, 201)
            free_ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "LIVRE DOIS", "telefone": "88999999107", "status": "ATIVO"},
            )
            self.assertEqual(free_ajudante_2.status_code, 201)
            inactive_ajudante = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "INATIVO", "telefone": "88999999108", "status": "DESATIVADO"},
            )
            self.assertEqual(inactive_ajudante.status_code, 201)

            created = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA OCUPADO",
                    "motorista_codigo": busy_motorista_codigo,
                    "veiculo": "BUS901",
                    "ajudantes": [str(busy_ajudante_1.json()["id"]), str(busy_ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 100,
                    "local_carregamento": "GRANJA",
                    "itens": [],
                },
            )
            self.assertEqual(created.status_code, 201)

            options = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            body = options.json()
            self.assertNotIn(busy_motorista_codigo, [item["codigo"] for item in body["motoristas"]])
            self.assertNotIn(inactive_motorista_codigo, [item["codigo"] for item in body["motoristas"]])
            self.assertIn(free_motorista_codigo, [item["codigo"] for item in body["motoristas"]])
            self.assertNotIn("BUS901", [item["placa"] for item in body["veiculos"]])
            self.assertNotIn("INA901", [item["placa"] for item in body["veiculos"]])
            self.assertIn("FRE901", [item["placa"] for item in body["veiculos"]])
            self.assertNotIn(str(busy_ajudante_1.json()["id"]), [item["id"] for item in body["ajudantes"]])
            self.assertNotIn(str(inactive_ajudante.json()["id"]), [item["id"] for item in body["ajudantes"]])
            self.assertIn(str(free_ajudante_1.json()["id"]), [item["id"] for item in body["ajudantes"]])

            conflict = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA OCUPADO",
                    "motorista_codigo": busy_motorista_codigo,
                    "veiculo": "FRE901",
                    "ajudantes": [str(free_ajudante_1.json()["id"]), str(free_ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 100,
                    "local_carregamento": "GRANJA",
                    "itens": [],
                },
            )
            self.assertEqual(conflict.status_code, 409)
            self.assertIn("Motorista", conflict.json()["detail"])

    def test_programacao_options_release_resources_when_any_final_status_is_set(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={"nome": "MOTORISTA FINAL", "codigo": "MF901", "senha": "1234", "telefone": "88999999201", "status": "ATIVO"},
            )
            self.assertEqual(motorista.status_code, 201)
            motorista_codigo = motorista.json()["data"]["codigo"]
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "FIN901", "modelo": "TRUCK", "capacidade_cx": 20, "status": "ATIVO"},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "FINAL UM", "telefone": "88999999202", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "FINAL DOIS", "telefone": "88999999203", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            created = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA FINAL",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "FIN901",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 100,
                    "local_carregamento": "GRANJA",
                    "itens": [],
                },
            )
            self.assertEqual(created.status_code, 201)
            async def marcar_status_inconsistente():
                async with async_session() as session:
                    obj = await session.get(ProgramacaoDB, int(created.json()["id"]))
                    obj.status = "FINALIZADA"
                    obj.status_operacional = "EM_ROTA"
                    obj.finalizada_no_app = 0
                    obj.prestacao_status = "PENDENTE"
                    await session.commit()

            asyncio.run(marcar_status_inconsistente())

            options = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            body = options.json()
            self.assertIn(motorista_codigo, [item["codigo"] for item in body["motoristas"]])
            self.assertIn("FIN901", [item["placa"] for item in body["veiculos"]])
            self.assertIn(str(ajudante_1.json()["id"]), [item["id"] for item in body["ajudantes"]])

            reuse = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA FINAL",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "FIN901",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 100,
                    "local_carregamento": "GRANJA",
                    "itens": [],
                },
            )
            self.assertEqual(reuse.status_code, 201)

    def test_admin_can_monitor_and_register_rota_gps(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA ROTAS",
                    "codigo": "MR903",
                    "senha": "1234",
                    "telefone": "88999999021",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            motorista_codigo = motorista.json()["data"]["codigo"]
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT903", "modelo": "TRUCK", "capacidade_cx": 25},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "ROTA UM", "telefone": "88999999022", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "ROTA DOIS", "telefone": "88999999023", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            programacao = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA ROTAS",
                    "motorista_codigo": "MR903",
                    "veiculo": "PGT903",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "CX",
                    "caixas_estimado": 12,
                    "local_carregamento": "GRANJA ROTAS",
                    "itens": [],
                },
            )
            self.assertEqual(programacao.status_code, 201)
            codigo = programacao.json()["codigo_programacao"]

            monitoramento = client.get("/api/v1/rotas/monitoramento", headers=headers)
            self.assertEqual(monitoramento.status_code, 200)
            rota = next(item for item in monitoramento.json() if item["codigo_programacao"] == codigo)
            self.assertEqual(rota["motorista"], "MOTORISTA ROTAS")
            self.assertEqual(rota["veiculo"], "PGT903")
            self.assertEqual(rota["status"], "ATIVA")
            self.assertIsNone(rota["lat"])
            self.assertIsNone(rota["lon"])

            gps = client.post(
                f"/api/v1/rotas/{codigo}/gps",
                headers=headers,
                json={
                    "lat": -3.7319,
                    "lon": -38.5267,
                    "speed": 45.5,
                    "accuracy": 8.2,
                    "recorded_at": "2026-05-07 10:00:00",
                },
            )
            self.assertEqual(gps.status_code, 201)
            self.assertEqual(gps.json()["codigo_programacao"], codigo)

            atualizado = client.get("/api/v1/rotas/monitoramento", headers=headers)
            self.assertEqual(atualizado.status_code, 200)
            rota_atualizada = next(item for item in atualizado.json() if item["codigo_programacao"] == codigo)
            self.assertAlmostEqual(rota_atualizada["lat"], -3.7319)
            self.assertAlmostEqual(rota_atualizada["lon"], -38.5267)
            self.assertEqual(rota_atualizada["speed"], 45.5)
            self.assertEqual(rota_atualizada["accuracy"], 8.2)
            self.assertEqual(rota_atualizada["recorded_at"], "2026-05-07 10:00:00")

    def test_home_dashboard_overview_and_route_preview(self):
        codigo = "HOME991"

        async def seed_home_data():
            async with async_session() as session:
                session.add(
                    ProgramacaoDB(
                        codigo_programacao=codigo,
                        data_criacao="2026-05-08",
                        motorista="MOTORISTA HOME",
                        motorista_codigo="MH991",
                        veiculo="PGH991",
                        equipe="AJUDANTE HOME UM|AJUDANTE HOME DOIS",
                        kg_estimado=75,
                        status="ATIVA",
                        prestacao_status="PENDENTE",
                        local_rota="SERRA",
                        local_carregamento="GRANJA HOME",
                        nf_numero="NF-HOME-991",
                        data_saida="2026-05-08",
                        hora_saida="07:45:00",
                        inicio_carregamento="06:10",
                        fim_carregamento="07:20",
                        total_caixas=3,
                        quilos=75,
                        adiantamento=20,
                        nf_kg=80,
                        nf_caixas=3,
                        nf_kg_carregado=75,
                        nf_saldo=5,
                        kg_carregado=75,
                        caixas_carregadas=3,
                        media=2.5,
                        aves_caixa_final=28,
                        km_inicial=1200,
                    )
                )
                session.add(
                    ProgramacaoItemDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-HOME-1",
                        nome_cliente="CLIENTE HOME",
                        qnt_caixas=3,
                        kg=75,
                        preco=12.5,
                        vendedor="VENDEDOR HOME",
                        pedido="PED-HOME-1",
                        produto="FRANGO",
                    )
                )
                session.add(
                    ProgramacaoItemControleDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-HOME-1",
                        pedido="PED-HOME-1",
                        status_pedido="ENTREGUE",
                        caixas_atual=3,
                        preco_atual=13,
                        valor_recebido=10,
                        forma_recebimento="PIX",
                        obs_recebimento="RECEBIDO NO APP",
                        mortalidade_aves=1,
                        media_aplicada=25,
                        peso_previsto=75,
                        alterado_em="2026-05-08 13:20:00",
                        alterado_por="MOTORISTA HOME",
                        alteracao_tipo="ENTREGA",
                        alteracao_detalhe="BAIXA PELO APP",
                        lat_evento=-3.7319,
                        lon_evento=-38.5267,
                        endereco_evento="RUA HOME, 123",
                        cidade_evento="FORTALEZA",
                        bairro_evento="CENTRO",
                        ordem_sugerida=1,
                        eta="13:20",
                        distancia=12.5,
                        confianca_localizacao=0.92,
                    )
                )
                session.add(
                    RecebimentoDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-HOME-1",
                        nome_cliente="CLIENTE HOME",
                        valor=80,
                        forma_pagamento="PIX",
                        data_registro="2026-05-08 12:00:00",
                    )
                )
                session.add(
                    DespesaDB(
                        codigo_programacao=codigo,
                        descricao="COMBUSTIVEL HOME",
                        valor=30,
                        data_registro="2026-05-08 11:00:00",
                        categoria="COMBUSTIVEL",
                        motorista="MOTORISTA HOME",
                        veiculo="PGH991",
                    )
                )
                session.add(
                    VendaImportadaDB(
                        pedido="PED-HOME-1",
                        data_venda="2026-05-08",
                        cliente="CLI-HOME-1",
                        nome_cliente="CLIENTE HOME",
                        produto="FRANGO",
                        vr_total=80,
                        qnt=75,
                    )
                )
                await session.commit()

        asyncio.run(seed_home_data())

        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            overview = client.get("/api/v1/home/overview?limit=100", headers=headers)
            self.assertEqual(overview.status_code, 200)
            body = overview.json()
            self.assertGreaterEqual(body["metrics"]["programacoes_ativas"], 1)
            self.assertGreaterEqual(body["metrics"]["vendas_importadas"], 1)
            self.assertGreaterEqual(body["metrics"]["clientes_ativos"], 1)
            self.assertGreaterEqual(body["pendencias"]["rotas_abertas"], 1)
            self.assertEqual(body["sistema"]["api"], "online")
            self.assertIn(codigo, [item["codigo_programacao"] for item in body["rotas"]])

            preview = client.get(f"/api/v1/home/rotas/{codigo}/preview", headers=headers)
            self.assertEqual(preview.status_code, 200)
            detail = preview.json()
            self.assertEqual(detail["programacao"]["codigo_programacao"], codigo)
            self.assertEqual(detail["programacao"]["num_nf"], "NF-HOME-991")
            self.assertEqual(detail["programacao"]["data_saida"], "2026-05-08")
            self.assertEqual(detail["programacao"]["hora_saida"], "07:45:00")
            self.assertEqual(detail["programacao"]["inicio_carregamento"], "06:10")
            self.assertEqual(detail["programacao"]["fim_carregamento"], "07:20")
            self.assertEqual(detail["programacao"]["local_carregamento"], "GRANJA HOME")
            self.assertEqual(detail["programacao"]["caixas_carregadas"], 3)
            self.assertEqual(detail["programacao"]["nf_caixas"], 3)
            self.assertEqual(detail["programacao"]["kg_carregado"], 75)
            self.assertEqual(detail["programacao"]["nf_saldo"], 5)
            self.assertEqual(detail["programacao"]["aves_caixa_final"], 28)
            self.assertEqual(detail["programacao"]["km_inicial"], 1200)
            self.assertEqual(detail["resumo"]["clientes"], 1)
            self.assertEqual(detail["resumo"]["caixas"], 3)
            self.assertEqual(detail["resumo"]["caixas_programadas"], 3)
            self.assertEqual(detail["resumo"]["caixas_entregues"], 3)
            self.assertEqual(detail["resumo"]["recebido"], 90)
            self.assertEqual(detail["resumo"]["despesas"], 30)
            self.assertEqual(detail["itens"][0]["status_pedido"], "ENTREGUE")
            self.assertEqual(detail["resumo"]["entregues"], 1)
            self.assertEqual(detail["resumo"]["com_localizacao"], 1)
            self.assertAlmostEqual(detail["itens"][0]["lat_evento"], -3.7319)
            self.assertAlmostEqual(detail["itens"][0]["lon_evento"], -38.5267)
            self.assertTrue(detail["itens"][0]["tem_localizacao"])
            self.assertEqual(detail["itens"][0]["alterado_em"], "2026-05-08 13:20:00")
            self.assertEqual(detail["itens"][0]["status_origem"], "APP MOTORISTA")
            self.assertEqual(detail["itens"][0]["endereco_evento"], "RUA HOME, 123")
            self.assertEqual(detail["recebimentos"][0]["forma_pagamento"], "PIX")
            self.assertEqual(detail["despesas"][0]["categoria"], "COMBUSTIVEL")

    def test_admin_can_read_escala_summary(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA ESCALA",
                    "codigo": "ME904",
                    "senha": "1234",
                    "telefone": "88999999031",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            motorista_codigo = motorista.json()["data"]["codigo"]
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT904", "modelo": "TRUCK", "capacidade_cx": 25},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "ESCALA UM", "telefone": "88999999032", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "ESCALA DOIS", "telefone": "88999999033", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            programacao = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA ESCALA",
                    "motorista_codigo": "ME904",
                    "veiculo": "PGT904",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERTAO",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 900,
                    "local_carregamento": "GRANJA ESCALA",
                    "itens": [],
                },
            )
            self.assertEqual(programacao.status_code, 201)
            programacao_id = programacao.json()["id"]

            async def set_escala_fields():
                async with async_session() as session:
                    item = await session.get(ProgramacaoDB, programacao_id)
                    item.status_operacional = "CARREGADA"
                    item.data_saida = "2026-05-06"
                    item.hora_saida = "08:00:00"
                    item.data_chegada = "2026-05-06"
                    item.hora_chegada = "16:30:00"
                    item.km_inicial = 100
                    item.km_final = 260
                    item.km_rodado = 160
                    item.litros = 20
                    item.media_km_l = 8
                    item.mortalidade_transbordo_aves = 4
                    await session.commit()

            asyncio.run(set_escala_fields())

            escala = client.get("/api/v1/escala/resumo?periodo=TODAS&status=CARREGADA", headers=headers)
            self.assertEqual(escala.status_code, 200)
            body = escala.json()
            self.assertEqual(body["kpis"]["rotas"], 1)
            self.assertEqual(body["kpis"]["motoristas"], 1)
            self.assertEqual(body["kpis"]["ajudantes"], 2)
            self.assertEqual(body["kpis"]["km_total"], 160.0)
            self.assertEqual(body["kpis"]["km_medio_motorista"], 160.0)
            self.assertEqual(body["kpis"]["media_km_l"], 8.0)
            self.assertEqual(body["kpis"]["horas_medias_motorista"], 8.5)
            self.assertEqual(body["kpis"]["mortalidade_media"], 4.0)
            self.assertEqual(body["motoristas"][0]["nome"], "MOTORISTA ESCALA")
            self.assertEqual(body["motoristas"][0]["local"], "SERTAO")
            self.assertEqual({item["nome"] for item in body["ajudantes"]}, {"AJUDANTE ESCALA UM", "AJUDANTE ESCALA DOIS"})
            self.assertEqual(body["periodo"], "TODAS")
            self.assertEqual(body["status"], "CARREGADA")

            pdf = client.get("/api/v1/escala/pdf?periodo=TODAS&status=CARREGADA", headers=headers)
            self.assertEqual(pdf.status_code, 200)
            self.assertEqual(pdf.headers["content-type"], "application/pdf")
            self.assertIn("ESCALA_", pdf.headers["content-disposition"])
            self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_escala_folga_blocks_programacao_resource(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA FOLGA WEB",
                    "codigo": "MF906",
                    "senha": "1234",
                    "telefone": "88999999061",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            motorista_codigo = motorista.json()["data"]["codigo"]
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT906", "modelo": "TRUCK", "capacidade_cx": 25},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "FOLGA WEB UM", "telefone": "88999999062", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "FOLGA WEB DOIS", "telefone": "88999999063", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            pessoas = client.get("/api/v1/escala/pessoas?tipo=MOTORISTA", headers=headers)
            self.assertEqual(pessoas.status_code, 200)
            self.assertIn("MOTORISTA FOLGA WEB", {item["pessoa_nome"] for item in pessoas.json()})

            folga = client.post(
                "/api/v1/escala/folgas",
                headers=headers,
                json={
                    "tipo": "MOTORISTA",
                    "pessoa_id": str(motorista.json()["id"]),
                    "pessoa_codigo": motorista_codigo,
                    "pessoa_nome": "MOTORISTA FOLGA WEB",
                    "data_inicio": "2026-05-27",
                    "data_fim": "2026-05-28",
                    "motivo": "DESCANSO",
                },
            )
            self.assertEqual(folga.status_code, 200)

            duplicate = client.post(
                "/api/v1/escala/folgas",
                headers=headers,
                json={
                    "tipo": "MOTORISTA",
                    "pessoa_codigo": motorista_codigo,
                    "pessoa_nome": "MOTORISTA FOLGA WEB",
                    "data_inicio": "2026-05-28",
                    "data_fim": "2026-05-29",
                },
            )
            self.assertEqual(duplicate.status_code, 409)

            options = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertNotIn("MOTORISTA FOLGA WEB", {item["nome"] for item in options.json()["motoristas"]})

            blocked = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA FOLGA WEB",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "PGT906",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERTAO",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 900,
                    "local_carregamento": "GRANJA FOLGA",
                    "itens": [],
                },
            )
            self.assertEqual(blocked.status_code, 409)
            self.assertIn("folga", blocked.json()["detail"].lower())

            resumo = client.get("/api/v1/escala/resumo?periodo=TODAS&status=TODOS", headers=headers)
            self.assertEqual(resumo.status_code, 200)
            self.assertTrue(any(item["pessoa_nome"] == "MOTORISTA FOLGA WEB" for item in resumo.json()["folgas"]))
            self.assertEqual(resumo.json()["kpis"]["folgas_motoristas"], 1)
            self.assertEqual(resumo.json()["kpis"]["folgas_ajudantes"], 0)

            encerrada = client.patch(f"/api/v1/escala/folgas/{folga.json()['id']}/encerrar", headers=headers)
            self.assertEqual(encerrada.status_code, 200)
            encerrada_novamente = client.patch(f"/api/v1/escala/folgas/{folga.json()['id']}/encerrar", headers=headers)
            self.assertEqual(encerrada_novamente.status_code, 409)
            options_after = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options_after.status_code, 200)
            self.assertIn("MOTORISTA FOLGA WEB", {item["nome"] for item in options_after.json()["motoristas"]})

            folga_ajudante = client.post(
                "/api/v1/escala/folgas",
                headers=headers,
                json={
                    "tipo": "AJUDANTE",
                    "pessoa_id": str(ajudante_1.json()["id"]),
                    "pessoa_nome": "AJUDANTE FOLGA WEB UM",
                    "data_inicio": "2026-05-27",
                    "data_fim": "2026-05-28",
                    "motivo": "DESCANSO",
                },
            )
            self.assertEqual(folga_ajudante.status_code, 200)
            options_aj = client.get("/api/v1/programacao/options", headers=headers)
            self.assertEqual(options_aj.status_code, 200)
            self.assertNotIn(
                str(ajudante_1.json()["id"]),
                {item["id"] for item in options_aj.json()["ajudantes"]},
            )

            blocked_aj = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA FOLGA WEB",
                    "motorista_codigo": motorista_codigo,
                    "veiculo": "PGT906",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERTAO",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 900,
                    "local_carregamento": "GRANJA FOLGA",
                    "itens": [],
                },
            )
            self.assertEqual(blocked_aj.status_code, 409)
            self.assertIn("folga", blocked_aj.json()["detail"].lower())

            resumo_aj = client.get("/api/v1/escala/resumo?periodo=TODAS&status=TODOS", headers=headers)
            self.assertEqual(resumo_aj.status_code, 200)
            self.assertEqual(resumo_aj.json()["kpis"]["folgas_motoristas"], 0)
            self.assertEqual(resumo_aj.json()["kpis"]["folgas_ajudantes"], 1)

    def test_admin_can_manage_recebimentos_flow(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA RECEBIMENTOS",
                    "codigo": "MR905",
                    "senha": "1234",
                    "telefone": "88999999041",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT905", "modelo": "TRUCK", "capacidade_cx": 25},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "REC UM", "telefone": "88999999042", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "REC DOIS", "telefone": "88999999043", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            programacao = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA RECEBIMENTOS",
                    "motorista_codigo": "MR905",
                    "veiculo": "PGT905",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 500,
                    "local_carregamento": "GRANJA RECEBIMENTOS",
                    "itens": [
                        {
                            "cod_cliente": "REC001",
                            "nome_cliente": "CLIENTE RECEBIMENTOS",
                            "produto": "FRANGO",
                            "qnt_caixas": 2,
                            "kg": 50,
                            "preco": 20,
                        }
                    ],
                },
            )
            self.assertEqual(programacao.status_code, 201)
            codigo = programacao.json()["codigo_programacao"]
            programacao_id = programacao.json()["id"]

            async def finalizar_rota():
                async with async_session() as session:
                    item = await session.get(ProgramacaoDB, programacao_id)
                    item.status_operacional = "FINALIZADA"
                    item.finalizada_no_app = 1
                    await session.commit()

            asyncio.run(finalizar_rota())

            options = client.get("/api/v1/recebimentos/programacoes", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertIn(codigo, [item["codigo_programacao"] for item in options.json()])

            cabecalho = client.put(
                f"/api/v1/recebimentos/{codigo}/cabecalho",
                headers=headers,
                json={
                    "data_saida": "2026-05-07",
                    "hora_saida": "07:30:00",
                    "data_chegada": "2026-05-08",
                    "hora_chegada": "09:00:00",
                    "diaria_motorista_valor": 80,
                },
            )
            self.assertEqual(cabecalho.status_code, 200)
            self.assertEqual(cabecalho.json()["diarias"]["qtd_diarias"], 1.5)
            self.assertEqual(cabecalho.json()["diarias"]["total_motorista"], 120)
            self.assertEqual(cabecalho.json()["diarias"]["total_ajudantes"], 210)

            despesas_diarias = client.get(f"/api/v1/despesas/{codigo}/bundle", headers=headers)
            self.assertEqual(despesas_diarias.status_code, 200)
            despesas_rows = despesas_diarias.json()["despesas"]
            despesas_by_desc = {item["descricao"]: item for item in despesas_rows}
            self.assertEqual(despesas_by_desc["DIARIAS MOTORISTA"]["valor"], 120)
            self.assertEqual(despesas_by_desc["DIARIAS AJUDANTES"]["valor"], 210)
            self.assertEqual(despesas_diarias.json()["financeiro"]["total_despesas"], 330)

            recebido = client.post(
                f"/api/v1/recebimentos/{codigo}/recebimentos",
                headers=headers,
                json={
                    "cod_cliente": "REC001",
                    "nome_cliente": "CLIENTE RECEBIMENTOS",
                    "valor": 123.45,
                    "forma_pagamento": "PIX",
                    "observacao": "PAGO",
                },
            )
            self.assertEqual(recebido.status_code, 201)
            self.assertEqual(recebido.json()["forma_pagamento"], "PIX")

            bundle = client.get(f"/api/v1/recebimentos/{codigo}/bundle", headers=headers)
            self.assertEqual(bundle.status_code, 200)
            self.assertEqual(bundle.json()["total_recebido"], 123.45)
            cliente = next(item for item in bundle.json()["clientes"] if item["cod_cliente"] == "REC001")
            self.assertEqual(cliente["valor"], 123.45)
            self.assertEqual(cliente["forma_pagamento"], "PIX")

            pdf = client.get(f"/api/v1/recebimentos/{codigo}/pdf", headers=headers)
            self.assertEqual(pdf.status_code, 200)
            self.assertEqual(pdf.headers["content-type"], "application/pdf")
            self.assertIn("RECEBIMENTOS_", pdf.headers["content-disposition"])
            self.assertTrue(pdf.content.startswith(b"%PDF"))

            manual = client.post(
                f"/api/v1/recebimentos/{codigo}/clientes/manual",
                headers=headers,
                json={"cod_cliente": "REC999", "nome_cliente": "CLIENTE MANUAL REC"},
            )
            self.assertEqual(manual.status_code, 201)
            bundle_manual = client.get(f"/api/v1/recebimentos/{codigo}/bundle", headers=headers)
            self.assertIn("REC999", [item["cod_cliente"] for item in bundle_manual.json()["clientes"]])

            zerado = client.delete(f"/api/v1/recebimentos/{codigo}/recebimentos/REC001", headers=headers)
            self.assertEqual(zerado.status_code, 200)
            cliente_zerado = next(item for item in zerado.json()["clientes"] if item["cod_cliente"] == "REC001")
            self.assertEqual(cliente_zerado["valor"], 0)
            self.assertEqual(zerado.json()["total_recebido"], 0)

            manual_recebido = client.post(
                f"/api/v1/recebimentos/{codigo}/recebimentos",
                headers=headers,
                json={
                    "cod_cliente": "REC998",
                    "nome_cliente": "CLIENTE MANUAL COM RECEBIMENTO",
                    "valor": 10,
                    "forma_pagamento": "DINHEIRO",
                    "observacao": "INSERIDO NO LANCAMENTO",
                },
            )
            self.assertEqual(manual_recebido.status_code, 201)
            self.assertEqual(manual_recebido.json()["cod_cliente"], "REC998")
            bundle_manual_recebido = client.get(f"/api/v1/recebimentos/{codigo}/bundle", headers=headers)
            cliente_manual_recebido = next(item for item in bundle_manual_recebido.json()["clientes"] if item["cod_cliente"] == "REC998")
            self.assertEqual(cliente_manual_recebido["nome_cliente"], "CLIENTE MANUAL COM RECEBIMENTO")
            self.assertEqual(cliente_manual_recebido["valor"], 10)

            manual_zerado = client.delete(f"/api/v1/recebimentos/{codigo}/recebimentos/REC998", headers=headers)
            self.assertEqual(manual_zerado.status_code, 200)
            self.assertEqual(manual_zerado.json()["total_recebido"], 0)

            async def fechar_prestacao():
                async with async_session() as session:
                    item = await session.get(ProgramacaoDB, programacao_id)
                    item.prestacao_status = "FECHADA"
                    await session.commit()

            asyncio.run(fechar_prestacao())
            recebimentos_fechado = client.get(f"/api/v1/recebimentos/{codigo}/bundle", headers=headers)
            self.assertEqual(recebimentos_fechado.status_code, 200)
            self.assertEqual(recebimentos_fechado.json()["cabecalho"]["prestacao_status"], "FECHADA")
            self.assertTrue(recebimentos_fechado.json()["cabecalho"]["fechada"])
            despesas_fechado = client.get(f"/api/v1/despesas/{codigo}/bundle", headers=headers)
            self.assertEqual(despesas_fechado.status_code, 200)
            self.assertEqual(despesas_fechado.json()["cabecalho"]["prestacao_status"], "FECHADA")
            blocked = client.post(
                f"/api/v1/recebimentos/{codigo}/recebimentos",
                headers=headers,
                json={
                    "cod_cliente": "REC001",
                    "nome_cliente": "CLIENTE RECEBIMENTOS",
                    "valor": 1,
                    "forma_pagamento": "DINHEIRO",
                },
            )
            self.assertEqual(blocked.status_code, 409)

    def test_admin_can_manage_despesas_flow(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA DESPESAS",
                    "codigo": "MD906",
                    "senha": "1234",
                    "telefone": "88999999051",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT906", "modelo": "TRUCK", "capacidade_cx": 25},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "DESP UM", "telefone": "88999999052", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "DESP DOIS", "telefone": "88999999053", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            programacao = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA DESPESAS",
                    "motorista_codigo": "MD906",
                    "veiculo": "PGT906",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERTAO",
                    "tipo_estimativa": "KG",
                    "kg_estimado": 800,
                    "local_carregamento": "GRANJA DESPESAS",
                    "adiantamento": 20,
                    "itens": [
                        {
                            "cod_cliente": "DESP001",
                            "nome_cliente": "CLIENTE DESPESAS",
                            "produto": "FRANGO",
                            "qnt_caixas": 4,
                            "kg": 80,
                            "preco": 20,
                        }
                    ],
                },
            )
            self.assertEqual(programacao.status_code, 201)
            codigo = programacao.json()["codigo_programacao"]
            programacao_id = programacao.json()["id"]

            async def finalizar_rota():
                async with async_session() as session:
                    item = await session.get(ProgramacaoDB, programacao_id)
                    item.status_operacional = "FINALIZADA"
                    item.finalizada_no_app = 1
                    await session.commit()

            asyncio.run(finalizar_rota())

            recebimento = client.post(
                f"/api/v1/recebimentos/{codigo}/recebimentos",
                headers=headers,
                json={
                    "cod_cliente": "DESP001",
                    "nome_cliente": "CLIENTE DESPESAS",
                    "valor": 200,
                    "forma_pagamento": "DINHEIRO",
                },
            )
            self.assertEqual(recebimento.status_code, 201)

            options = client.get("/api/v1/despesas/programacoes", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertIn(codigo, [item["codigo_programacao"] for item in options.json()])

            despesa = client.post(
                f"/api/v1/despesas/{codigo}/despesas",
                headers=headers,
                json={
                    "descricao": "COMBUSTIVEL",
                    "valor": 50,
                    "categoria": "COMBUSTIVEL",
                    "observacao": "POSTO",
                },
            )
            self.assertEqual(despesa.status_code, 201)
            despesa_id = despesa.json()["id"]

            atualizada = client.patch(
                f"/api/v1/despesas/{codigo}/despesas/{despesa_id}",
                headers=headers,
                json={"valor": 70, "observacao": "POSTO ATUALIZADO"},
            )
            self.assertEqual(atualizada.status_code, 200)
            self.assertEqual(atualizada.json()["valor"], 70)

            rota = client.put(
                f"/api/v1/despesas/{codigo}/rota",
                headers=headers,
                json={"km_inicial": 100, "km_final": 220, "litros": 20, "rota_observacao": "OK"},
            )
            self.assertEqual(rota.status_code, 200)
            self.assertEqual(rota.json()["rota"]["km_rodado"], 120)
            self.assertEqual(rota.json()["rota"]["media_km_l"], 6)
            self.assertAlmostEqual(rota.json()["rota"]["custo_km"], 0.58, places=2)

            nf = client.put(
                f"/api/v1/despesas/{codigo}/nf",
                headers=headers,
                json={
                    "nf_numero": "NF-DESP-1",
                    "nf_kg": 100,
                    "nf_preco": 10,
                    "nf_caixas": 5,
                    "nf_kg_carregado": 100,
                    "nf_kg_vendido": 80,
                    "nf_media_carregada": 20,
                    "nf_caixa_final": 4,
                    "mortalidade_transbordo_aves": 2,
                    "mortalidade_transbordo_kg": 1.5,
                    "obs_transbordo": "SEM DIVERGENCIA",
                },
            )
            self.assertEqual(nf.status_code, 200)
            self.assertEqual(nf.json()["nf"]["nf_numero"], "NF-DESP-1")
            self.assertEqual(nf.json()["nf"]["nf_saldo"], 20)
            self.assertEqual(nf.json()["nf"]["total_compra"], 1000)
            self.assertEqual(nf.json()["nf"]["receita_estimada"], 1600)

            financeiro = client.put(
                f"/api/v1/despesas/{codigo}/financeiro",
                headers=headers,
                json={"adiantamento": 20, "pix_motorista": 10, "cedulas": {"50": 1, "20": 1}},
            )
            self.assertEqual(financeiro.status_code, 200)
            body = financeiro.json()
            self.assertEqual(body["financeiro"]["total_recebido"], 200)
            self.assertEqual(body["financeiro"]["total_despesas"], 70)
            self.assertEqual(body["financeiro"]["valor_dinheiro"], 70)
            self.assertEqual(body["financeiro"]["total_devolvido"], 80)
            self.assertEqual(body["financeiro"]["diferenca"], 70)
            self.assertEqual(body["financeiro"]["resultado_liquido"], 70)

            pdf = client.get(f"/api/v1/despesas/{codigo}/pdf", headers=headers)
            self.assertEqual(pdf.status_code, 200)
            self.assertEqual(pdf.headers["content-type"], "application/pdf")
            self.assertIn("PRESTACAO_", pdf.headers["content-disposition"])
            self.assertTrue(pdf.content.startswith(b"%PDF"))
            self.assertIn(b"PRESTACAO DE CONTAS", pdf.content)

            finalizada = client.post(f"/api/v1/despesas/{codigo}/finalizar", headers=headers)
            self.assertEqual(finalizada.status_code, 200)
            self.assertEqual(finalizada.json()["cabecalho"]["prestacao_status"], "FECHADA")
            self.assertTrue(finalizada.json()["cabecalho"]["fechada"])

            bundle_fechado = client.get(f"/api/v1/despesas/{codigo}/bundle", headers=headers)
            self.assertEqual(bundle_fechado.status_code, 200)
            self.assertEqual(bundle_fechado.json()["cabecalho"]["prestacao_status"], "FECHADA")
            self.assertTrue(bundle_fechado.json()["cabecalho"]["fechada"])

            pdf_fechado = client.get(f"/api/v1/despesas/{codigo}/pdf", headers=headers)
            self.assertEqual(pdf_fechado.status_code, 200)
            self.assertTrue(pdf_fechado.content.startswith(b"%PDF"))

            blocked = client.post(
                f"/api/v1/despesas/{codigo}/despesas",
                headers=headers,
                json={"descricao": "BLOQUEADA", "valor": 1, "categoria": "OUTRA"},
            )
            self.assertEqual(blocked.status_code, 409)

    def test_admin_can_read_centro_custos_summary(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA CENTRO CUSTOS",
                    "codigo": "MCC907",
                    "senha": "1234",
                    "telefone": "88999999061",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            for placa in ("PGC901", "PGC902"):
                veiculo = client.post(
                    "/api/v1/cadastros/veiculos",
                    headers=headers,
                    json={"placa": placa, "modelo": "TRUCK", "capacidade_cx": 25},
                )
                self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "CC UM", "telefone": "88999999062", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "CC DOIS", "telefone": "88999999063", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)
            ajudantes_ids = [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])]

            def criar_rota(placa: str, despesa_valor: float, km_final: float, nf_kg_carregado: float):
                programacao = client.post(
                    "/api/v1/programacao/",
                    headers=headers,
                    json={
                        "motorista": "MOTORISTA CENTRO CUSTOS",
                        "motorista_codigo": "MCC907",
                        "veiculo": placa,
                        "ajudantes": ajudantes_ids,
                        "local_rota": "SERRA",
                        "tipo_estimativa": "KG",
                        "kg_estimado": nf_kg_carregado,
                        "local_carregamento": "GRANJA CENTRO",
                        "itens": [],
                    },
                )
                self.assertEqual(programacao.status_code, 201)
                codigo = programacao.json()["codigo_programacao"]
                despesa = client.post(
                    f"/api/v1/despesas/{codigo}/despesas",
                    headers=headers,
                    json={"descricao": "CUSTO ROTA", "valor": despesa_valor, "categoria": "ROTA"},
                )
                self.assertEqual(despesa.status_code, 201)
                rota = client.put(
                    f"/api/v1/despesas/{codigo}/rota",
                    headers=headers,
                    json={"km_inicial": 0, "km_final": km_final, "litros": 10},
                )
                self.assertEqual(rota.status_code, 200)
                nf = client.put(
                    f"/api/v1/despesas/{codigo}/nf",
                    headers=headers,
                    json={
                        "nf_numero": f"NF-{codigo}",
                        "nf_kg": nf_kg_carregado,
                        "nf_preco": 10,
                        "nf_caixas": 1,
                        "nf_kg_carregado": nf_kg_carregado,
                        "nf_kg_vendido": 0,
                    },
                )
                self.assertEqual(nf.status_code, 200)
                async def liberar_recursos_rota():
                    async with async_session() as session:
                        item = await session.get(ProgramacaoDB, programacao.json()["id"])
                        item.status_operacional = "FINALIZADA"
                        item.status = "FINALIZADA"
                        await session.commit()

                asyncio.run(liberar_recursos_rota())
                return codigo

            criar_rota("PGC901", 100, 100, 200)
            criar_rota("PGC901", 50, 50, 100)
            criar_rota("PGC902", 30, 100, 50)

            options = client.get("/api/v1/centro-custos/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertIn("PGC901", options.json()["veiculos"])
            self.assertIn("PGC902", options.json()["veiculos"])

            resumo = client.get(
                "/api/v1/centro-custos/resumo?periodo=TODAS&veiculo=PGC901&metric=DESPESA_TOTAL",
                headers=headers,
            )
            self.assertEqual(resumo.status_code, 200)
            body = resumo.json()
            self.assertEqual(body["periodo"], "TODAS")
            self.assertEqual(body["veiculo"], "PGC901")
            self.assertEqual(body["kpis"]["veiculos"], 1)
            self.assertEqual(body["kpis"]["rotas"], 2)
            self.assertEqual(body["kpis"]["km_total"], 150)
            self.assertEqual(body["kpis"]["kg_carregado"], 300)
            self.assertEqual(body["kpis"]["despesas_total"], 150)
            self.assertEqual(body["kpis"]["custo_km_global"], 1)
            self.assertEqual(body["kpis"]["custo_kg_global"], 0.5)
            self.assertEqual(len(body["rows"]), 1)
            row = body["rows"][0]
            self.assertEqual(row["veiculo"], "PGC901")
            self.assertEqual(row["ticket_rota"], 75)
            self.assertEqual(body["chart"][0]["label"], "PGC901")
            self.assertEqual(body["chart"][0]["value"], 150)

    def test_mortalidade_uses_real_sources_for_kg_and_value(self):
        codigo = "PGMORTREAL01"
        photo_fd, photo_name = tempfile.mkstemp(suffix=".jpg")
        os.close(photo_fd)
        photo_path = Path(photo_name)
        photo_path.write_bytes(b"fake-image-bytes")

        async def seed_mortalidade():
            async with async_session() as session:
                cols = await session.execute(text("PRAGMA table_info(programacao_itens_controle)"))
                if "aves_por_caixa" not in {row[1] for row in cols.fetchall()}:
                    await session.execute(text("ALTER TABLE programacao_itens_controle ADD COLUMN aves_por_caixa INTEGER DEFAULT 0"))
                session.add(
                    ProgramacaoDB(
                        codigo_programacao=codigo,
                        codigo=codigo,
                        data="2026-05-26",
                        data_criacao="2026-05-26",
                        motorista="MOTORISTA MORT",
                        veiculo="MRT001",
                        nf_numero="NF-MORT-1",
                        num_nf="NF-MORT-1",
                        nf_preco=5,
                        preco_nf=5,
                        media=4,
                        mortalidade_transbordo_aves=2,
                        mortalidade_transbordo_kg=7,
                    )
                )
                session.add(
                    ProgramacaoItemDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI1",
                        nome_cliente="CLIENTE MORT",
                        pedido="PED1",
                        qnt_caixas=10,
                        kg=120,
                        preco=9,
                    )
                )
                await session.flush()
                await session.execute(
                    text(
                        """
                        INSERT INTO programacao_itens_controle
                            (codigo_programacao, cod_cliente, pedido, status_pedido, caixas_atual,
                             preco_atual, mortalidade_aves, media_aplicada, peso_previsto,
                             aves_por_caixa, timestamp_entrega, updated_at)
                        VALUES
                            (:codigo, 'CLI1', 'PED1', 'ENTREGUE', 10,
                             11, 3, 0, 120,
                             4, '2026-05-26 10:00:00', '2026-05-26 10:00:00')
                        """
                    ),
                    {"codigo": codigo},
                )
                await session.execute(
                    text(
                        """
                        INSERT INTO rota_fotos (
                            id_foto, codigo_programacao, categoria, tipo_registro,
                            path_local, storage_path, arquivo_nome, mime_type, registrado_em
                        ) VALUES (
                            'FOTO-MORT-REAL-01', :codigo, 'DOA', 'MORTALIDADE',
                            :path, :path, :nome, 'image/jpeg', '2026-05-26 10:05:00'
                        )
                        """
                    ),
                    {"codigo": codigo, "path": str(photo_path), "nome": photo_path.name},
                )
                await session.commit()

        try:
            asyncio.run(seed_mortalidade())

            with TestClient(app, base_url="http://testserver") as client:
                headers = self._auth_headers(client)
                response = client.get(
                    f"/api/v1/despesas/mortalidade/fotos?periodo=TODAS&codigo_programacao={codigo}",
                    headers=headers,
                )
                photo_response = client.get(
                    "/api/v1/despesas/mortalidade/fotos/FOTO-MORT-REAL-01/arquivo",
                    headers=headers,
                )
        finally:
            try:
                photo_path.unlink()
            except OSError:
                pass

        self.assertEqual(response.status_code, 200)
        self.assertEqual(photo_response.status_code, 200)
        self.assertEqual(photo_response.content, b"fake-image-bytes")
        body = response.json()
        self.assertEqual(body["kpis"]["mortalidade_cliente_aves"], 3)
        self.assertEqual(body["kpis"]["mortalidade_doa_aves"], 2)
        self.assertEqual(body["kpis"]["mortalidade_cliente_kg"], 9.0)
        self.assertEqual(body["kpis"]["mortalidade_doa_kg"], 7.0)
        self.assertEqual(body["kpis"]["valor_cliente"], 99.0)
        self.assertEqual(body["kpis"]["valor_operacao"], 35.0)
        self.assertEqual(body["kpis"]["valor_afetado"], 134.0)
        cliente = next(item for item in body["fotos"] if item["escopo"] == "CLIENTE")
        self.assertEqual(cliente["fonte_kg"], "PESO_PREVISTO_CAIXAS")
        self.assertEqual(cliente["fonte_preco"], "CONTROLE_PRECO_ATUAL")

    def test_centro_custos_uses_predicted_sale_when_receipts_are_partial(self):
        codigo = "PGFINPARCIAL01"

        async def seed_partial_financial_route():
            async with async_session() as session:
                session.add(
                    ProgramacaoDB(
                        codigo_programacao=codigo,
                        codigo=codigo,
                        data="2026-05-26",
                        data_criacao="2026-05-26",
                        motorista="MOTORISTA PARCIAL",
                        veiculo="PFP001",
                        local_rota="SERTAO",
                        nf_kg=1000,
                        nf_preco=5,
                        nf_kg_carregado=1000,
                        km_rodado=100,
                    )
                )
                session.add_all(
                    [
                        ProgramacaoItemDB(
                            codigo_programacao=codigo,
                            cod_cliente="CLI1",
                            nome_cliente="CLIENTE UM",
                            pedido="P1",
                            qnt_caixas=10,
                            kg=500,
                            preco=10,
                        ),
                        ProgramacaoItemDB(
                            codigo_programacao=codigo,
                            cod_cliente="CLI2",
                            nome_cliente="CLIENTE DOIS",
                            pedido="P2",
                            qnt_caixas=10,
                            kg=500,
                            preco=10,
                        ),
                        ProgramacaoItemControleDB(
                            codigo_programacao=codigo,
                            cod_cliente="CLI1",
                            pedido="P1",
                            status_pedido="ENTREGUE",
                            caixas_atual=10,
                            peso_previsto=500,
                            preco_atual=10,
                            valor_recebido=100,
                        ),
                        ProgramacaoItemControleDB(
                            codigo_programacao=codigo,
                            cod_cliente="CLI2",
                            pedido="P2",
                            status_pedido="ENTREGUE",
                            caixas_atual=10,
                            peso_previsto=500,
                            preco_atual=10,
                            valor_recebido=0,
                        ),
                    ]
                )
                await session.commit()

        asyncio.run(seed_partial_financial_route())

        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)
            response = client.get(
                f"/api/v1/centro-custos/financeiro?periodo=TODAS&veiculo=PFP001&limit=5000",
                headers=headers,
            )

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.json()["rows"] if item["codigo_programacao"] == codigo)
        self.assertEqual(row["venda"], 10000)
        self.assertEqual(row["venda_confirmada"], 100)
        self.assertEqual(row["venda_prevista"], 10000)
        self.assertEqual(row["fonte_venda"], "ITENS_COM_RECEBIMENTO_PARCIAL")
        self.assertTrue(any("Recebimento parcial" in alerta for alerta in row["alertas"]))

    def test_admin_can_generate_relatorios_and_route_status(self):
        codigo = "REL990"

        async def seed_relatorio_data():
            async with async_session() as session:
                session.add(
                    ProgramacaoDB(
                        codigo_programacao=codigo,
                        data_criacao="2026-05-08",
                        motorista="MOTORISTA RELATORIO",
                        veiculo="PGR990",
                        equipe="AJUDANTE REL UM|AJUDANTE REL DOIS",
                        kg_estimado=120,
                        status="ATIVA",
                        prestacao_status="PENDENTE",
                        local_rota="SERRA",
                        local_carregamento="GRANJA REL",
                        nf_numero="NF-REL-990",
                        data_saida="2026-05-08",
                        hora_saida="07:00:00",
                        data_chegada="2026-05-08",
                        hora_chegada="16:00:00",
                        adiantamento=20,
                        nf_kg=120,
                        nf_caixas=10,
                        nf_kg_carregado=120,
                        nf_kg_vendido=100,
                        nf_saldo=20,
                        km_inicial=10,
                        km_final=90,
                        km_rodado=80,
                        litros=20,
                        media_km_l=4,
                        custo_km=0.5,
                        ced_50_qtd=2,
                        valor_dinheiro=100,
                    )
                )
                session.add(
                    ProgramacaoItemDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-REL-1",
                        nome_cliente="CLIENTE RELATORIO",
                        qnt_caixas=2,
                        kg=100,
                        preco=750,
                        vendedor="VEND REL",
                    )
                )
                session.add(
                    RecebimentoDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-REL-1",
                        nome_cliente="CLIENTE RELATORIO",
                        valor=150,
                        forma_pagamento="DINHEIRO",
                        data_registro="2026-05-08 16:10:00",
                    )
                )
                session.add(
                    DespesaDB(
                        codigo_programacao=codigo,
                        descricao="DESPESA REL",
                        valor=40,
                        data_registro="2026-05-08 15:00:00",
                        categoria="RELCAT",
                        motorista="MOTORISTA RELATORIO",
                        veiculo="PGR990",
                    )
                )
                session.add(
                    ProgramacaoItemControleDB(
                        codigo_programacao=codigo,
                        cod_cliente="CLI-REL-1",
                        mortalidade_aves=3,
                    )
                )
                session.add(
                    ProgramacaoDB(
                        codigo_programacao="REL-CX-991",
                        data_criacao="2026-05-08",
                        motorista="MOTORISTA TRANSBORDO REL",
                        veiculo="PGR991",
                        equipe="AJUDANTE TRANSBORDO UM|AJUDANTE TRANSBORDO DOIS",
                        kg_estimado=0,
                        tipo_estimativa="CX",
                        caixas_estimado=30,
                        operacao_tipo="TRANSBORDO",
                        transbordo_modalidade="EMPRESA_BUSCA",
                        transbordo_grupo="REL-CX-991",
                        status="FINALIZADA",
                        status_operacional="FINALIZADA",
                        finalizada_no_app=1,
                        prestacao_status="PENDENTE",
                        local_rota="SERRA",
                        local_carregamento="GRANJA TRANSBORDO REL",
                        nf_numero="NF-REL-CX-991",
                    )
                )
                await session.commit()

        asyncio.run(seed_relatorio_data())

        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            options = client.get("/api/v1/relatorios/options", headers=headers)
            self.assertEqual(options.status_code, 200)
            self.assertIn("Prestacao de Contas", options.json()["tipos"])

            programacoes = client.get(
                "/api/v1/relatorios/programacoes?tipo=Programacoes&codigo=REL990",
                headers=headers,
            )
            self.assertEqual(programacoes.status_code, 200)
            self.assertEqual(programacoes.json()[0]["codigo_programacao"], codigo)

            transbordo_programacoes = client.get(
                "/api/v1/relatorios/programacoes?tipo=Nota%20Fiscal%20%2F%20Transbordo&codigo=REL-CX-991",
                headers=headers,
            )
            self.assertEqual(transbordo_programacoes.status_code, 200)
            transbordo_option = transbordo_programacoes.json()[0]
            self.assertEqual(transbordo_option["codigo_programacao"], "REL-CX-991")
            self.assertEqual(transbordo_option["tipo_estimativa"], "CX")
            self.assertEqual(transbordo_option["operacao_tipo"], "TRANSBORDO")
            self.assertEqual(transbordo_option["transbordo_modalidade"], "EMPRESA_BUSCA")
            self.assertEqual(transbordo_option["transbordo_grupo"], "REL-CX-991")

            resumo_prog = client.get(
                f"/api/v1/relatorios/resumo?tipo=Programacoes&programacao={codigo}",
                headers=headers,
            )
            self.assertEqual(resumo_prog.status_code, 200)
            body_prog = resumo_prog.json()
            self.assertEqual(body_prog["programacao"], codigo)
            self.assertIn("RELATORIO DE PROGRAMACAO", body_prog["text"])
            self.assertEqual(body_prog["rows"][0]["preco"], 7.5)

            prestacao = client.get(
                f"/api/v1/relatorios/resumo?tipo=Prestacao%20de%20Contas&programacao={codigo}",
                headers=headers,
            )
            self.assertEqual(prestacao.status_code, 200)
            body_prest = prestacao.json()
            self.assertIn("RELATORIO DE PRESTACAO DE CONTAS", body_prest["text"])
            self.assertEqual(body_prest["kpis"][2]["value"], "R$ 30,00")
            self.assertTrue(any(row["bloco"] == "RECEBIMENTO" for row in body_prest["rows"]))
            self.assertTrue(any(row["bloco"] == "DESPESA" for row in body_prest["rows"]))

            rotina = client.get(
                "/api/v1/relatorios/resumo?tipo=Rotina%20Motorista%2FAjudantes&motorista=MOTORISTA%20RELATORIO",
                headers=headers,
            )
            self.assertEqual(rotina.status_code, 200)
            self.assertTrue(any(row["nome"] == "MOTORISTA RELATORIO" for row in rotina.json()["rows"]))

            km = client.get("/api/v1/relatorios/resumo?tipo=KM%20de%20Veiculos", headers=headers)
            self.assertEqual(km.status_code, 200)
            self.assertTrue(any(row["veiculo"] == "PGR990" for row in km.json()["rows"]))

            despesas = client.get("/api/v1/relatorios/resumo?tipo=Despesas", headers=headers)
            self.assertEqual(despesas.status_code, 200)
            self.assertTrue(any(row["categoria"] == "RELCAT" and row["total"] == 40 for row in despesas.json()["rows"]))

            mortalidade = client.get(
                "/api/v1/relatorios/resumo?tipo=Mortalidade%20Motorista&codigo=REL990",
                headers=headers,
            )
            self.assertEqual(mortalidade.status_code, 200)
            self.assertEqual(mortalidade.json()["rows"][0]["mortalidade_total"], 3)

            excel = client.get(f"/api/v1/relatorios/{codigo}/exportar-excel", headers=headers)
            self.assertEqual(excel.status_code, 200)
            self.assertIn("spreadsheetml", excel.headers["content-type"])

            pdf = client.get(f"/api/v1/relatorios/pdf?tipo=Programacoes&programacao={codigo}", headers=headers)
            self.assertEqual(pdf.status_code, 200)
            self.assertEqual(pdf.headers["content-type"], "application/pdf")

            finalizada = client.post(f"/api/v1/relatorios/{codigo}/finalizar-rota", headers=headers)
            self.assertEqual(finalizada.status_code, 200)
            self.assertEqual(finalizada.json()["status"], "FINALIZADA")

            reaberta = client.post(f"/api/v1/relatorios/{codigo}/reabrir-rota", headers=headers)
            self.assertEqual(reaberta.status_code, 200)
            self.assertEqual(reaberta.json()["status"], "ATIVA")

    def test_nf_transbordo_report_uses_transfer_weight_for_destination(self):
        async def seed_nf_transbordo_data():
            async with async_session() as session:
                await session.execute(
                    text(
                        """
                        CREATE TABLE IF NOT EXISTS transferencias (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            codigo_origem TEXT,
                            codigo_destino TEXT,
                            cod_cliente TEXT,
                            pedido TEXT,
                            qtd_caixas INTEGER,
                            qtd_convertida INTEGER,
                            status TEXT,
                            snapshot TEXT,
                            obs TEXT,
                            motorista_origem TEXT,
                            motorista_destino TEXT,
                            criado_em TEXT,
                            atualizado_em TEXT
                        )
                        """
                    )
                )
                session.add(
                    ProgramacaoDB(
                        codigo_programacao="NFTR-ROOT",
                        data_criacao="2026-05-09",
                        motorista="MOTORISTA RAIZ",
                        veiculo="PGR991",
                        equipe="AJUDANTE RAIZ",
                        status="ATIVA",
                        prestacao_status="PENDENTE",
                        operacao_tipo="TRANSBORDO",
                        transbordo_modalidade="DISTRIBUICAO",
                        nf_numero="NF-TR-991",
                        nf_kg=1000,
                        nf_preco=10,
                        nf_kg_carregado=1000,
                        kg_carregado=1000,
                        nf_caixas=100,
                        caixas_carregadas=100,
                        total_caixas=100,
                    )
                )
                session.add(
                    ProgramacaoDB(
                        codigo_programacao="NFTR-DEST",
                        data_criacao="2026-05-09",
                        motorista="MOTORISTA DESTINO",
                        veiculo="PGR992",
                        equipe="AJUDANTE DESTINO",
                        status="ATIVA",
                        prestacao_status="PENDENTE",
                        operacao_tipo="VENDA",
                    )
                )
                session.add(
                    ProgramacaoDB(
                        codigo_programacao="NFTR-PARC",
                        data_criacao="2026-05-09",
                        motorista="MOTORISTA PARCIAL",
                        veiculo="PGR993",
                        equipe="AJUDANTE PARCIAL",
                        status="ATIVA",
                        prestacao_status="PENDENTE",
                        operacao_tipo="VENDA",
                    )
                )
                session.add(
                    ProgramacaoItemDB(
                        codigo_programacao="NFTR-DEST",
                        cod_cliente="CLI-NFTR",
                        nome_cliente="CLIENTE NF TRANSBORDO",
                        qnt_caixas=20,
                        kg=0,
                        preco=10,
                        pedido="PED-NFTR",
                        carga_raiz_programacao="NFTR-ROOT",
                        carga_origem_imediata="NFTR-ROOT",
                        transferencia_origem_id="1",
                    )
                )
                session.add(
                    ProgramacaoItemControleDB(
                        codigo_programacao="NFTR-DEST",
                        cod_cliente="CLI-NFTR",
                        pedido="PED-NFTR",
                        status_pedido="ENTREGUE",
                        caixas_atual=0,
                    )
                )
                session.add(
                    RecebimentoDB(
                        codigo_programacao="NFTR-DEST",
                        cod_cliente="CLI-NFTR",
                        nome_cliente="CLIENTE NF TRANSBORDO",
                        valor=200,
                        forma_pagamento="PIX",
                        data_registro="2026-05-09 12:00:00",
                    )
                )
                await session.execute(
                    text(
                        """
                        INSERT INTO transferencias (
                            codigo_origem, codigo_destino, cod_cliente, pedido, qtd_caixas,
                            qtd_convertida, status, snapshot, obs, motorista_origem,
                            motorista_destino, criado_em, atualizado_em
                        ) VALUES (
                            'NFTR-ROOT', 'NFTR-DEST', 'CLI-NFTR', 'PED-NFTR', 20,
                            20, 'CONVERTIDA', '{"carga_raiz_programacao":"NFTR-ROOT"}',
                            'TESTE NF TRANSBORDO', 'MOTORISTA RAIZ', 'MOTORISTA DESTINO',
                            '2026-05-09 10:00:00', '2026-05-09 10:30:00'
                        )
                        """
                    )
                )
                await session.execute(
                    text(
                        """
                        INSERT INTO transferencias (
                            codigo_origem, codigo_destino, cod_cliente, pedido, qtd_caixas,
                            qtd_convertida, status, snapshot, obs, motorista_origem,
                            motorista_destino, criado_em, atualizado_em
                        ) VALUES (
                            'NFTR-ROOT', 'NFTR-PARC', 'CLI-NFTR-P', 'PED-NFTR-P', 20,
                            5, 'ACEITA', '{"carga_raiz_programacao":"NFTR-ROOT"}',
                            'TESTE NF TRANSBORDO PARCIAL', 'MOTORISTA RAIZ', 'MOTORISTA PARCIAL',
                            '2026-05-09 11:00:00', '2026-05-09 11:30:00'
                        )
                        """
                    )
                )
                await session.commit()

        asyncio.run(seed_nf_transbordo_data())

        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            response = client.get(
                "/api/v1/relatorios/resumo?tipo=Nota%20Fiscal%20%2F%20Transbordo&nf=NF-TR-991",
                headers=headers,
            )
            self.assertEqual(response.status_code, 200)
            body = response.json()
            rows_by_codigo = {row["codigo_programacao"]: row for row in body["rows"]}
            self.assertEqual(rows_by_codigo["NFTR-DEST"]["kg_entregue"], 200)
            self.assertEqual(rows_by_codigo["NFTR-PARC"]["kg_entregue"], 50)
            self.assertEqual(rows_by_codigo["NFTR-DEST"]["kg_transferido_entrada"], 200)
            self.assertEqual(rows_by_codigo["NFTR-PARC"]["kg_transferido_entrada"], 50)
            self.assertEqual(rows_by_codigo["NFTR-ROOT"]["kg_transferido_saida"], 250)
            self.assertEqual(body["kpis"][2]["value"], "250.00")
            transfer_section = next(section for section in body["sections"] if section["title"] == "Fluxo de Transbordo e Transferencias")
            transfer_by_destino = {row["destino"]: row for row in transfer_section["rows"]}
            self.assertEqual(transfer_by_destino["NFTR-DEST"]["kg_estimado"], 200)
            self.assertEqual(transfer_by_destino["NFTR-PARC"]["caixas"], 20)
            self.assertEqual(transfer_by_destino["NFTR-PARC"]["caixas_convertidas"], 5)
            self.assertEqual(transfer_by_destino["NFTR-PARC"]["caixas_saldo"], 15)
            self.assertEqual(transfer_by_destino["NFTR-PARC"]["kg_estimado"], 200)
            self.assertEqual(transfer_by_destino["NFTR-PARC"]["kg_convertido_estimado"], 50)

            prestacao = client.get(
                "/api/v1/relatorios/resumo?tipo=Prestacao%20de%20Contas&programacao=NFTR-PARC",
                headers=headers,
            )
            self.assertEqual(prestacao.status_code, 200)
            transbordo_rows = [row for row in prestacao.json()["rows"] if row["bloco"] == "TRANSBORDO"]
            self.assertEqual(len(transbordo_rows), 1)
            self.assertEqual(transbordo_rows[0]["caixas"], 20)
            self.assertEqual(transbordo_rows[0]["caixas_convertidas"], 5)
            self.assertEqual(transbordo_rows[0]["caixas_saldo"], 15)
            self.assertEqual(transbordo_rows[0]["kg_estimado"], 200)
            self.assertEqual(transbordo_rows[0]["kg_convertido_estimado"], 50)
            self.assertIn("conv 5 / saldo 15", transbordo_rows[0]["descricao"])

            async def load_pdf_transferencias():
                async with async_session() as session:
                    return await transferencias_operacionais_pdf(session, "NFTR-PARC")

            pdf_transferencias = asyncio.run(load_pdf_transferencias())
            self.assertEqual(len(pdf_transferencias), 1)
            self.assertEqual(pdf_transferencias[0]["caixas"], 20)
            self.assertEqual(pdf_transferencias[0]["caixas_convertidas"], 5)
            self.assertEqual(pdf_transferencias[0]["caixas_saldo"], 15)
            self.assertEqual(pdf_transferencias[0]["kg"], 200)
            self.assertEqual(pdf_transferencias[0]["kg_convertido"], 50)

            async def load_centro_custos_transferencias():
                async with async_session() as session:
                    return await transferencias_compra_por_programacao(session, ["NFTR-ROOT", "NFTR-DEST", "NFTR-PARC"])

            centro_saida, centro_entrada = asyncio.run(load_centro_custos_transferencias())
            self.assertEqual(centro_saida["NFTR-ROOT"], 2500)
            self.assertEqual(centro_entrada["NFTR-DEST"], 2000)
            self.assertEqual(centro_entrada["NFTR-PARC"], 500)

            centro_financeiro = client.get(
                "/api/v1/centro-custos/financeiro?periodo=TODAS&veiculo=TODOS&limit=5000",
                headers=headers,
            )
            self.assertEqual(centro_financeiro.status_code, 200)
            centro_body = centro_financeiro.json()
            centro_rows = centro_body["rows"]
            root_row = next(row for row in centro_rows if row["codigo_programacao"] == "NFTR-ROOT")
            self.assertTrue(root_row["has_children"])
            self.assertEqual([row["codigo_programacao"] for row in root_row["filhos"]], ["NFTR-DEST", "NFTR-PARC"])
            self.assertNotIn("NFTR-DEST", [row["codigo_programacao"] for row in centro_rows if row["codigo_programacao"] != "NFTR-ROOT"])
            self.assertEqual(root_row["compra"], 2500)
            self.assertEqual(root_row["venda"], 200)
            self.assertEqual(root_row["lucro_bruto"], -2300)
            self.assertEqual(root_row["fonte_venda"], "TRANSBORDO_CONSOLIDADO")
            self.assertFalse(any("Sem venda" in alerta for alerta in root_row["alertas"]))
            self.assertEqual(root_row["filhos"][0]["parent_codigo"], "NFTR-ROOT")
            self.assertEqual(root_row["filhos"][0]["nivel"], 1)

            centro_financeiro_veiculo = client.get(
                "/api/v1/centro-custos/financeiro?periodo=TODAS&veiculo=PGR991&limit=5000",
                headers=headers,
            )
            self.assertEqual(centro_financeiro_veiculo.status_code, 200)
            root_filtrado = centro_financeiro_veiculo.json()["rows"][0]
            self.assertEqual(root_filtrado["codigo_programacao"], "NFTR-ROOT")
            self.assertEqual([row["codigo_programacao"] for row in root_filtrado["filhos"]], ["NFTR-DEST", "NFTR-PARC"])

            rotina_destino = client.get(
                "/api/v1/relatorios/resumo?tipo=Rotina%20Motorista%2FAjudantes&motorista=MOTORISTA%20DESTINO",
                headers=headers,
            )
            self.assertEqual(rotina_destino.status_code, 200)
            motorista_destino = next(row for row in rotina_destino.json()["rows"] if row["tipo"] == "MOTORISTA")
            self.assertEqual(motorista_destino["kg"], 200)

            rotina_parcial = client.get(
                "/api/v1/relatorios/resumo?tipo=Rotina%20Motorista%2FAjudantes&motorista=MOTORISTA%20PARCIAL",
                headers=headers,
            )
            self.assertEqual(rotina_parcial.status_code, 200)
            motorista_parcial = next(row for row in rotina_parcial.json()["rows"] if row["tipo"] == "MOTORISTA")
            self.assertEqual(motorista_parcial["kg"], 50)

            home_preview = client.get("/api/v1/home/rotas/NFTR-PARC/preview", headers=headers)
            self.assertEqual(home_preview.status_code, 200)
            self.assertEqual(home_preview.json()["programacao"]["nf_kg_vendido"], 50)
            self.assertEqual(home_preview.json()["programacao"]["kg_transferido_convertido"], 50)
            self.assertEqual(home_preview.json()["resumo"]["kg"], 50)

            home_overview = client.get("/api/v1/home/overview?limit=100", headers=headers)
            self.assertEqual(home_overview.status_code, 200)
            home_rows = {row["codigo_programacao"]: row for row in home_overview.json()["rotas"]}
            self.assertEqual(home_rows["NFTR-PARC"]["nf_kg_vendido"], 50)
            self.assertEqual(home_rows["NFTR-PARC"]["kg_transferido_convertido"], 50)

            programacao_parcial = client.get("/api/v1/programacao/NFTR-PARC", headers=headers)
            self.assertEqual(programacao_parcial.status_code, 200)
            self.assertEqual(programacao_parcial.json()["total_caixas"], 20)
            self.assertEqual(programacao_parcial.json()["quilos"], 50)

            programacoes = client.get("/api/v1/programacao/", headers=headers)
            self.assertEqual(programacoes.status_code, 200)
            prog_rows = {row["codigo_programacao"]: row for row in programacoes.json()}
            self.assertEqual(prog_rows["NFTR-PARC"]["total_caixas"], 20)
            self.assertEqual(prog_rows["NFTR-PARC"]["quilos"], 50)

    def test_admin_can_use_backup_and_system_tools(self):
        async def seed_system_tools_data():
            async with async_session() as session:
                session.add(
                    VendaImportadaDB(
                        pedido="BKP-1",
                        data_venda="2026-05-08",
                        cliente="CLI-BKP",
                        nome_cliente="CLIENTE BACKUP",
                        vendedor="VENDEDOR BACKUP",
                        produto="FRANGO",
                        vr_total=100,
                        qnt=10,
                        cidade="FORTALEZA",
                        valor_unitario=10,
                    )
                )
                session.add(
                    SistemaLogDB(
                        tipo_acao="TESTE_ANTIGO",
                        descricao="LOG ANTIGO",
                        usuario="TESTE",
                        status="OK",
                        resultado_texto="",
                        executado_em="2000-01-01 00:00:00",
                    )
                )
                await session.commit()

        asyncio.run(seed_system_tools_data())
        backup_path = None

        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            overview = client.get("/api/v1/system-tools/overview", headers=headers)
            self.assertEqual(overview.status_code, 200)
            self.assertIn("programacoes", overview.json()["info"]["registros_por_tabela"])

            integrity = client.get("/api/v1/system-tools/integridade", headers=headers)
            self.assertEqual(integrity.status_code, 200)
            self.assertTrue(integrity.json()["ok"])

            created = client.post("/api/v1/system-tools/backups", headers=headers)
            self.assertEqual(created.status_code, 200)
            created_body = created.json()
            self.assertTrue(created_body["arquivo"].startswith("banco_de_dados_"))
            backup_path = Path(created_body["caminho"])
            self.assertTrue(backup_path.exists())

            backups = client.get("/api/v1/system-tools/backups", headers=headers)
            self.assertEqual(backups.status_code, 200)
            self.assertIn(created_body["arquivo"], [item["arquivo"] for item in backups.json()])

            download = client.get(f"/api/v1/system-tools/backups/{created_body['arquivo']}/download", headers=headers)
            self.assertEqual(download.status_code, 200)
            self.assertTrue(download.content.startswith(b"SQLite format 3"))
            download.close()

            vendas = client.get("/api/v1/system-tools/vendas-importadas/export", headers=headers)
            self.assertEqual(vendas.status_code, 200)
            self.assertIn("spreadsheetml", vendas.headers["content-type"])

            logs = client.get("/api/v1/system-tools/logs", headers=headers)
            self.assertEqual(logs.status_code, 200)
            self.assertTrue(any(item["tipo_acao"] == "BACKUP" for item in logs.json()))

            cleared = client.delete("/api/v1/system-tools/logs?dias=1", headers=headers)
            self.assertEqual(cleared.status_code, 200)
            self.assertGreaterEqual(cleared.json()["linhas_deletadas"], 1)

        if backup_path and backup_path.exists():
            try:
                backup_path.unlink()
            except PermissionError:
                pass

    def test_admin_can_import_mark_and_link_vendas_to_programacao(self):
        with TestClient(app, base_url="http://testserver") as client:
            headers = self._auth_headers(client)

            motorista = client.post(
                "/api/v1/cadastros/motoristas",
                headers=headers,
                json={
                    "nome": "MOTORISTA IMPORTACAO",
                    "codigo": "MI902",
                    "senha": "1234",
                    "telefone": "88999999011",
                    "status": "ATIVO",
                },
            )
            self.assertEqual(motorista.status_code, 201)
            veiculo = client.post(
                "/api/v1/cadastros/veiculos",
                headers=headers,
                json={"placa": "PGT902", "modelo": "TRUCK", "capacidade_cx": 30},
            )
            self.assertEqual(veiculo.status_code, 201)
            ajudante_1 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "IMPORT UM", "telefone": "88999999012", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_1.status_code, 201)
            ajudante_2 = client.post(
                "/api/v1/cadastros/ajudantes",
                headers=headers,
                json={"nome": "AJUDANTE", "sobrenome": "IMPORT DOIS", "telefone": "88999999013", "status": "ATIVO"},
            )
            self.assertEqual(ajudante_2.status_code, 201)

            programacao = client.post(
                "/api/v1/programacao/",
                headers=headers,
                json={
                    "motorista": "MOTORISTA IMPORTACAO",
                    "motorista_codigo": "MI902",
                    "veiculo": "PGT902",
                    "ajudantes": [str(ajudante_1.json()["id"]), str(ajudante_2.json()["id"])],
                    "local_rota": "SERRA",
                    "tipo_estimativa": "CX",
                    "caixas_estimado": 10,
                    "local_carregamento": "GRANJA IMPORT",
                    "itens": [],
                },
            )
            self.assertEqual(programacao.status_code, 201)
            codigo = programacao.json()["codigo_programacao"]

            imported = client.post(
                "/api/v1/importar-vendas/importar",
                headers=headers,
                json={
                    "rows": [
                        {
                            "pedido": "1001",
                            "data_venda": "07/05/2026",
                            "cliente": "CLIMP1",
                            "nome_cliente": "CLIENTE IMPORTADO",
                            "produto": "FRANGO",
                            "vr_total": 120,
                            "qnt": 3,
                            "cidade": "SOBRAL",
                            "vendedor": "VENDEDOR IMPORT",
                        },
                        {
                            "pedido": "1001",
                            "data_venda": "07/05/2026",
                            "cliente": "CLIMP1",
                            "nome_cliente": "CLIENTE IMPORTADO",
                            "produto": "FRANGO",
                            "vr_total": 120,
                            "qnt": 3,
                        },
                        {
                            "pedido": "",
                            "cliente": "SEM PEDIDO",
                            "nome_cliente": "INVALIDO",
                            "produto": "FRANGO",
                        },
                    ]
                },
            )
            self.assertEqual(imported.status_code, 201)
            self.assertEqual(imported.json()["importadas"], 1)
            self.assertEqual(imported.json()["ignoradas"], 1)
            self.assertEqual(imported.json()["invalidas"], 1)

            vendas = client.get("/api/v1/importar-vendas/", headers=headers)
            self.assertEqual(vendas.status_code, 200)
            venda = next(item for item in vendas.json() if item["pedido"] == "1001")
            venda_id = venda["id"]
            self.assertEqual(venda["data_venda"], "2026-05-07")

            toggled = client.post(f"/api/v1/importar-vendas/{venda_id}/toggle-selecao", headers=headers)
            self.assertEqual(toggled.status_code, 200)
            self.assertEqual(toggled.json()["selecionada"], 1)

            vinculo_options = client.get("/api/v1/importar-vendas/programacoes-vinculo", headers=headers)
            self.assertEqual(vinculo_options.status_code, 200)
            self.assertIn(codigo, vinculo_options.json())

            linked = client.post(
                "/api/v1/importar-vendas/vincular",
                headers=headers,
                json={"codigo_programacao": codigo, "ids": [venda_id], "caixas_por_venda": {str(venda_id): 4}},
            )
            self.assertEqual(linked.status_code, 200)
            self.assertEqual(linked.json()["vendas_vinculadas"], 1)
            self.assertEqual(linked.json()["itens_adicionados"], 1)
            self.assertEqual(linked.json()["total_caixas"], 4)

            loaded = client.get(f"/api/v1/programacao/{codigo}", headers=headers)
            self.assertEqual(loaded.status_code, 200)
            self.assertEqual(loaded.json()["total_caixas"], 4)
            self.assertEqual(len(loaded.json()["itens"]), 1)
            item = loaded.json()["itens"][0]
            self.assertEqual(item["cod_cliente"], "CLIMP1")
            self.assertEqual(item["qnt_caixas"], 4)
            self.assertEqual(item["preco"], 40)
            self.assertEqual(item["endereco"], "SOBRAL")

            financeiro = client.get("/api/v1/centro-custos/financeiro?periodo=TODAS&veiculo=TODOS&limit=5000", headers=headers)
            self.assertEqual(financeiro.status_code, 200)
            financeiro_body = financeiro.json()
            row = next(item for item in financeiro_body["rows"] if item["codigo_programacao"] == codigo)
            self.assertEqual(row["venda"], 120)
            self.assertEqual(row["tipo_estimativa"], "CX")
            self.assertEqual(row["operacao_tipo"], "TRANSBORDO")
            self.assertEqual(row["transbordo_modalidade"], "EMPRESA_BUSCA")
            self.assertEqual(row["transbordo_grupo"], codigo)
            self.assertEqual(financeiro_body["kpis"]["venda_total"], sum(item["venda"] for item in financeiro_body["rows"]))

            livres = client.get("/api/v1/importar-vendas/", headers=headers)
            self.assertEqual(livres.status_code, 200)
            self.assertNotIn(venda_id, [item["id"] for item in livres.json()])


if __name__ == "__main__":
    unittest.main()
